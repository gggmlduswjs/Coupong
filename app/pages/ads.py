"""
ê´‘ê³  ë³´ê³ ì„œ ì—…ë¡œë“œ + ë¶„ì„ í˜ì´ì§€
===============================
ê³„ì •ë³„ ê´‘ê³  ì„±ê³¼ / ê´‘ê³ ë¹„ ì •ì‚° Excel ì—…ë¡œë“œ â†’ DB ì €ì¥ â†’ ì „ì²´ ë¶„ì„.
"""

import logging
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import streamlit as st
from sqlalchemy import text

from app.dashboard_utils import (
    query_df,
    fmt_krw,
    fmt_money_df,
    engine,
)

logger = logging.getLogger(__name__)
ROOT = Path(__file__).resolve().parent.parent.parent


def render(selected_account, accounts_df, account_names):
    """ê´‘ê³  í˜ì´ì§€"""

    st.title("ê´‘ê³ ")

    tab_upload, tab_analysis = st.tabs(["ì—…ë¡œë“œ", "ë¶„ì„"])

    with tab_upload:
        _render_upload(accounts_df, account_names)

    with tab_analysis:
        _render_analysis(accounts_df, account_names)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ì—…ë¡œë“œ íƒ­
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _render_upload(accounts_df, account_names):
    """ê³„ì •ë³„ ê´‘ê³  ë³´ê³ ì„œ ì—…ë¡œë“œ"""

    st.caption(
        "WING > ê´‘ê³  ê´€ë¦¬ > ê´‘ê³  ë³´ê³ ì„œì—ì„œ ë‹¤ìš´ë¡œë“œí•œ Excelì„ ê³„ì •ë³„ë¡œ ì—…ë¡œë“œí•˜ì„¸ìš”."
    )

    # â”€â”€ ê³„ì • ì„ íƒ â”€â”€
    upload_account = st.selectbox(
        "ê³„ì • ì„ íƒ",
        account_names,
        key="ad_upload_account",
    )

    if not upload_account:
        st.info("ê³„ì •ì„ ì„ íƒí•˜ì„¸ìš”.")
        return

    acct_row = accounts_df[accounts_df["account_name"] == upload_account].iloc[0]
    acct_id = int(acct_row["id"])
    vendor_id = acct_row.get("vendor_id", "")

    st.caption(f"Vendor ID: **{vendor_id}**")

    # â”€â”€ 2ê°œ ì„¹ì…˜ ë‚˜ë€íˆ â”€â”€
    col_perf, col_spend = st.columns(2)

    # â”€â”€ ì¢Œ: ë§¤ì¶œ ì„±ì¥ ê´‘ê³  ë³´ê³ ì„œ (= ìƒí’ˆê´‘ê³ ) â”€â”€
    with col_perf:
        st.subheader("ê´‘ê³  ì„±ê³¼ ë³´ê³ ì„œ")
        st.caption(
            "ë§¤ì¶œ ì„±ì¥ ê´‘ê³  ë³´ê³ ì„œ ë‹¤ìš´ë¡œë“œ\n\n"
            "ì„¤ì •: ê¸°ê°„ êµ¬ë¶„ **ì¼ë³„**, ë³´ê³ ì„œ êµ¬ì¡° **ìº í˜ì¸>ê´‘ê³ ê·¸ë£¹>ìƒí’ˆ>í‚¤ì›Œë“œ**"
        )

        perf_files = st.file_uploader(
            "ê´‘ê³  ì„±ê³¼ Excel",
            type=["xlsx", "xls"],
            key=f"ad_perf_{upload_account}",
            accept_multiple_files=True,
        )

        if perf_files and st.button(
            f"ì„±ê³¼ ë³´ê³ ì„œ ë™ê¸°í™” ({len(perf_files)}ê°œ)",
            key="btn_perf_sync",
            type="primary",
            use_container_width=True,
        ):
            _sync_performance(perf_files, acct_id)

    # â”€â”€ ìš°: ì¼ë³„ ê´‘ê³ ë¹„ ì •ì‚° â”€â”€
    with col_spend:
        st.subheader("ê´‘ê³ ë¹„ ì •ì‚° ë³´ê³ ì„œ")
        st.caption(
            "ê´‘ê³ ë¹„ ì •ì‚° ë³´ê³ ì„œ > ì¼ë³„ ê´‘ê³ ë¹„ ì •ì‚°ë‚´ì—­ ë‹¤ìš´ë¡œë“œ\n\n"
            "íŒŒì¼ëª…ì— vendor_idê°€ ìë™ í¬í•¨ë©ë‹ˆë‹¤."
        )

        spend_files = st.file_uploader(
            "ê´‘ê³ ë¹„ ì •ì‚° Excel",
            type=["xlsx", "xls"],
            key=f"ad_spend_{upload_account}",
            accept_multiple_files=True,
        )

        if spend_files and st.button(
            f"ì •ì‚° ë³´ê³ ì„œ ë™ê¸°í™” ({len(spend_files)}ê°œ)",
            key="btn_spend_sync",
            type="primary",
            use_container_width=True,
        ):
            _sync_spend(spend_files)

    st.divider()

    # â”€â”€ í˜„ì¬ ê³„ì • ì—…ë¡œë“œ í˜„í™© â”€â”€
    st.subheader(f"{upload_account} ì—…ë¡œë“œ í˜„í™©")
    _render_account_status(acct_id, upload_account)


def _sync_performance(files, account_id):
    """ê´‘ê³  ì„±ê³¼ ë³´ê³ ì„œ ë™ê¸°í™”"""
    from scripts.sync_ad_performance import AdPerformanceSync

    syncer = AdPerformanceSync()
    results = []
    progress = st.progress(0, text="ë™ê¸°í™” ì¤€ë¹„ ì¤‘...")

    for i, f in enumerate(files):
        progress.progress(i / len(files), text=f"ì²˜ë¦¬ ì¤‘: {f.name}")
        tmp_path = ROOT / "data" / "reports" / f.name
        tmp_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path.write_bytes(f.getvalue())

        try:
            result = syncer.sync_file(str(tmp_path), account_id=account_id)
            results.append(result)
        except Exception as e:
            results.append({"file": f.name, "error": str(e), "parsed": 0, "saved": 0})

    progress.progress(1.0, text="ì™„ë£Œ!")
    _show_sync_results("ê´‘ê³  ì„±ê³¼", results)
    query_df.clear()


def _sync_spend(files):
    """ê´‘ê³ ë¹„ ì •ì‚° ë³´ê³ ì„œ ë™ê¸°í™”"""
    from scripts.sync_ad_spend import AdSpendSync

    syncer = AdSpendSync()
    results = []
    progress = st.progress(0, text="ë™ê¸°í™” ì¤€ë¹„ ì¤‘...")

    for i, f in enumerate(files):
        progress.progress(i / len(files), text=f"ì²˜ë¦¬ ì¤‘: {f.name}")
        tmp_path = ROOT / "data" / "reports" / f.name
        tmp_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path.write_bytes(f.getvalue())

        try:
            result = syncer.sync_file(str(tmp_path))
            results.append(result)
        except Exception as e:
            results.append({"file": f.name, "error": str(e), "parsed": 0, "saved": 0})

    progress.progress(1.0, text="ì™„ë£Œ!")
    _show_sync_results("ê´‘ê³ ë¹„ ì •ì‚°", results)
    query_df.clear()


def _show_sync_results(label: str, results: list):
    """ë™ê¸°í™” ê²°ê³¼ í‘œì‹œ"""
    total_parsed = sum(r.get("parsed", 0) for r in results)
    total_saved = sum(r.get("saved", 0) for r in results)
    errors = [r for r in results if r.get("error")]

    if total_saved > 0:
        st.success(f"{label}: {len(results)}ê°œ íŒŒì¼, íŒŒì‹± {total_parsed:,}ê±´, ì €ì¥ {total_saved:,}ê±´")
    elif not errors:
        st.warning(f"{label}: ìƒˆë¡œ ì €ì¥í•  ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤ (ì´ë¯¸ ë™ê¸°í™”ë¨)")

    for r in errors:
        st.error(f"{r['file']}: {r['error']}")

    if results:
        rows = []
        for r in results:
            rows.append({
                "íŒŒì¼": r.get("file", ""),
                "ê³„ì •": r.get("account", "-"),
                "ê¸°ê°„": r.get("period", "-"),
                "íŒŒì‹±": r.get("parsed", 0),
                "ì €ì¥": r.get("saved", 0),
                "ì˜¤ë¥˜": r.get("error", ""),
            })
        st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)


def _render_account_status(acct_id: int, acct_name: str):
    """ë‹¨ì¼ ê³„ì •ì˜ ì—…ë¡œë“œ í˜„í™©"""
    try:
        perf = query_df(f"""
            SELECT MIN(ad_date) as ì‹œì‘ì¼, MAX(ad_date) as ì¢…ë£Œì¼,
                COUNT(*) as ë ˆì½”ë“œìˆ˜,
                SUM(ad_spend) as ê´‘ê³ ë¹„, SUM(total_revenue) as ì „í™˜ë§¤ì¶œ
            FROM ad_performances WHERE account_id = {acct_id}
        """)
        spend = query_df(f"""
            SELECT MIN(ad_date) as ì‹œì‘ì¼, MAX(ad_date) as ì¢…ë£Œì¼,
                COUNT(*) as ë ˆì½”ë“œìˆ˜,
                SUM(total_charge) as ì´ë¹„ìš©
            FROM ad_spends WHERE account_id = {acct_id}
        """)
    except Exception:
        st.info("ì•„ì§ ì—…ë¡œë“œëœ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        return

    c1, c2 = st.columns(2)
    with c1:
        p = perf.iloc[0] if not perf.empty else None
        p_cnt = int(p["ë ˆì½”ë“œìˆ˜"]) if p is not None and pd.notna(p["ë ˆì½”ë“œìˆ˜"]) else 0
        if p_cnt > 0:
            st.metric("ê´‘ê³  ì„±ê³¼", f"{p_cnt:,}ê±´")
            st.caption(f"{p['ì‹œì‘ì¼']} ~ {p['ì¢…ë£Œì¼']}")
        else:
            st.metric("ê´‘ê³  ì„±ê³¼", "0ê±´")
            st.caption("ì—…ë¡œë“œ í•„ìš”")

    with c2:
        s = spend.iloc[0] if not spend.empty else None
        s_cnt = int(s["ë ˆì½”ë“œìˆ˜"]) if s is not None and pd.notna(s["ë ˆì½”ë“œìˆ˜"]) else 0
        if s_cnt > 0:
            st.metric("ê´‘ê³ ë¹„ ì •ì‚°", f"{s_cnt:,}ê±´")
            st.caption(f"{s['ì‹œì‘ì¼']} ~ {s['ì¢…ë£Œì¼']}")
        else:
            st.metric("ê´‘ê³ ë¹„ ì •ì‚°", "0ê±´")
            st.caption("ì—…ë¡œë“œ í•„ìš”")


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ë¶„ì„ íƒ­
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _render_analysis(accounts_df, account_names):
    """ì „ì²´ ê´‘ê³  ë¶„ì„"""

    # ë°ì´í„° ì¡´ì¬ í™•ì¸
    try:
        p_cnt = int(query_df("SELECT COUNT(*) as c FROM ad_performances").iloc[0]["c"])
        s_cnt = int(query_df("SELECT COUNT(*) as c FROM ad_spends").iloc[0]["c"])
    except Exception:
        p_cnt, s_cnt = 0, 0

    if p_cnt == 0 and s_cnt == 0:
        st.info("ê´‘ê³  ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤. 'ì—…ë¡œë“œ' íƒ­ì—ì„œ ë³´ê³ ì„œë¥¼ ì—…ë¡œë“œí•˜ì„¸ìš”.")
        return

    # â”€â”€ ìƒë‹¨ ì»¨íŠ¸ë¡¤ â”€â”€
    ctrl1, ctrl2 = st.columns([3, 3])
    with ctrl1:
        period_opt = st.selectbox("ê¸°ê°„", ["1ì£¼", "2ì£¼", "1ê°œì›”", "3ê°œì›”"], index=2, key="ad_period")
    with ctrl2:
        acct_filter = st.selectbox("ê³„ì •", ["ì „ì²´"] + account_names, key="ad_acct_filter")

    period_map = {"1ì£¼": 7, "2ì£¼": 14, "1ê°œì›”": 30, "3ê°œì›”": 90}
    days_back = period_map[period_opt]
    d_to = date.today()
    d_from = d_to - timedelta(days=days_back)
    d_from_s = d_from.isoformat()
    d_to_s = d_to.isoformat()

    # ê³„ì • í•„í„°
    acct_where_perf = ""
    acct_where_spend = ""
    if acct_filter != "ì „ì²´":
        _aid_row = accounts_df[accounts_df["account_name"] == acct_filter]
        if _aid_row.empty:
            st.error(f"ê³„ì • '{acct_filter}'ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
            return
        _aid = int(_aid_row.iloc[0]["id"])
        acct_where_perf = f"AND ap.account_id = {_aid}"
        acct_where_spend = f"AND ad.account_id = {_aid}"

    st.divider()

    # â”€â”€ KPI â”€â”€
    _render_kpi(d_from_s, d_to_s, acct_where_perf, acct_where_spend, p_cnt, s_cnt)

    st.divider()

    # â”€â”€ ì¼ë³„ ì¶”ì´ â”€â”€
    if p_cnt > 0:
        _render_daily_chart(d_from_s, d_to_s, acct_where_perf)

    # â”€â”€ í•˜ë‹¨ ë¶„ì„ íƒ­ â”€â”€
    if acct_filter == "ì „ì²´":
        sub1, sub2, sub3, sub4 = st.tabs(["ê³„ì •ë³„ ë¹„êµ", "ìº í˜ì¸ë³„", "ìƒí’ˆë³„ TOP", "íš¨ìœ¨ ë¦¬í¬íŠ¸"])
    else:
        sub1, sub2, sub3, sub4 = st.tabs(["ìº í˜ì¸ë³„", "ìƒí’ˆë³„ TOP", "í‚¤ì›Œë“œë³„", "íš¨ìœ¨ ë¦¬í¬íŠ¸"])

    if acct_filter == "ì „ì²´":
        with sub1:
            _render_account_compare(d_from_s, d_to_s, p_cnt, s_cnt)
        with sub2:
            _render_campaign_table(d_from_s, d_to_s, acct_where_perf, p_cnt)
        with sub3:
            _render_product_top(d_from_s, d_to_s, acct_where_perf, p_cnt)
    else:
        with sub1:
            _render_campaign_table(d_from_s, d_to_s, acct_where_perf, p_cnt)
        with sub2:
            _render_product_top(d_from_s, d_to_s, acct_where_perf, p_cnt)
        with sub3:
            _render_keyword_table(d_from_s, d_to_s, acct_where_perf, p_cnt)

    with sub4:
        _render_efficiency_report(d_from_s, d_to_s, acct_where_perf, p_cnt)


def _render_kpi(d_from, d_to, aw_perf, aw_spend, p_cnt, s_cnt):
    """KPI ì¹´ë“œ"""
    spend_total = 0
    revenue_total = 0
    clicks = 0
    impressions = 0
    orders = 0

    if p_cnt > 0:
        kpi = query_df(f"""
            SELECT
                COALESCE(SUM(ap.ad_spend), 0) as spend,
                COALESCE(SUM(ap.total_revenue), 0) as revenue,
                COALESCE(SUM(ap.clicks), 0) as clicks,
                COALESCE(SUM(ap.impressions), 0) as impressions,
                COALESCE(SUM(ap.total_orders), 0) as orders
            FROM ad_performances ap
            WHERE ap.ad_date BETWEEN '{d_from}' AND '{d_to}' {aw_perf}
        """)
        if not kpi.empty:
            k = kpi.iloc[0]
            spend_total = int(k["spend"])
            revenue_total = int(k["revenue"])
            clicks = int(k["clicks"])
            impressions = int(k["impressions"])
            orders = int(k["orders"])

    # ì •ì‚° ë°ì´í„°ê°€ ìˆìœ¼ë©´ ê´‘ê³ ë¹„ëŠ” ì •ì‚° ê¸°ì¤€ ì‚¬ìš© (ë” ì •í™•)
    settle_total = 0
    if s_cnt > 0:
        settle = query_df(f"""
            SELECT COALESCE(SUM(ad.total_charge), 0) as total
            FROM ad_spends ad
            WHERE ad.ad_date BETWEEN '{d_from}' AND '{d_to}' {aw_spend}
        """)
        if not settle.empty:
            settle_total = int(settle.iloc[0]["total"])

    ad_cost = settle_total if settle_total > 0 else spend_total
    roas = round(revenue_total / ad_cost * 100) if ad_cost > 0 else 0
    ctr = round(clicks / impressions * 100, 2) if impressions > 0 else 0
    cpc = round(ad_cost / clicks) if clicks > 0 else 0

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("ê´‘ê³ ë¹„", fmt_krw(ad_cost))
    k2.metric("ì „í™˜ë§¤ì¶œ", fmt_krw(revenue_total))
    k3.metric("ROAS", f"{roas}%")
    k4.metric("ì „í™˜ì£¼ë¬¸", f"{orders:,}ê±´")
    k5.metric("í´ë¦­ìˆ˜", f"{clicks:,}")
    k6.metric("CTR", f"{ctr}%")

    st.caption(f"{d_from} ~ {d_to}  |  CPC í‰ê· : {cpc:,}ì›")


def _render_daily_chart(d_from, d_to, aw_perf):
    """ì¼ë³„ ê´‘ê³ ë¹„ vs ì „í™˜ë§¤ì¶œ ì°¨íŠ¸"""
    daily = query_df(f"""
        SELECT ap.ad_date as ë‚ ì§œ,
            SUM(ap.ad_spend) as ê´‘ê³ ë¹„,
            SUM(ap.total_revenue) as ì „í™˜ë§¤ì¶œ,
            SUM(ap.clicks) as í´ë¦­ìˆ˜,
            SUM(ap.total_orders) as ì£¼ë¬¸ìˆ˜
        FROM ad_performances ap
        WHERE ap.ad_date BETWEEN '{d_from}' AND '{d_to}' {aw_perf}
        GROUP BY ap.ad_date ORDER BY ap.ad_date
    """)
    if daily.empty:
        return

    daily["ë‚ ì§œ"] = pd.to_datetime(daily["ë‚ ì§œ"])

    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Bar(x=daily["ë‚ ì§œ"], y=daily["ê´‘ê³ ë¹„"], name="ê´‘ê³ ë¹„", marker_color="#ff6b6b", opacity=0.7),
        secondary_y=False,
    )
    fig.add_trace(
        go.Bar(x=daily["ë‚ ì§œ"], y=daily["ì „í™˜ë§¤ì¶œ"], name="ì „í™˜ë§¤ì¶œ", marker_color="#4dabf7", opacity=0.7),
        secondary_y=False,
    )
    roas_daily = (daily["ì „í™˜ë§¤ì¶œ"] / daily["ê´‘ê³ ë¹„"].replace(0, pd.NA) * 100).fillna(0)
    fig.add_trace(
        go.Scatter(x=daily["ë‚ ì§œ"], y=roas_daily, name="ROAS(%)", line=dict(color="#51cf66", width=2)),
        secondary_y=True,
    )
    fig.update_layout(
        title="ì¼ë³„ ê´‘ê³ ë¹„ vs ì „í™˜ë§¤ì¶œ",
        barmode="group",
        height=350,
        margin=dict(t=40, b=10, l=10, r=10),
        legend=dict(orientation="h", y=1.12),
    )
    fig.update_yaxes(title_text="ê¸ˆì•¡", secondary_y=False)
    fig.update_yaxes(title_text="ROAS(%)", secondary_y=True)

    st.plotly_chart(fig, use_container_width=True)


def _render_account_compare(d_from, d_to, p_cnt, s_cnt):
    """ê³„ì •ë³„ ë¹„êµ"""
    if p_cnt > 0:
        acct_perf = query_df(f"""
            SELECT a.account_name as ê³„ì •,
                SUM(ap.impressions) as ë…¸ì¶œìˆ˜,
                SUM(ap.clicks) as í´ë¦­ìˆ˜,
                SUM(ap.ad_spend) as ê´‘ê³ ë¹„,
                SUM(ap.total_orders) as ì „í™˜ì£¼ë¬¸,
                SUM(ap.total_revenue) as ì „í™˜ë§¤ì¶œ,
                ROUND(SUM(ap.total_revenue) * 100.0 / NULLIF(SUM(ap.ad_spend), 0), 0) as "ROAS(%)",
                ROUND(SUM(ap.clicks) * 100.0 / NULLIF(SUM(ap.impressions), 0), 2) as "CTR(%)"
            FROM ad_performances ap
            JOIN accounts a ON ap.account_id = a.id
            WHERE ap.ad_date BETWEEN '{d_from}' AND '{d_to}'
            GROUP BY ap.account_id, a.account_name ORDER BY ê´‘ê³ ë¹„ DESC
        """)
        if not acct_perf.empty:
            st.subheader("ê³„ì •ë³„ ê´‘ê³  ì„±ê³¼")

            # ì°¨íŠ¸ + í…Œì´ë¸”
            chart_col, pie_col = st.columns([3, 2])
            with chart_col:
                st.bar_chart(acct_perf.set_index("ê³„ì •")[["ê´‘ê³ ë¹„", "ì „í™˜ë§¤ì¶œ"]])
            with pie_col:
                import plotly.express as px
                _pie = acct_perf[acct_perf["ê´‘ê³ ë¹„"] > 0]
                if not _pie.empty:
                    fig = px.pie(_pie, values="ê´‘ê³ ë¹„", names="ê³„ì •", title="ê´‘ê³ ë¹„ ë¹„ì¤‘",
                                 hole=0.4, color_discrete_sequence=px.colors.qualitative.Set2)
                    fig.update_layout(margin=dict(t=40, b=10, l=10, r=10), height=300, showlegend=True)
                    st.plotly_chart(fig, use_container_width=True)

            st.dataframe(fmt_money_df(acct_perf), hide_index=True, use_container_width=True)

    if s_cnt > 0:
        acct_spend = query_df(f"""
            SELECT a.account_name as ê³„ì •,
                COUNT(DISTINCT ad.campaign_id) as ìº í˜ì¸ìˆ˜,
                SUM(ad.spent_amount) as ì†Œì§„ì•¡,
                SUM(ad.total_charge) as ì´ë¹„ìš©
            FROM ad_spends ad
            JOIN accounts a ON ad.account_id = a.id
            WHERE ad.ad_date BETWEEN '{d_from}' AND '{d_to}'
            GROUP BY ad.account_id, a.account_name ORDER BY ì´ë¹„ìš© DESC
        """)
        if not acct_spend.empty:
            st.subheader("ê³„ì •ë³„ ê´‘ê³ ë¹„ ì •ì‚°")
            st.dataframe(fmt_money_df(acct_spend), hide_index=True, use_container_width=True)


def _render_campaign_table(d_from, d_to, aw_perf, p_cnt):
    """ìº í˜ì¸ë³„ ì„±ê³¼"""
    if p_cnt == 0:
        st.info("ê´‘ê³  ì„±ê³¼ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        return

    campaigns = query_df(f"""
        SELECT ap.campaign_name, ap.campaign_id,
            SUM(ap.impressions) as ë…¸ì¶œìˆ˜,
            SUM(ap.clicks) as í´ë¦­ìˆ˜,
            ROUND(SUM(ap.clicks) * 100.0 / NULLIF(SUM(ap.impressions), 0), 2) as "CTR(%)",
            SUM(ap.ad_spend) as ê´‘ê³ ë¹„,
            SUM(ap.total_orders) as ì „í™˜ì£¼ë¬¸,
            SUM(ap.total_revenue) as ì „í™˜ë§¤ì¶œ,
            ROUND(SUM(ap.total_revenue) * 100.0 / NULLIF(SUM(ap.ad_spend), 0), 0) as "ROAS(%)"
        FROM ad_performances ap
        WHERE ap.ad_date BETWEEN '{d_from}' AND '{d_to}' {aw_perf}
            AND ap.campaign_name != ''
        GROUP BY ap.campaign_name, ap.campaign_id ORDER BY ê´‘ê³ ë¹„ DESC
        LIMIT 30
    """)
    if not campaigns.empty:
        campaigns["ìº í˜ì¸"] = campaigns.apply(
            lambda r: f"{r['campaign_name']} ({r['campaign_id']})" if r["campaign_id"] else r["campaign_name"], axis=1)
        campaigns = campaigns.drop(columns=["campaign_name", "campaign_id"])
        cols = ["ìº í˜ì¸"] + [c for c in campaigns.columns if c != "ìº í˜ì¸"]
        campaigns = campaigns[cols]
    if not campaigns.empty:
        st.dataframe(fmt_money_df(campaigns), hide_index=True, use_container_width=True)
        xl = _df_to_excel_bytes(campaigns, "ìº í˜ì¸ë³„")
        st.download_button("Excel ë‹¤ìš´ë¡œë“œ", xl, "campaigns.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_campaigns")
    else:
        st.info("ìº í˜ì¸ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")


def _render_product_top(d_from, d_to, aw_perf, p_cnt):
    """ìƒí’ˆë³„ TOP ì„±ê³¼"""
    if p_cnt == 0:
        st.info("ê´‘ê³  ì„±ê³¼ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        return

    products = query_df(f"""
        SELECT ap.campaign_name, ap.campaign_id, ap.product_name as ìƒí’ˆëª…,
            SUM(ap.impressions) as ë…¸ì¶œìˆ˜,
            SUM(ap.clicks) as í´ë¦­ìˆ˜,
            SUM(ap.ad_spend) as ê´‘ê³ ë¹„,
            SUM(ap.total_orders) as ì „í™˜ì£¼ë¬¸,
            SUM(ap.total_revenue) as ì „í™˜ë§¤ì¶œ,
            ROUND(SUM(ap.total_revenue) * 100.0 / NULLIF(SUM(ap.ad_spend), 0), 0) as "ROAS(%)"
        FROM ad_performances ap
        WHERE ap.ad_date BETWEEN '{d_from}' AND '{d_to}' {aw_perf}
            AND ap.product_name != ''
        GROUP BY ap.campaign_name, ap.campaign_id, ap.coupang_product_id, ap.product_name ORDER BY ì „í™˜ë§¤ì¶œ DESC
        LIMIT 20
    """)
    if not products.empty:
        products["ìº í˜ì¸"] = products.apply(
            lambda r: f"{r['campaign_name']} ({r['campaign_id']})" if r["campaign_id"] else r["campaign_name"], axis=1)
        products = products.drop(columns=["campaign_name", "campaign_id"])
        cols = ["ìº í˜ì¸"] + [c for c in products.columns if c != "ìº í˜ì¸"]
        products = products[cols]
    if not products.empty:
        products.insert(0, "#", range(1, len(products) + 1))
        st.dataframe(fmt_money_df(products), hide_index=True, use_container_width=True)
        xl = _df_to_excel_bytes(products, "ìƒí’ˆë³„TOP")
        st.download_button("Excel ë‹¤ìš´ë¡œë“œ", xl, "products_ad.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_products_ad")
    else:
        st.info("ìƒí’ˆ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")


def _render_keyword_table(d_from, d_to, aw_perf, p_cnt):
    """í‚¤ì›Œë“œë³„ ì„±ê³¼ (ê°œë³„ ê³„ì • ì„ íƒ ì‹œ)"""
    if p_cnt == 0:
        st.info("ê´‘ê³  ì„±ê³¼ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        return

    keywords = query_df(f"""
        SELECT ap.campaign_name, ap.campaign_id,
            ap.keyword as í‚¤ì›Œë“œ,
            ap.match_type as ë§¤ì¹˜ìœ í˜•,
            SUM(ap.impressions) as ë…¸ì¶œìˆ˜,
            SUM(ap.clicks) as í´ë¦­ìˆ˜,
            ROUND(SUM(ap.clicks) * 100.0 / NULLIF(SUM(ap.impressions), 0), 2) as "CTR(%)",
            SUM(ap.ad_spend) as ê´‘ê³ ë¹„,
            SUM(ap.total_orders) as ì „í™˜ì£¼ë¬¸,
            SUM(ap.total_revenue) as ì „í™˜ë§¤ì¶œ,
            ROUND(SUM(ap.total_revenue) * 100.0 / NULLIF(SUM(ap.ad_spend), 0), 0) as "ROAS(%)"
        FROM ad_performances ap
        WHERE ap.ad_date BETWEEN '{d_from}' AND '{d_to}' {aw_perf}
            AND ap.keyword != ''
        GROUP BY ap.campaign_name, ap.campaign_id, ap.keyword, ap.match_type ORDER BY ê´‘ê³ ë¹„ DESC
        LIMIT 30
    """)
    if not keywords.empty:
        keywords["ìº í˜ì¸"] = keywords.apply(
            lambda r: f"{r['campaign_name']} ({r['campaign_id']})" if r["campaign_id"] else r["campaign_name"], axis=1)
        keywords = keywords.drop(columns=["campaign_name", "campaign_id"])
        cols = ["ìº í˜ì¸"] + [c for c in keywords.columns if c != "ìº í˜ì¸"]
        keywords = keywords[cols]
    if not keywords.empty:
        st.dataframe(fmt_money_df(keywords), hide_index=True, use_container_width=True)
        xl = _df_to_excel_bytes(keywords, "í‚¤ì›Œë“œë³„")
        st.download_button("Excel ë‹¤ìš´ë¡œë“œ", xl, "keywords.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_keywords")
    else:
        st.info("í‚¤ì›Œë“œ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# Excel ìœ í‹¸ë¦¬í‹°
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    """DataFrame â†’ Excel bytes (ë‹¨ì¼ ì‹œíŠ¸)"""
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        ws = writer.sheets[sheet_name]
        _style_excel_header(ws, len(df.columns), len(df), sheet_name)
    return buf.getvalue()


def _style_excel_header(ws, num_cols: int, num_rows: int, title: str = ""):
    """Excel ì‹œíŠ¸ í—¤ë” ìŠ¤íƒ€ì¼ë§ (export_order_sheets.py íŒ¨í„´)"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # íƒ€ì´í‹€ í–‰
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(num_cols, 1))
        cell = ws.cell(row=1, column=1)
        cell.value = title
        cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(horizontal="center")

    # í—¤ë” í–‰ (row 2)
    for ci in range(1, num_cols + 1):
        c = ws.cell(row=2, column=ci)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    # ë°ì´í„° í–‰ í…Œë‘ë¦¬
    for ri in range(3, 3 + num_rows):
        for ci in range(1, num_cols + 1):
            ws.cell(row=ri, column=ci).border = thin_border

    # ì—´ ë„ˆë¹„ ìë™ ì¡°ì • (ìµœì†Œ 12, ìµœëŒ€ 40)
    for ci in range(1, num_cols + 1):
        header_val = str(ws.cell(row=2, column=ci).value or "")
        max_len = max(len(header_val) * 2, 12)  # í•œê¸€ 2ë°°
        for ri in range(3, min(3 + num_rows, 53)):  # ìµœëŒ€ 50í–‰ ìƒ˜í”Œ
            val = str(ws.cell(row=ri, column=ci).value or "")
            max_len = max(max_len, min(len(val) + 2, 40))
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].width = min(max_len, 40)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# íš¨ìœ¨ ë¦¬í¬íŠ¸
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _render_efficiency_report(d_from, d_to, aw_perf, p_cnt):
    """ê´‘ê³  íš¨ìœ¨ ë¦¬í¬íŠ¸ â€” ë¹„íš¨ìœ¨ í‚¤ì›Œë“œ/ìƒí’ˆ/ìº í˜ì¸ ë¶„ì„ + Excel ë‹¤ìš´ë¡œë“œ"""
    if p_cnt == 0:
        st.info("ê´‘ê³  ì„±ê³¼ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        return

    # â”€â”€ ì´ ê´‘ê³ ë¹„ â”€â”€
    total_kpi = query_df(f"""
        SELECT COALESCE(SUM(ap.ad_spend), 0) as total_spend
        FROM ad_performances ap
        WHERE ap.ad_date BETWEEN '{d_from}' AND '{d_to}' {aw_perf}
    """)
    total_spend = int(total_kpi.iloc[0]["total_spend"]) if not total_kpi.empty else 0

    if total_spend == 0:
        st.info("ì„ íƒí•œ ê¸°ê°„ì— ê´‘ê³ ë¹„ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        return

    # â”€â”€ í‚¤ì›Œë“œë³„ ì§‘ê³„ â”€â”€
    kw_df = query_df(f"""
        SELECT ap.campaign_name, ap.campaign_id,
            ap.keyword as í‚¤ì›Œë“œ, ap.match_type as ë§¤ì¹˜ìœ í˜•,
            SUM(ap.impressions) as ë…¸ì¶œìˆ˜,
            SUM(ap.clicks) as í´ë¦­ìˆ˜,
            SUM(ap.ad_spend) as ê´‘ê³ ë¹„,
            SUM(ap.total_orders) as ì „í™˜ì£¼ë¬¸,
            SUM(ap.total_revenue) as ì „í™˜ë§¤ì¶œ,
            ROUND(SUM(ap.ad_spend) * 1.0 / NULLIF(SUM(ap.clicks), 0), 0) as "CPC",
            ROUND(SUM(ap.total_revenue) * 100.0 / NULLIF(SUM(ap.ad_spend), 0), 0) as "ROAS(%)"
        FROM ad_performances ap
        WHERE ap.ad_date BETWEEN '{d_from}' AND '{d_to}' {aw_perf}
            AND ap.keyword != '' AND ap.keyword != '-'
        GROUP BY ap.campaign_name, ap.campaign_id, ap.keyword, ap.match_type
        HAVING SUM(ap.ad_spend) > 0
        ORDER BY ê´‘ê³ ë¹„ DESC
    """)
    if not kw_df.empty:
        kw_df["ìº í˜ì¸"] = kw_df.apply(
            lambda r: f"{r['campaign_name']} ({r['campaign_id']})" if r["campaign_id"] else r["campaign_name"], axis=1)
        kw_df = kw_df.drop(columns=["campaign_name", "campaign_id"])
        cols = ["ìº í˜ì¸"] + [c for c in kw_df.columns if c != "ìº í˜ì¸"]
        kw_df = kw_df[cols]

    # â”€â”€ ìƒí’ˆë³„ ì§‘ê³„ â”€â”€
    prod_df = query_df(f"""
        SELECT ap.campaign_name, ap.campaign_id, ap.product_name as ìƒí’ˆëª…,
            SUM(ap.impressions) as ë…¸ì¶œìˆ˜,
            SUM(ap.clicks) as í´ë¦­ìˆ˜,
            SUM(ap.ad_spend) as ê´‘ê³ ë¹„,
            SUM(ap.total_orders) as ì „í™˜ì£¼ë¬¸,
            SUM(ap.total_revenue) as ì „í™˜ë§¤ì¶œ,
            ROUND(SUM(ap.total_revenue) * 100.0 / NULLIF(SUM(ap.ad_spend), 0), 0) as "ROAS(%)"
        FROM ad_performances ap
        WHERE ap.ad_date BETWEEN '{d_from}' AND '{d_to}' {aw_perf}
            AND ap.product_name != ''
        GROUP BY ap.campaign_name, ap.campaign_id, ap.coupang_product_id, ap.product_name
        HAVING SUM(ap.ad_spend) > 0
        ORDER BY ê´‘ê³ ë¹„ DESC
    """)
    if not prod_df.empty:
        prod_df["ìº í˜ì¸"] = prod_df.apply(
            lambda r: f"{r['campaign_name']} ({r['campaign_id']})" if r["campaign_id"] else r["campaign_name"], axis=1)
        prod_df = prod_df.drop(columns=["campaign_name", "campaign_id"])
        cols = ["ìº í˜ì¸"] + [c for c in prod_df.columns if c != "ìº í˜ì¸"]
        prod_df = prod_df[cols]

    # â”€â”€ ìº í˜ì¸ë³„ ì§‘ê³„ â”€â”€
    camp_df = query_df(f"""
        SELECT ap.campaign_name, ap.campaign_id,
            SUM(ap.impressions) as ë…¸ì¶œìˆ˜,
            SUM(ap.clicks) as í´ë¦­ìˆ˜,
            SUM(ap.ad_spend) as ê´‘ê³ ë¹„,
            SUM(ap.total_orders) as ì „í™˜ì£¼ë¬¸,
            SUM(ap.total_revenue) as ì „í™˜ë§¤ì¶œ,
            ROUND(SUM(ap.total_revenue) * 100.0 / NULLIF(SUM(ap.ad_spend), 0), 0) as "ROAS(%)"
        FROM ad_performances ap
        WHERE ap.ad_date BETWEEN '{d_from}' AND '{d_to}' {aw_perf}
            AND ap.campaign_name != ''
        GROUP BY ap.campaign_name, ap.campaign_id
        HAVING SUM(ap.ad_spend) > 0
        ORDER BY ê´‘ê³ ë¹„ DESC
    """)
    if not camp_df.empty:
        camp_df["ìº í˜ì¸"] = camp_df.apply(
            lambda r: f"{r['campaign_name']} ({r['campaign_id']})" if r["campaign_id"] else r["campaign_name"], axis=1)
        camp_df = camp_df.drop(columns=["campaign_name", "campaign_id"])
        cols = ["ìº í˜ì¸"] + [c for c in camp_df.columns if c != "ìº í˜ì¸"]
        camp_df = camp_df[cols]

    # â”€â”€ ìº í˜ì¸ í•„í„° â”€â”€
    all_campaigns = sorted(set(
        (kw_df["ìº í˜ì¸"].unique().tolist() if not kw_df.empty else [])
        + (prod_df["ìº í˜ì¸"].unique().tolist() if not prod_df.empty else [])
        + (camp_df["ìº í˜ì¸"].unique().tolist() if not camp_df.empty else [])
    ))
    camp_filter = st.selectbox(
        "ìº í˜ì¸ í•„í„°", ["ì „ì²´"] + all_campaigns, key="eff_camp_filter")

    if camp_filter != "ì „ì²´":
        kw_df = kw_df[kw_df["ìº í˜ì¸"] == camp_filter] if not kw_df.empty else kw_df
        prod_df = prod_df[prod_df["ìº í˜ì¸"] == camp_filter] if not prod_df.empty else prod_df
        camp_df = camp_df[camp_df["ìº í˜ì¸"] == camp_filter] if not camp_df.empty else camp_df
        # ì„ íƒ ìº í˜ì¸ ê¸°ì¤€ ì´ ê´‘ê³ ë¹„ ì¬ê³„ì‚°
        total_spend = int(kw_df["ê´‘ê³ ë¹„"].sum()) + int(prod_df[~prod_df["ìƒí’ˆëª…"].isin([""])]["ê´‘ê³ ë¹„"].sum()) if not kw_df.empty else 0
        if not camp_df.empty:
            total_spend = int(camp_df["ê´‘ê³ ë¹„"].sum())

    st.divider()

    # â”€â”€ ë¶„ë¥˜ â”€â”€
    zero_conv_kw = kw_df[kw_df["ì „í™˜ì£¼ë¬¸"] == 0] if not kw_df.empty else pd.DataFrame()
    low_roas_kw = kw_df[(kw_df["ì „í™˜ì£¼ë¬¸"] > 0) & (kw_df["ROAS(%)"].fillna(0) < 200)] if not kw_df.empty else pd.DataFrame()
    good_kw = kw_df[(kw_df["ì „í™˜ì£¼ë¬¸"] > 0) & (kw_df["ROAS(%)"].fillna(0) >= 200)].sort_values("ì „í™˜ë§¤ì¶œ", ascending=False) if not kw_df.empty else pd.DataFrame()

    zero_conv_prod = prod_df[prod_df["ì „í™˜ì£¼ë¬¸"] == 0] if not prod_df.empty else pd.DataFrame()
    good_prod = prod_df[(prod_df["ì „í™˜ì£¼ë¬¸"] > 0) & (prod_df["ROAS(%)"].fillna(0) >= 200)].sort_values("ì „í™˜ë§¤ì¶œ", ascending=False) if not prod_df.empty else pd.DataFrame()

    # â”€â”€ (A) ìƒë‹¨ KPI â”€â”€
    wasted_spend = int(zero_conv_kw["ê´‘ê³ ë¹„"].sum()) if not zero_conv_kw.empty else 0
    waste_pct = round(wasted_spend / total_spend * 100, 1) if total_spend > 0 else 0
    efficient_kw_count = len(good_kw)

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("ì´ ê´‘ê³ ë¹„", fmt_krw(total_spend))
    k2.metric("ë‚­ë¹„ ê´‘ê³ ë¹„ (ì „í™˜ 0ê±´)", fmt_krw(wasted_spend))
    k3.metric("ë‚­ë¹„ ë¹„ìœ¨", f"{waste_pct}%")
    k4.metric("íš¨ìœ¨ í‚¤ì›Œë“œ ìˆ˜", f"{efficient_kw_count}ê°œ")

    st.divider()

    # â”€â”€ (B) í‚¤ì›Œë“œ ë¹„íš¨ìœ¨ ë¶„ì„ â”€â”€
    st.subheader("í‚¤ì›Œë“œ ë¶„ì„")

    if not kw_df.empty:
        if not zero_conv_kw.empty:
            st.markdown(f"**ì „í™˜ 0ê±´ í‚¤ì›Œë“œ** ({len(zero_conv_kw)}ê°œ, ë‚­ë¹„ {fmt_krw(wasted_spend)})")
            st.dataframe(fmt_money_df(zero_conv_kw.head(30)), hide_index=True, use_container_width=True)

        if not low_roas_kw.empty:
            low_roas_spend = int(low_roas_kw["ê´‘ê³ ë¹„"].sum())
            st.markdown(f"**ROAS 200% ë¯¸ë§Œ â€” ì ì í‚¤ì›Œë“œ** ({len(low_roas_kw)}ê°œ, ê´‘ê³ ë¹„ {fmt_krw(low_roas_spend)})")
            st.dataframe(fmt_money_df(low_roas_kw), hide_index=True, use_container_width=True)

        if not good_kw.empty:
            st.markdown(f"**íš¨ìœ¨ ì¢‹ì€ í‚¤ì›Œë“œ (ROAS >= 200%)** ({len(good_kw)}ê°œ)")
            st.dataframe(fmt_money_df(good_kw), hide_index=True, use_container_width=True)
    else:
        st.info("í‚¤ì›Œë“œ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")

    st.divider()

    # â”€â”€ (C) ìƒí’ˆ ë¹„íš¨ìœ¨ ë¶„ì„ â”€â”€
    st.subheader("ìƒí’ˆ ë¶„ì„")

    if not prod_df.empty:
        if not zero_conv_prod.empty:
            wasted_prod = int(zero_conv_prod["ê´‘ê³ ë¹„"].sum())
            st.markdown(f"**ì „í™˜ 0ê±´ ìƒí’ˆ** ({len(zero_conv_prod)}ê°œ, ë‚­ë¹„ {fmt_krw(wasted_prod)})")
            st.dataframe(fmt_money_df(zero_conv_prod), hide_index=True, use_container_width=True)

        if not good_prod.empty:
            st.markdown(f"**íš¨ìœ¨ ì¢‹ì€ ìƒí’ˆ (ROAS >= 200%)** ({len(good_prod)}ê°œ)")
            st.dataframe(fmt_money_df(good_prod), hide_index=True, use_container_width=True)
    else:
        st.info("ìƒí’ˆ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")

    st.divider()

    # â”€â”€ (D) ìº í˜ì¸ ë¹„íš¨ìœ¨ ë¶„ì„ â”€â”€
    st.subheader("ìº í˜ì¸ ROAS ë¹„êµ")

    if not camp_df.empty:
        st.dataframe(fmt_money_df(camp_df), hide_index=True, use_container_width=True)
    else:
        st.info("ìº í˜ì¸ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")

    # â”€â”€ (E) ê°œì„  ì œì•ˆ â”€â”€
    st.divider()
    st.subheader("ê°œì„  ì œì•ˆ")

    recs_df = _generate_recommendations(kw_df, prod_df, camp_df)

    if not recs_df.empty:
        # ìš”ì•½ ì¹´ë“œ
        n_stop = len(recs_df[recs_df["ì¡°ì¹˜"].str.contains("ì¤‘ì§€")])
        n_adjust = len(recs_df[recs_df["ì¡°ì¹˜"].str.contains("í•˜í–¥|ì ê²€|ì •ë¦¬", regex=True)])
        n_expand = len(recs_df[recs_df["ì¡°ì¹˜"].str.contains("í™•ëŒ€|ìƒí–¥", regex=True)])

        c1, c2, c3 = st.columns(3)
        c1.metric("ì¦‰ì‹œ ì¤‘ì§€ ê¶Œì¥", f"{n_stop}ê±´")
        c2.metric("ì…ì°°ê°€/í‚¤ì›Œë“œ ì¡°ì •", f"{n_adjust}ê±´")
        c3.metric("ì˜ˆì‚° í™•ëŒ€ ê²€í† ", f"{n_expand}ê±´")

        # ìš°ì„ ìˆœìœ„ ì•„ì´ì½˜ ë§¤í•‘
        display_df = recs_df.copy()
        icon_map = {"ë†’ìŒ": "ğŸ”´ ë†’ìŒ", "ì¤‘ê°„": "ğŸŸ¡ ì¤‘ê°„", "ë‚®ìŒ": "ğŸŸ¢ ë‚®ìŒ"}
        display_df["ìš°ì„ ìˆœìœ„"] = display_df["ìš°ì„ ìˆœìœ„"].map(icon_map)

        st.dataframe(
            fmt_money_df(display_df[["ëŒ€ìƒìœ í˜•", "ìº í˜ì¸", "ì´ë¦„", "ROAS(%)", "ê´‘ê³ ë¹„", "ì¡°ì¹˜", "ìš°ì„ ìˆœìœ„", "ì‚¬ìœ "]]),
            hide_index=True,
            use_container_width=True,
        )
    else:
        st.info("ì¡°ì¹˜ê°€ í•„ìš”í•œ í•­ëª©ì´ ì—†ìŠµë‹ˆë‹¤.")

    st.divider()

    # â”€â”€ (F) Excel ë‹¤ìš´ë¡œë“œ â”€â”€
    xl_bytes = _create_efficiency_excel(
        total_spend, wasted_spend, waste_pct, efficient_kw_count,
        zero_conv_kw, good_kw, zero_conv_prod, good_prod, camp_df,
        recs_df, d_from, d_to,
    )
    st.download_button(
        "íš¨ìœ¨ ë¦¬í¬íŠ¸ Excel ë‹¤ìš´ë¡œë“œ",
        xl_bytes,
        f"ê´‘ê³ _íš¨ìœ¨_ë¦¬í¬íŠ¸_{d_from}_{d_to}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_efficiency_excel",
        type="primary",
        use_container_width=True,
    )


def _create_efficiency_excel(total_spend, wasted_spend, waste_pct, efficient_kw_count,
                              zero_conv_kw, good_kw, zero_conv_prod, good_prod,
                              camp_df, recs_df, d_from, d_to):
    """íš¨ìœ¨ ë¦¬í¬íŠ¸ ë©€í‹°ì‹œíŠ¸ Excel ìƒì„±"""
    from io import BytesIO

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        period_label = f"{d_from} ~ {d_to}"

        # Sheet 1: ìš”ì•½
        summary = pd.DataFrame([
            {"í•­ëª©": "ì´ ê´‘ê³ ë¹„", "ê°’": f"{total_spend:,}ì›"},
            {"í•­ëª©": "ë‚­ë¹„ ê´‘ê³ ë¹„ (ì „í™˜ 0ê±´ í‚¤ì›Œë“œ)", "ê°’": f"{wasted_spend:,}ì›"},
            {"í•­ëª©": "ë‚­ë¹„ ë¹„ìœ¨", "ê°’": f"{waste_pct}%"},
            {"í•­ëª©": "íš¨ìœ¨ í‚¤ì›Œë“œ ìˆ˜ (ROAS>=200%)", "ê°’": f"{efficient_kw_count}ê°œ"},
            {"í•­ëª©": "ë¶„ì„ ê¸°ê°„", "ê°’": period_label},
        ])
        summary.to_excel(writer, sheet_name="ìš”ì•½", index=False, startrow=1)
        ws = writer.sheets["ìš”ì•½"]
        _style_excel_header(ws, 2, len(summary), f"ê´‘ê³  íš¨ìœ¨ ìš”ì•½ ({period_label})")
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        # Sheet 2: ë¹„íš¨ìœ¨ í‚¤ì›Œë“œ
        if not zero_conv_kw.empty:
            zero_conv_kw.to_excel(writer, sheet_name="ë¹„íš¨ìœ¨ í‚¤ì›Œë“œ", index=False, startrow=1)
            ws = writer.sheets["ë¹„íš¨ìœ¨ í‚¤ì›Œë“œ"]
            _style_excel_header(ws, len(zero_conv_kw.columns), len(zero_conv_kw),
                                f"ì „í™˜ 0ê±´ í‚¤ì›Œë“œ ({period_label})")
        else:
            pd.DataFrame({"ë©”ì‹œì§€": ["ì „í™˜ 0ê±´ í‚¤ì›Œë“œ ì—†ìŒ"]}).to_excel(
                writer, sheet_name="ë¹„íš¨ìœ¨ í‚¤ì›Œë“œ", index=False)

        # Sheet 3: íš¨ìœ¨ í‚¤ì›Œë“œ
        if not good_kw.empty:
            good_kw.to_excel(writer, sheet_name="íš¨ìœ¨ í‚¤ì›Œë“œ", index=False, startrow=1)
            ws = writer.sheets["íš¨ìœ¨ í‚¤ì›Œë“œ"]
            _style_excel_header(ws, len(good_kw.columns), len(good_kw),
                                f"íš¨ìœ¨ í‚¤ì›Œë“œ ROAS>=200% ({period_label})")
        else:
            pd.DataFrame({"ë©”ì‹œì§€": ["ROAS>=200% í‚¤ì›Œë“œ ì—†ìŒ"]}).to_excel(
                writer, sheet_name="íš¨ìœ¨ í‚¤ì›Œë“œ", index=False)

        # Sheet 4: ë¹„íš¨ìœ¨ ìƒí’ˆ
        if not zero_conv_prod.empty:
            zero_conv_prod.to_excel(writer, sheet_name="ë¹„íš¨ìœ¨ ìƒí’ˆ", index=False, startrow=1)
            ws = writer.sheets["ë¹„íš¨ìœ¨ ìƒí’ˆ"]
            _style_excel_header(ws, len(zero_conv_prod.columns), len(zero_conv_prod),
                                f"ì „í™˜ 0ê±´ ìƒí’ˆ ({period_label})")
        else:
            pd.DataFrame({"ë©”ì‹œì§€": ["ì „í™˜ 0ê±´ ìƒí’ˆ ì—†ìŒ"]}).to_excel(
                writer, sheet_name="ë¹„íš¨ìœ¨ ìƒí’ˆ", index=False)

        # Sheet 5: íš¨ìœ¨ ìƒí’ˆ
        if not good_prod.empty:
            good_prod.to_excel(writer, sheet_name="íš¨ìœ¨ ìƒí’ˆ", index=False, startrow=1)
            ws = writer.sheets["íš¨ìœ¨ ìƒí’ˆ"]
            _style_excel_header(ws, len(good_prod.columns), len(good_prod),
                                f"íš¨ìœ¨ ìƒí’ˆ ROAS>=200% ({period_label})")
        else:
            pd.DataFrame({"ë©”ì‹œì§€": ["ROAS>=200% ìƒí’ˆ ì—†ìŒ"]}).to_excel(
                writer, sheet_name="íš¨ìœ¨ ìƒí’ˆ", index=False)

        # Sheet 6: ìº í˜ì¸ ë¹„êµ
        if not camp_df.empty:
            camp_df.to_excel(writer, sheet_name="ìº í˜ì¸ ë¹„êµ", index=False, startrow=1)
            ws = writer.sheets["ìº í˜ì¸ ë¹„êµ"]
            _style_excel_header(ws, len(camp_df.columns), len(camp_df),
                                f"ìº í˜ì¸ë³„ ì„±ê³¼ ({period_label})")
        else:
            pd.DataFrame({"ë©”ì‹œì§€": ["ìº í˜ì¸ ë°ì´í„° ì—†ìŒ"]}).to_excel(
                writer, sheet_name="ìº í˜ì¸ ë¹„êµ", index=False)

        # Sheet 7: ê°œì„  ì œì•ˆ
        if recs_df is not None and not recs_df.empty:
            recs_df.to_excel(writer, sheet_name="ê°œì„  ì œì•ˆ", index=False, startrow=1)
            ws = writer.sheets["ê°œì„  ì œì•ˆ"]
            _style_excel_header(ws, len(recs_df.columns), len(recs_df),
                                f"ê°œì„  ì œì•ˆ ({period_label})")
        else:
            pd.DataFrame({"ë©”ì‹œì§€": ["ì¡°ì¹˜ê°€ í•„ìš”í•œ í•­ëª© ì—†ìŒ"]}).to_excel(
                writer, sheet_name="ê°œì„  ì œì•ˆ", index=False)

    return buf.getvalue()


def _generate_recommendations(kw_df, prod_df, camp_df):
    """í‚¤ì›Œë“œ/ìƒí’ˆ/ìº í˜ì¸ë³„ ì¡°ì¹˜ ê¶Œì¥ ëª©ë¡ ìƒì„±"""
    recs = []

    # â”€â”€ í‚¤ì›Œë“œ ì¡°ì¹˜ â”€â”€
    if not kw_df.empty:
        for _, r in kw_df.iterrows():
            conv = int(r.get("ì „í™˜ì£¼ë¬¸", 0) or 0)
            spend = int(r.get("ê´‘ê³ ë¹„", 0) or 0)
            roas = float(r.get("ROAS(%)", 0) or 0)
            clicks = int(r.get("í´ë¦­ìˆ˜", 0) or 0)
            impressions = int(r.get("ë…¸ì¶œìˆ˜", 0) or 0)
            ctr = (clicks / impressions * 100) if impressions > 0 else 0
            name = r.get("í‚¤ì›Œë“œ", "")
            campaign = r.get("ìº í˜ì¸", "")

            if conv == 0 and ctr >= 5:
                recs.append({"ëŒ€ìƒìœ í˜•": "í‚¤ì›Œë“œ", "ìº í˜ì¸": campaign, "ì´ë¦„": name,
                             "ROAS(%)": roas, "ê´‘ê³ ë¹„": spend,
                             "ì¡°ì¹˜": "ìƒí’ˆí˜ì´ì§€ ì ê²€", "ìš°ì„ ìˆœìœ„": "ë†’ìŒ",
                             "ì‚¬ìœ ": f"CTR {ctr:.1f}%ë¡œ ë†’ìœ¼ë‚˜ ì „í™˜ 0ê±´"})
            elif conv == 0 and spend >= 5000:
                recs.append({"ëŒ€ìƒìœ í˜•": "í‚¤ì›Œë“œ", "ìº í˜ì¸": campaign, "ì´ë¦„": name,
                             "ROAS(%)": roas, "ê´‘ê³ ë¹„": spend,
                             "ì¡°ì¹˜": "í‚¤ì›Œë“œ ì¤‘ì§€", "ìš°ì„ ìˆœìœ„": "ë†’ìŒ",
                             "ì‚¬ìœ ": f"ì „í™˜ 0ê±´, ê´‘ê³ ë¹„ {spend:,}ì› ì†Œì§„"})
            elif conv == 0 and spend < 5000:
                recs.append({"ëŒ€ìƒìœ í˜•": "í‚¤ì›Œë“œ", "ìº í˜ì¸": campaign, "ì´ë¦„": name,
                             "ROAS(%)": roas, "ê´‘ê³ ë¹„": spend,
                             "ì¡°ì¹˜": "ëª¨ë‹ˆí„°ë§", "ìš°ì„ ìˆœìœ„": "ë‚®ìŒ",
                             "ì‚¬ìœ ": "ì „í™˜ 0ê±´, ë°ì´í„° ë¶€ì¡±"})
            elif roas < 100:
                recs.append({"ëŒ€ìƒìœ í˜•": "í‚¤ì›Œë“œ", "ìº í˜ì¸": campaign, "ì´ë¦„": name,
                             "ROAS(%)": roas, "ê´‘ê³ ë¹„": spend,
                             "ì¡°ì¹˜": "ì…ì°°ê°€ í•˜í–¥ ë˜ëŠ” ì¤‘ì§€", "ìš°ì„ ìˆœìœ„": "ë†’ìŒ",
                             "ì‚¬ìœ ": f"ROAS {roas:.0f}% ì ì"})
            elif roas < 200:
                recs.append({"ëŒ€ìƒìœ í˜•": "í‚¤ì›Œë“œ", "ìº í˜ì¸": campaign, "ì´ë¦„": name,
                             "ROAS(%)": roas, "ê´‘ê³ ë¹„": spend,
                             "ì¡°ì¹˜": "ì…ì°°ê°€ í•˜í–¥ ê²€í† ", "ìš°ì„ ìˆœìœ„": "ì¤‘ê°„",
                             "ì‚¬ìœ ": f"ROAS {roas:.0f}% ì €íš¨ìœ¨"})
            elif roas >= 500 and clicks >= 10:
                recs.append({"ëŒ€ìƒìœ í˜•": "í‚¤ì›Œë“œ", "ìº í˜ì¸": campaign, "ì´ë¦„": name,
                             "ROAS(%)": roas, "ê´‘ê³ ë¹„": spend,
                             "ì¡°ì¹˜": "ì…ì°°ê°€ ìƒí–¥ ê²€í† ", "ìš°ì„ ìˆœìœ„": "ì¤‘ê°„",
                             "ì‚¬ìœ ": f"ROAS {roas:.0f}%, í´ë¦­ {clicks}íšŒ â€” í™•ëŒ€ ì—¬ì§€"})

    # â”€â”€ ìƒí’ˆ ì¡°ì¹˜ â”€â”€
    if not prod_df.empty:
        for _, r in prod_df.iterrows():
            conv = int(r.get("ì „í™˜ì£¼ë¬¸", 0) or 0)
            spend = int(r.get("ê´‘ê³ ë¹„", 0) or 0)
            roas = float(r.get("ROAS(%)", 0) or 0)
            name = r.get("ìƒí’ˆëª…", "")
            campaign = r.get("ìº í˜ì¸", "")

            if conv == 0 and spend >= 10000:
                recs.append({"ëŒ€ìƒìœ í˜•": "ìƒí’ˆ", "ìº í˜ì¸": campaign, "ì´ë¦„": name,
                             "ROAS(%)": roas, "ê´‘ê³ ë¹„": spend,
                             "ì¡°ì¹˜": "ê´‘ê³  ì¤‘ì§€", "ìš°ì„ ìˆœìœ„": "ë†’ìŒ",
                             "ì‚¬ìœ ": f"ì „í™˜ 0ê±´, ê´‘ê³ ë¹„ {spend:,}ì› ì†Œì§„"})
            elif conv > 0 and roas < 100:
                recs.append({"ëŒ€ìƒìœ í˜•": "ìƒí’ˆ", "ìº í˜ì¸": campaign, "ì´ë¦„": name,
                             "ROAS(%)": roas, "ê´‘ê³ ë¹„": spend,
                             "ì¡°ì¹˜": "ì˜ˆì‚° ì¶•ì†Œ ë˜ëŠ” ì¤‘ì§€", "ìš°ì„ ìˆœìœ„": "ë†’ìŒ",
                             "ì‚¬ìœ ": f"ROAS {roas:.0f}% ì ì"})
            elif conv > 0 and roas >= 500:
                recs.append({"ëŒ€ìƒìœ í˜•": "ìƒí’ˆ", "ìº í˜ì¸": campaign, "ì´ë¦„": name,
                             "ROAS(%)": roas, "ê´‘ê³ ë¹„": spend,
                             "ì¡°ì¹˜": "ì˜ˆì‚° í™•ëŒ€ ê²€í† ", "ìš°ì„ ìˆœìœ„": "ì¤‘ê°„",
                             "ì‚¬ìœ ": f"ROAS {roas:.0f}% ê³ íš¨ìœ¨"})

    # â”€â”€ ìº í˜ì¸ ì¡°ì¹˜ â”€â”€
    if not camp_df.empty:
        for _, r in camp_df.iterrows():
            roas = float(r.get("ROAS(%)", 0) or 0)
            spend = int(r.get("ê´‘ê³ ë¹„", 0) or 0)
            name = r.get("ìº í˜ì¸", "")

            if roas < 100:
                recs.append({"ëŒ€ìƒìœ í˜•": "ìº í˜ì¸", "ìº í˜ì¸": name, "ì´ë¦„": name,
                             "ROAS(%)": roas, "ê´‘ê³ ë¹„": spend,
                             "ì¡°ì¹˜": "ìº í˜ì¸ ì˜ˆì‚° ì¶•ì†Œ", "ìš°ì„ ìˆœìœ„": "ë†’ìŒ",
                             "ì‚¬ìœ ": f"ROAS {roas:.0f}% ì ì"})
            elif roas < 200:
                recs.append({"ëŒ€ìƒìœ í˜•": "ìº í˜ì¸", "ìº í˜ì¸": name, "ì´ë¦„": name,
                             "ROAS(%)": roas, "ê´‘ê³ ë¹„": spend,
                             "ì¡°ì¹˜": "í‚¤ì›Œë“œ ì •ë¦¬ í•„ìš”", "ìš°ì„ ìˆœìœ„": "ì¤‘ê°„",
                             "ì‚¬ìœ ": f"ROAS {roas:.0f}% ì €íš¨ìœ¨"})
            elif roas >= 300:
                recs.append({"ëŒ€ìƒìœ í˜•": "ìº í˜ì¸", "ìº í˜ì¸": name, "ì´ë¦„": name,
                             "ROAS(%)": roas, "ê´‘ê³ ë¹„": spend,
                             "ì¡°ì¹˜": "ì˜ˆì‚° í™•ëŒ€ ê²€í† ", "ìš°ì„ ìˆœìœ„": "ì¤‘ê°„",
                             "ì‚¬ìœ ": f"ROAS {roas:.0f}% ê³ íš¨ìœ¨"})

    if not recs:
        return pd.DataFrame()

    recs_df = pd.DataFrame(recs)
    # ìš°ì„ ìˆœìœ„ ì •ë ¬: ë†’ìŒ â†’ ì¤‘ê°„ â†’ ë‚®ìŒ, ê°™ì€ ìˆœìœ„ ë‚´ ê´‘ê³ ë¹„ ë‚´ë¦¼ì°¨ìˆœ
    priority_order = {"ë†’ìŒ": 0, "ì¤‘ê°„": 1, "ë‚®ìŒ": 2}
    recs_df["_sort"] = recs_df["ìš°ì„ ìˆœìœ„"].map(priority_order)
    recs_df = (recs_df.sort_values(["_sort", "ê´‘ê³ ë¹„"], ascending=[True, False])
               .drop(columns="_sort").reset_index(drop=True))
    return recs_df

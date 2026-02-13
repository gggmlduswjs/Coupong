"""
주문/배송 페이지
================
결제완료 → 발주서 → 배송 → 출고/극동 워크플로우.
"""
import io
import json
import logging
import re
from datetime import date, datetime, timedelta

import pandas as pd
import streamlit as st
from sqlalchemy import text as sa_text
from st_aggrid import AgGrid, GridOptionsBuilder

from app.api.coupang_wing_client import CoupangWingError
from app.constants import (
    is_gift_item,
    match_publisher_from_text,
    resolve_distributor,
)
from app.dashboard_utils import (
    create_wing_client,
    engine,
    fmt_krw,
    fmt_money_df,
    query_df,
    query_df_cached,
    render_grid,
    run_sql,
)
logger = logging.getLogger(__name__)

# ── 주문 DB 저장용 UPSERT SQL ──
_UPSERT_ORDER_SQL = """
    INSERT INTO orders
        (account_id, shipment_box_id, order_id, vendor_item_id,
         status, ordered_at, paid_at,
         orderer_name, receiver_name, receiver_addr, receiver_post_code,
         product_id, seller_product_id, seller_product_name, vendor_item_name,
         shipping_count, cancel_count, hold_count_for_cancel,
         sales_price, order_price, discount_price, shipping_price,
         delivery_company_name, invoice_number, shipment_type,
         delivered_date, confirm_date,
         refer, canceled, listing_id, raw_json, updated_at)
    VALUES
        (:account_id, :shipment_box_id, :order_id, :vendor_item_id,
         :status, :ordered_at, :paid_at,
         :orderer_name, :receiver_name, :receiver_addr, :receiver_post_code,
         :product_id, :seller_product_id, :seller_product_name, :vendor_item_name,
         :shipping_count, :cancel_count, :hold_count_for_cancel,
         :sales_price, :order_price, :discount_price, :shipping_price,
         :delivery_company_name, :invoice_number, :shipment_type,
         :delivered_date, :confirm_date,
         :refer, :canceled, :listing_id, :raw_json, :updated_at)
    ON CONFLICT (account_id, shipment_box_id, vendor_item_id) DO UPDATE SET
        status=EXCLUDED.status, ordered_at=EXCLUDED.ordered_at, paid_at=EXCLUDED.paid_at,
        orderer_name=EXCLUDED.orderer_name, receiver_name=EXCLUDED.receiver_name,
        receiver_addr=EXCLUDED.receiver_addr, receiver_post_code=EXCLUDED.receiver_post_code,
        product_id=EXCLUDED.product_id, seller_product_id=EXCLUDED.seller_product_id,
        seller_product_name=EXCLUDED.seller_product_name, vendor_item_name=EXCLUDED.vendor_item_name,
        shipping_count=EXCLUDED.shipping_count, cancel_count=EXCLUDED.cancel_count,
        hold_count_for_cancel=EXCLUDED.hold_count_for_cancel,
        sales_price=EXCLUDED.sales_price, order_price=EXCLUDED.order_price,
        discount_price=EXCLUDED.discount_price, shipping_price=EXCLUDED.shipping_price,
        delivery_company_name=EXCLUDED.delivery_company_name, invoice_number=EXCLUDED.invoice_number,
        shipment_type=EXCLUDED.shipment_type, delivered_date=EXCLUDED.delivered_date,
        confirm_date=EXCLUDED.confirm_date, refer=EXCLUDED.refer, canceled=EXCLUDED.canceled,
        listing_id=EXCLUDED.listing_id, raw_json=EXCLUDED.raw_json, updated_at=EXCLUDED.updated_at
"""


def _parse_dt(val):
    """날짜/시간 문자열 파싱"""
    if not val:
        return None
    return str(val)[:19]


def _extract_price(val):
    """v4 plain int / v5 {units, nanos} 파싱"""
    if val is None:
        return 0
    if isinstance(val, dict):
        return int(val.get("units", 0) or 0)
    return int(val or 0)


def _save_ordersheets_to_db(acct, ordersheets, status):
    """WING API 응답 → orders 테이블 UPSERT (백그라운드 스레드, match_listing 생략)"""
    if not ordersheets:
        return

    def _do_save():
        account_id = int(acct["id"])
        try:
            with engine.connect() as conn:
                for os_data in ordersheets:
                    shipment_box_id = os_data.get("shipmentBoxId")
                    order_id = os_data.get("orderId")
                    if not shipment_box_id or not order_id:
                        continue
                    order_items = os_data.get("orderItems", [])
                    if not order_items:
                        order_items = [os_data]
                    orderer = os_data.get("orderer") or {}
                    receiver = os_data.get("receiver") or {}
                    addr1 = receiver.get("addr1", "") or ""
                    addr2 = receiver.get("addr2", "") or ""
                    receiver_addr = f"{addr1} {addr2}".strip()
                    for item in order_items:
                        v_item_id = item.get("vendorItemId") or os_data.get("vendorItemId")
                        sp_id = item.get("sellerProductId") or os_data.get("sellerProductId")
                        sp_name = item.get("sellerProductName") or os_data.get("sellerProductName", "")
                        params = {
                            "account_id": account_id,
                            "shipment_box_id": int(shipment_box_id),
                            "order_id": int(order_id),
                            "vendor_item_id": int(v_item_id) if v_item_id else 0,
                            "status": status,
                            "ordered_at": _parse_dt(os_data.get("orderedAt")),
                            "paid_at": _parse_dt(os_data.get("paidAt")),
                            "orderer_name": orderer.get("name", ""),
                            "receiver_name": receiver.get("name", ""),
                            "receiver_addr": receiver_addr,
                            "receiver_post_code": receiver.get("postCode", ""),
                            "product_id": int(item.get("productId") or 0) or None,
                            "seller_product_id": int(sp_id) if sp_id else None,
                            "seller_product_name": sp_name,
                            "vendor_item_name": item.get("vendorItemName") or "",
                            "shipping_count": int(item.get("shippingCount", 0) or 0),
                            "cancel_count": int(item.get("cancelCount", 0) or 0),
                            "hold_count_for_cancel": int(item.get("holdCountForCancel", 0) or 0),
                            "sales_price": _extract_price(item.get("salesPrice")),
                            "order_price": _extract_price(item.get("orderPrice")),
                            "discount_price": _extract_price(item.get("discountPrice")),
                            "shipping_price": _extract_price(os_data.get("shippingPrice")),
                            "delivery_company_name": os_data.get("deliveryCompanyName", ""),
                            "invoice_number": os_data.get("invoiceNumber", ""),
                            "shipment_type": os_data.get("shipmentType", ""),
                            "delivered_date": _parse_dt(os_data.get("deliveredDate")),
                            "confirm_date": _parse_dt(item.get("confirmDate")),
                            "refer": os_data.get("refer", ""),
                            "canceled": bool(item.get("canceled", False)),
                            "listing_id": None,
                            "raw_json": json.dumps(os_data, ensure_ascii=False, default=str)[:5000],
                            "updated_at": datetime.utcnow().isoformat(),
                        }
                        try:
                            conn.execute(sa_text(_UPSERT_ORDER_SQL), params)
                        except Exception:
                            pass
                conn.commit()
        except Exception as e:
            logger.warning(f"주문 DB 저장 오류: {e}")

    import threading
    threading.Thread(target=_do_save, daemon=True).start()


def render(selected_account, accounts_df, account_names):
    st.title("주문 관리")

    # ── DB 기반 주문 조회 (단일 쿼리, 즉시 로드) ──

    @st.cache_data(ttl=30)
    def _load_all_orders_from_db():
        """DB에서 최근 30일 전체 주문 1회 조회 — 이후 Python에서 필터"""
        _from = (date.today() - timedelta(days=30)).isoformat()
        return query_df("""
            SELECT a.account_name AS "계정",
                   o.shipment_box_id AS "묶음배송번호",
                   o.order_id AS "주문번호",
                   o.seller_product_name AS "상품명",
                   o.vendor_item_name AS "옵션명",
                   o.shipping_count AS "수량",
                   o.order_price AS "결제금액",
                   to_char(o.ordered_at, 'YYYY-MM-DD') AS "주문일",
                   o.receiver_name AS "수취인",
                   o.status AS "상태",
                   o.delivery_company_name AS "택배사",
                   o.invoice_number AS "운송장번호",
                   to_char(o.delivered_date, 'YYYY-MM-DD') AS "배송완료일",
                   COALESCE(o.canceled, false) AS "취소",
                   o.account_id AS "_account_id",
                   o.vendor_item_id AS "_vendor_item_id",
                   o.seller_product_id AS "_seller_product_id",
                   o.order_price AS "_order_price_raw",
                   to_char(o.ordered_at, 'YYYY-MM-DD HH24:MI:SS') AS "주문일시",
                   o.orderer_name AS "구매자",
                   '' AS "구매자전화번호",
                   '' AS "수취인전화번호",
                   o.receiver_post_code AS "우편번호",
                   o.receiver_addr AS "수취인주소",
                   '' AS "배송메세지",
                   COALESCE(o.shipping_price, 0) AS "배송비",
                   0 AS "도서산간추가배송비",
                   COALESCE(o.refer, '') AS "결제위치",
                   false AS "분리배송가능",
                   '' AS "주문시출고예정일",
                   '' AS "배송비구분",
                   COALESCE(o.sales_price, 0) AS "판매단가",
                   '' AS "최초등록상품옵션명",
                   '' AS "업체상품코드",
                   '' AS "개인통관번호",
                   '' AS "통관용전화번호"
            FROM orders o
            JOIN accounts a ON o.account_id = a.id
            WHERE o.ordered_at >= :date_from
            ORDER BY o.ordered_at DESC
        """, {"date_from": _from})

    def _sync_live_orders():
        """WING API 병렬 호출 → DB 저장 → 페이지 새로고침"""
        from concurrent.futures import ThreadPoolExecutor, as_completed

        _today = date.today()
        _from = (_today - timedelta(days=7)).isoformat()
        _to = _today.isoformat()

        acct_clients = []
        for _, acct in accounts_df.iterrows():
            client = create_wing_client(acct)
            if client:
                acct_clients.append((acct, client))

        def _fetch_one(acct, client, status):
            try:
                return acct, status, client.get_all_ordersheets(_from, _to, status=status)
            except Exception:
                return acct, status, []

        total = 0
        with ThreadPoolExecutor(max_workers=10) as pool:
            futures = []
            for acct, client in acct_clients:
                for status in ["ACCEPT", "INSTRUCT"]:
                    futures.append(pool.submit(_fetch_one, acct, client, status))
            for f in as_completed(futures):
                acct, status, ordersheets = f.result()
                if ordersheets:
                    _save_ordersheets_to_db(acct, ordersheets, status)
                    total += len(ordersheets)
        return total

    def _clear_order_caches():
        """캐시 초기화"""
        _load_all_orders_from_db.clear()
        st.cache_data.clear()

    # ── 상단 컨트롤 ──
    _top_c1, _top_c2 = st.columns(2)
    with _top_c1:
        if st.button("실시간 동기화", key="btn_live_refresh", use_container_width=True):
            with st.spinner("WING API 조회 중..."):
                _synced = _sync_live_orders()
            _clear_order_caches()
            st.success(f"동기화 완료: {_synced}건")
            import time; time.sleep(0.5)
            st.rerun()
    with _top_c2:
        if st.button("전체 동기화 (7일)", key="btn_sync_orders", use_container_width=True):
            try:
                from scripts.sync_orders import OrderSync
                _syncer = OrderSync()
                _sync_accounts = _syncer._get_accounts()
                _sync_bar = st.progress(0, text="주문 동기화 시작...")
                _sync_results = []
                for _si, _sa in enumerate(_sync_accounts):
                    _sync_bar.progress((_si) / len(_sync_accounts), text=f"[{_sa['account_name']}] 동기화 중...")
                    _sr = _syncer.sync_account(
                        _sa,
                        date_from=date.today() - timedelta(days=7),
                        date_to=date.today(),
                    )
                    _sync_results.append(_sr)
                _sync_bar.progress(1.0, text="동기화 완료!")
                _total_upserted = sum(r["upserted"] for r in _sync_results)
                _total_fetched = sum(r["fetched"] for r in _sync_results)
                _clear_order_caches()
                st.success(f"동기화 완료: {len(_sync_accounts)}개 계정, {_total_fetched}건 조회, {_total_upserted}건 저장")
                for _sr in _sync_results:
                    st.caption(f"  [{_sr['account']}] 조회 {_sr['fetched']} / 저장 {_sr['upserted']} / 매칭 {_sr['matched']}")
            except Exception as e:
                st.error(f"동기화 오류: {e}")

    # ── 공통 유틸 ──
    _status_map = {
        "ACCEPT": "결제완료", "INSTRUCT": "상품준비중", "DEPARTURE": "출고완료",
        "DELIVERING": "배송중", "FINAL_DELIVERY": "배송완료", "NONE_TRACKING": "추적불가",
    }

    def _ord_fmt_krw(val):
        val = int(val)
        if abs(val) >= 100_000_000:
            return f"{val / 100_000_000:.1f}억"
        elif abs(val) >= 10_000:
            return f"{val / 10_000:.0f}만"
        else:
            return f"{val:,}"

    _ord_date_to_str = date.today().isoformat()
    _ord_date_from_str = (date.today() - timedelta(days=30)).isoformat()

    # ── DB에서 즉시 로드 (단일 쿼리, 30초 캐시) ──
    _all_orders = _load_all_orders_from_db()

    def _filter_status(df, status):
        if df.empty:
            return pd.DataFrame()
        return df[df["상태"] == status].copy()

    def _kpi_count(df, status):
        sub = _filter_status(df, status)
        if sub.empty:
            return {}
        return sub.groupby("계정")["묶음배송번호"].nunique().to_dict()

    _accept_all = _filter_status(_all_orders, "ACCEPT")
    _instruct_live = _filter_status(_all_orders, "INSTRUCT")
    _kpi_departure = _kpi_count(_all_orders, "DEPARTURE")
    _kpi_delivering = _kpi_count(_all_orders, "DELIVERING")
    _kpi_final = _kpi_count(_all_orders, "FINAL_DELIVERY")
    _instruct_all = _instruct_live[~_instruct_live["취소"]].copy() if not _instruct_live.empty else pd.DataFrame()

    # KPI 계정별 집계
    # 발주서(묶음배송) 단위 집계
    _kpi_accept = _accept_all.groupby("계정")["묶음배송번호"].nunique().to_dict() if not _accept_all.empty else {}
    _kpi_instruct = _instruct_all.groupby("계정")["묶음배송번호"].nunique().to_dict() if not _instruct_all.empty else {}

    # ── 실시간 주문 현황 KPI ──
    _kc1, _kc2, _kc3, _kc4, _kc5 = st.columns(5)

    def _render_kpi(col, label, counts):
        total = sum(counts.values())
        col.metric(label, f"{total:,}건")
        if counts:
            parts = [f"{k}: {v}" for k, v in sorted(counts.items())]
            col.caption(" | ".join(parts))

    _render_kpi(_kc1, "결제완료", _kpi_accept)
    _render_kpi(_kc2, "상품준비중", _kpi_instruct)
    _render_kpi(_kc3, "배송지시", _kpi_departure)
    _render_kpi(_kc4, "배송중", _kpi_delivering)
    _render_kpi(_kc5, "배송완료(30일)", _kpi_final)

    st.divider()

    # 묶음배송 단위 집계 (배송 탭에서 사용)
    if not _instruct_all.empty:
        _inst_by_box = _instruct_all.groupby(["계정", "묶음배송번호", "주문번호", "주문일", "수취인"]).agg(
            상품명=("상품명", lambda x: " / ".join(x.unique())),
            수량=("수량", "sum"),
            결제금액=("_order_price_raw", "sum"),
        ).reset_index()
    else:
        _inst_by_box = pd.DataFrame()

    # ── 4 탭: 워크플로우 순서 ──
    _ord_tab1, _ord_tab2, _ord_tab3, _ord_tab4 = st.tabs(["결제완료", "발주서", "배송", "출고/극동"])

    # ══════════════════════════════════════
    # 탭1: 결제완료 (ACCEPT)
    # ══════════════════════════════════════
    with _ord_tab1:
        st.caption("DB 기반 조회 (실시간 동기화 버튼으로 갱신)")

        # 계정 필터
        _t1_acct = st.selectbox("계정", ["전체"] + account_names, key="tab1_acct")
        _t1_data = _accept_all.copy() if not _accept_all.empty else pd.DataFrame()
        if not _t1_data.empty and _t1_acct != "전체":
            _t1_data = _t1_data[_t1_data["계정"] == _t1_acct]

        _accept_total = len(_t1_data)
        _accept_amount = int(_t1_data["결제금액"].sum()) if not _t1_data.empty else 0
        _accept_by_acct = _t1_data.groupby("계정").size().to_dict() if not _t1_data.empty else {}

        _ak1, _ak2, _ak3 = st.columns(3)
        _ak1.metric("결제완료 주문", f"{_accept_total:,}건")
        _ak2.metric("총 금액", f"₩{_ord_fmt_krw(_accept_amount)}")
        if _accept_by_acct:
            _acct_parts = [f"{k}: {v}" for k, v in sorted(_accept_by_acct.items())]
            _ak3.metric("계정별", " | ".join(_acct_parts))

        st.divider()

        if _t1_data.empty:
            st.info("결제완료(ACCEPT) 상태의 주문이 없습니다.")
        else:
            _accept_display = _t1_data[["계정", "묶음배송번호", "주문번호", "상품명", "옵션명", "수량", "결제금액", "주문일", "수취인"]].copy()
            _accept_display["결제금액"] = _accept_display["결제금액"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "0")

            gb = GridOptionsBuilder.from_dataframe(_accept_display)
            gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=20)
            gb.configure_default_column(resizable=True, sorteable=True, filterable=True)
            gb.configure_column("상품명", width=250)
            gb.configure_column("옵션명", width=200)
            grid_opts = gb.build()
            AgGrid(_accept_display, gridOptions=grid_opts, height=450, theme="streamlit", key="tab1_accept_grid")

            st.divider()

            st.info("ACCEPT 주문을 상품준비중(INSTRUCT)으로 일괄 변경합니다.")
            _ack_unique = _t1_data[["계정", "묶음배송번호"]].drop_duplicates()
            _ack_total_count = len(_ack_unique)

            if st.button(f"상품준비중 처리 ({_ack_total_count}건)", type="primary", key="btn_ack_all_v2"):
                _acct_groups = _t1_data.groupby("_account_id")
                _total_success = 0
                _total_fail = 0

                for _aid, _grp in _acct_groups:
                    _acct_name = _grp.iloc[0]["계정"]
                    _acct_row = accounts_df[accounts_df["id"] == _aid]
                    if _acct_row.empty:
                        st.error(f"[{_acct_name}] 계정 정보를 찾을 수 없습니다.")
                        continue
                    _client = create_wing_client(_acct_row.iloc[0])
                    if not _client:
                        st.error(f"[{_acct_name}] WING API 클라이언트 생성 실패")
                        continue

                    _ack_ids = _grp["묶음배송번호"].unique().tolist()
                    try:
                        _ack_result = _client.acknowledge_ordersheets([int(x) for x in _ack_ids])

                        _success_ids = []
                        _fail_items = []
                        if isinstance(_ack_result, dict) and "data" in _ack_result:
                            _resp_data = _ack_result["data"]
                            _resp_code = _resp_data.get("responseCode")
                            _resp_list = _resp_data.get("responseList", [])

                            for _item in _resp_list:
                                if _item.get("succeed"):
                                    _success_ids.append(_item["shipmentBoxId"])
                                else:
                                    _fail_items.append(_item)

                            if _resp_code == 0:
                                st.success(f"[{_acct_name}] 완료: {len(_success_ids)}건")
                            elif _resp_code == 1:
                                st.warning(f"[{_acct_name}] 부분 성공: {len(_success_ids)}건 성공, {len(_fail_items)}건 실패")
                                for _fi in _fail_items:
                                    st.error(f"  {_fi.get('shipmentBoxId')}: {_fi.get('resultMessage', '')}")
                            elif _resp_code == 99:
                                st.error(f"[{_acct_name}] 전체 실패: {_resp_data.get('responseMessage', '')}")
                            else:
                                _success_ids = [int(x) for x in _ack_ids]
                                st.success(f"[{_acct_name}] 완료: {len(_success_ids)}건")
                        else:
                            _success_ids = [int(x) for x in _ack_ids]
                            st.success(f"[{_acct_name}] 완료: {len(_success_ids)}건")

                        _total_success += len(_success_ids)
                        _total_fail += len(_fail_items)

                    except CoupangWingError as e:
                        st.error(f"[{_acct_name}] API 오류: {e}")
                        _total_fail += len(_ack_ids)

                if _total_success > 0:
                    _clear_order_caches()
                    st.rerun()

            # ── 주문 취소 ──
            with st.expander("주문 취소", expanded=False):
                st.caption("ACCEPT/INSTRUCT 상태의 주문을 취소합니다.")

                _cancel_acct = st.selectbox("취소할 계정", account_names, key="tab1_cancel_acct")
                _cancel_acct_row = None
                if _cancel_acct and not accounts_df.empty:
                    _mask = accounts_df["account_name"] == _cancel_acct
                    if _mask.any():
                        _cancel_acct_row = accounts_df[_mask].iloc[0]

                if _cancel_acct_row is not None:
                    _cancel_account_id = int(_cancel_acct_row["id"])
                    _cancel_client = create_wing_client(_cancel_acct_row)

                    # 공유 데이터에서 계정 필터
                    _cancel_frames = []
                    if not _accept_all.empty:
                        _cancel_frames.append(_accept_all)
                    if not _instruct_live.empty:
                        _cancel_frames.append(_instruct_live)
                    _cancel_all = pd.concat(_cancel_frames, ignore_index=True) if _cancel_frames else pd.DataFrame()
                    _cancelable = pd.DataFrame()
                    if not _cancel_all.empty:
                        _cancel_acct_df = _cancel_all[
                            (_cancel_all["_account_id"] == _cancel_account_id) & (~_cancel_all["취소"])
                        ].copy()
                        if not _cancel_acct_df.empty:
                            _cancelable = _cancel_acct_df.rename(columns={"_vendor_item_id": "옵션ID"})[
                                ["주문번호", "옵션ID", "상품명", "수량", "결제금액", "상태", "주문일"]
                            ].copy()

                    if _cancelable.empty:
                        st.info(f"[{_cancel_acct}] 취소 가능한 주문이 없습니다.")
                    else:
                        _cancelable_display = _cancelable.copy()
                        _cancelable_display["상태"] = _cancelable_display["상태"].map(lambda x: _status_map.get(x, x))
                        st.dataframe(_cancelable_display, width="stretch", hide_index=True)

                        _cancel_reasons = {
                            "SOLD_OUT": "재고 소진",
                            "PRICE_ERROR": "가격 오류",
                            "PRODUCT_ERROR": "상품 정보 오류",
                            "OTHER": "기타 사유",
                        }
                        _sel_reason = st.selectbox("취소 사유", list(_cancel_reasons.keys()),
                                                    format_func=lambda x: _cancel_reasons[x],
                                                    key="tab1_cancel_reason")
                        _cancel_detail = st.text_input("상세 사유", value=_cancel_reasons[_sel_reason], key="tab1_cancel_detail")

                        _confirm_cancel = st.checkbox(
                            f"{len(_cancelable)}건을 정말 취소하시겠습니까? (되돌릴 수 없음)",
                            key="tab1_cancel_confirm",
                        )
                        if _confirm_cancel:
                            if st.button(f"주문 취소 ({len(_cancelable)}건)", type="secondary", key="btn_cancel_ord"):
                                if _cancel_client:
                                    try:
                                        _cancel_groups = _cancelable.groupby("주문번호")
                                        _cancel_count = 0
                                        for _oid, _group in _cancel_groups:
                                            _vids = [int(x) for x in _group["옵션ID"].tolist() if pd.notna(x)]
                                            _cnts = [int(x) for x in _group["수량"].tolist()]
                                            if _vids:
                                                _cancel_client.cancel_order(
                                                    order_id=int(_oid),
                                                    vendor_item_ids=_vids,
                                                    receipt_counts=_cnts,
                                                    cancel_reason_category=_sel_reason,
                                                    cancel_reason=_cancel_detail,
                                                )
                                                _cancel_count += len(_vids)
                                        st.success(f"취소 요청 완료: {_cancel_count}건")
                                        _clear_order_caches()
                                        st.rerun()
                                    except CoupangWingError as e:
                                        st.error(f"API 오류: {e}")
                                else:
                                    st.error("WING API 클라이언트를 생성할 수 없습니다.")

    # ══════════════════════════════════════
    # 탭2: 발주서
    # ══════════════════════════════════════
    with _ord_tab2:
            st.caption("INSTRUCT 주문 기반 거래처별 발주서 생성")

            # 상품준비중(INSTRUCT)만 발주서 대상 (API 실시간)
            _dist_orders = _instruct_all.copy() if not _instruct_all.empty else pd.DataFrame()

            # 사은품/증정품 필터링
            if not _dist_orders.empty:
                _before = len(_dist_orders)
                _dist_orders = _dist_orders[~_dist_orders["옵션명"].apply(lambda x: is_gift_item(str(x)))].copy()
                _gift_cnt = _before - len(_dist_orders)
                if _gift_cnt > 0:
                    st.caption(f"사은품/증정품 {_gift_cnt}건 제외됨")

            if _dist_orders.empty:
                st.info("발주서 대상 주문이 없습니다.")
            else:
                # 출판사 매칭 (거래처 그룹핑용)
                _pub_list = query_df_cached("SELECT name FROM publishers WHERE is_active = true ORDER BY LENGTH(name) DESC")
                _pub_names = _pub_list["name"].tolist() if not _pub_list.empty else []

                def _match_pub(row):
                    result = match_publisher_from_text(str(row.get("옵션명") or ""), _pub_names)
                    if not result:
                        result = match_publisher_from_text(str(row.get("상품명") or ""), _pub_names)
                    return result

                # ISBN 조회: listings.isbn → books (직접 매칭)
                _isbn_lookup = query_df_cached("""
                    SELECT l.coupang_product_id,
                           l.isbn as isbn,
                           b.title as db_title,
                           l.product_name as listing_name
                    FROM listings l
                    LEFT JOIN books b ON l.isbn = b.isbn AND l.isbn IS NOT NULL AND l.isbn != ''
                    WHERE l.coupang_product_id IS NOT NULL
                """)
                _isbn_map = {}
                if not _isbn_lookup.empty:
                    for _, _r in _isbn_lookup.iterrows():
                        _isbn_map[str(_r["coupang_product_id"])] = {
                            "isbn": str(_r["isbn"]) if pd.notna(_r["isbn"]) else "",
                            "title": str(_r["db_title"]) if pd.notna(_r["db_title"]) else "",
                            "listing_name": str(_r["listing_name"]) if pd.notna(_r["listing_name"]) else "",
                        }

                # 도서명/ISBN: 1) books.title 2) listing.product_name 3) 옵션명
                def _resolve_book_info(row):
                    spid = str(row.get("_seller_product_id", ""))
                    info = _isbn_map.get(spid, {})
                    isbn = info.get("isbn", "")
                    title = info.get("title", "")
                    if not title:
                        title = info.get("listing_name", "")
                    if not title:
                        title = str(row.get("옵션명", "")).strip()
                    return pd.Series({"도서명": title, "ISBN": isbn})

                _dist_orders[["도서명", "ISBN"]] = _dist_orders.apply(_resolve_book_info, axis=1)
                _dist_df = _dist_orders

                # ISBN 매칭 현황 표시
                _isbn_found = _dist_df["ISBN"].apply(lambda x: bool(x and str(x).strip())).sum()
                _isbn_total = len(_dist_df)
                _isbn_missing = _isbn_total - _isbn_found
                if _isbn_missing > 0:
                    _missing_names = _dist_df[~_dist_df["ISBN"].apply(lambda x: bool(x and str(x).strip()))]["도서명"].unique()[:5]
                    st.warning(f"ISBN 미매칭: {_isbn_missing}/{_isbn_total}건 — 리스팅 동기화(상품관리) 또는 fill_isbn 실행 필요. 예: {', '.join(_missing_names[:3])}")

                # 발주서 날짜 범위: INSTRUCT 주문의 실제 주문일 범위
                if "주문일" in _dist_df.columns and not _dist_df.empty:
                    _dist_dates = _dist_df["주문일"].dropna()
                    if not _dist_dates.empty:
                        _ord_date_from_str = str(_dist_dates.min())
                        _ord_date_to_str = str(_dist_dates.max())

                _dist_df["출판사"] = _dist_df.apply(_match_pub, axis=1)
                _dist_df["거래처"] = _dist_df["출판사"].apply(resolve_distributor)

                # 거래처별 요약
                _dist_summary = _dist_df.groupby("거래처").agg(
                    건수=("도서명", "count"),
                    수량합계=("수량", "sum"),
                    금액합계=("결제금액", "sum"),
                ).reset_index().sort_values("건수", ascending=False)
                _dist_summary["금액합계"] = _dist_summary["금액합계"].apply(lambda x: f"{int(x):,}")

                st.dataframe(_dist_summary, hide_index=True, width="stretch")

                # Excel 다운로드 (ISBN 기반 그룹핑)
                _dist_df["_group_key"] = _dist_df.apply(
                    lambda r: r["ISBN"] if r.get("ISBN") else r["도서명"], axis=1
                )
                _agg = _dist_df.groupby(["거래처", "출판사", "_group_key"]).agg(
                    도서명=("도서명", "first"),
                    ISBN=("ISBN", "first"),
                    주문수량=("수량", "sum"),
                ).reset_index().drop(columns=["_group_key"])
                _agg = _agg.sort_values(["거래처", "출판사", "도서명"])

                _dist_names_sorted = _dist_summary["거래처"].tolist()

                _xl_buf = io.BytesIO()
                with pd.ExcelWriter(_xl_buf, engine="openpyxl") as writer:
                    from openpyxl.styles import Font as _Font, PatternFill as _PF, Alignment as _AL, Border as _Bdr, Side as _Sd

                    _hf = _PF(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    _sf = _PF(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                    _bdr = _Bdr(left=_Sd(style='thin'), right=_Sd(style='thin'),
                                top=_Sd(style='thin'), bottom=_Sd(style='thin'))

                    # 전체 목록 시트 (옵션명 원본 | 수량)
                    _raw_agg = _dist_df.groupby("옵션명").agg(수량=("수량", "sum")).reset_index().sort_values("옵션명")
                    _raw_agg.to_excel(writer, sheet_name="전체", index=False, startrow=1)
                    _ws0 = writer.sheets["전체"]
                    _ws0.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
                    _ws0.cell(row=1, column=1).value = f"전체 주문 목록 ({_ord_date_from_str} ~ {_ord_date_to_str})"
                    _ws0.cell(row=1, column=1).font = _Font(bold=True, size=13)
                    for ci in range(1, 3):
                        c = _ws0.cell(row=2, column=ci)
                        c.fill = _hf
                        c.font = _Font(bold=True, color="FFFFFF")
                    _ws0.column_dimensions["A"].width = 70
                    _ws0.column_dimensions["B"].width = 8

                    # 요약 시트
                    _agg_summary = _agg.groupby("거래처").agg(
                        품목수=("도서명", "count"), 총수량=("주문수량", "sum")
                    ).reset_index().sort_values("총수량", ascending=False)
                    _agg_summary.to_excel(writer, sheet_name="요약", index=False, startrow=1)
                    _ws1 = writer.sheets["요약"]
                    _ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
                    _ws1.cell(row=1, column=1).value = f"거래처별 요약 ({_ord_date_from_str} ~ {_ord_date_to_str})"
                    _ws1.cell(row=1, column=1).font = _Font(bold=True, size=13)
                    for ci in range(1, 4):
                        c = _ws1.cell(row=2, column=ci)
                        c.fill = _hf
                        c.font = _Font(bold=True, color="FFFFFF")

                    # 거래처별 시트 (ISBN | 도서명 | 출판사 | 주문수량)
                    _dist_order = ["제일", "대성", "일신", "서부", "북전", "동아", "강우사", "대원", "일반"]
                    _all_dists = sorted(_agg["거래처"].unique(),
                                        key=lambda d: _dist_order.index(d) if d in _dist_order else 99)
                    for _dname in _all_dists:
                        _sdf = _agg[_agg["거래처"] == _dname][["ISBN", "도서명", "출판사", "주문수량"]].copy()
                        if _sdf.empty:
                            continue
                        _sdf = _sdf.sort_values(["출판사", "도서명"])
                        _safe = _dname[:31].replace("/", "_").replace("\\", "_")
                        _sdf.to_excel(writer, sheet_name=_safe, index=False, startrow=1)
                        ws = writer.sheets[_safe]
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
                        ws.cell(row=1, column=1).value = f"[{_dname}] 발주서 ({_ord_date_from_str} ~ {_ord_date_to_str})"
                        ws.cell(row=1, column=1).font = _Font(bold=True, size=13)
                        ws.cell(row=1, column=1).alignment = _AL(horizontal="center")
                        for ci in range(1, 5):
                            c = ws.cell(row=2, column=ci)
                            c.fill = _hf
                            c.font = _Font(bold=True, color="FFFFFF")
                            c.border = _bdr
                        for ri in range(3, 3 + len(_sdf)):
                            for ci in range(1, 5):
                                ws.cell(row=ri, column=ci).border = _bdr
                            ws.cell(row=ri, column=4).alignment = _AL(horizontal="center")
                        _sr = 3 + len(_sdf)
                        ws.cell(row=_sr, column=1, value="합계").font = _Font(bold=True)
                        ws.cell(row=_sr, column=1).fill = _sf
                        ws.cell(row=_sr, column=4, value=int(_sdf["주문수량"].sum())).font = _Font(bold=True)
                        ws.cell(row=_sr, column=4).fill = _sf
                        for ci in range(1, 5):
                            ws.cell(row=_sr, column=ci).border = _bdr
                        ws.column_dimensions["A"].width = 16
                        ws.column_dimensions["B"].width = 55
                        ws.column_dimensions["C"].width = 14
                        ws.column_dimensions["D"].width = 10

                _xl_buf.seek(0)

                st.download_button(
                    "발주서 Excel 다운로드 (거래처별 시트)",
                    _xl_buf.getvalue(),
                    file_name=f"주문{_ord_date_to_str[5:7]}{_ord_date_to_str[8:10]}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dist_xlsx_dl",
                    type="primary",
                    use_container_width=True,
                )

                # 도서별 합산 목록
                st.subheader("도서별 주문 합산")
                _dist_filter = st.multiselect(
                    "거래처 필터", _dist_names_sorted,
                    default=_dist_names_sorted, key="dist_filter",
                )
                _filtered_agg = _agg[_agg["거래처"].isin(_dist_filter)] if _dist_filter else _agg
                _show_agg = _filtered_agg[["거래처", "ISBN", "출판사", "도서명", "주문수량"]].copy()

                gb = GridOptionsBuilder.from_dataframe(_show_agg)
                gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=20)
                gb.configure_default_column(resizable=True, sorteable=True, filterable=True)
                gb.configure_column("도서명", width=350)
                gb.configure_column("주문수량", width=80)
                grid_opts = gb.build()
                AgGrid(_show_agg, gridOptions=grid_opts, height=500, theme="streamlit", key="dist_grid")


    # ══════════════════════════════════════
    # 탭3: 배송
    # ══════════════════════════════════════
    with _ord_tab3:
        st.caption("상품준비중 주문 배송 처리: 주문 확인 → 배송리스트 → 송장 업로드")

        # ── Step 워크플로우 표시 ──
        _s1, _s2, _s3 = st.columns(3)
        with _s1:
            st.markdown("**STEP 1** 주문 확인")
        with _s2:
            st.markdown("**STEP 2** 배송리스트")
        with _s3:
            st.markdown("**STEP 3** 송장 업로드")

        st.divider()

        # ── STEP 1: 상품준비중 주문 확인 ──
        st.subheader("STEP 1: 상품준비중 주문 확인")

        _inst_total = len(_inst_by_box)
        _inst_amount = int(_inst_by_box["결제금액"].sum()) if not _inst_by_box.empty else 0

        _ik1, _ik2 = st.columns(2)
        _ik1.metric("상품준비중 주문", f"{_inst_total:,}건")
        _ik2.metric("총 금액", f"₩{_ord_fmt_krw(_inst_amount)}")

        if _inst_by_box.empty:
            st.info("상품준비중(INSTRUCT) 상태의 주문이 없습니다.")
        else:
            _inst_display = _inst_by_box[["계정", "묶음배송번호", "주문번호", "상품명", "수량", "결제금액", "주문일", "수취인"]].copy()
            _inst_display["결제금액"] = _inst_display["결제금액"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "0")

            gb = GridOptionsBuilder.from_dataframe(_inst_display)
            gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=20)
            gb.configure_default_column(resizable=True, sorteable=True, filterable=True)
            gb.configure_column("상품명", width=350)
            grid_opts = gb.build()
            AgGrid(_inst_display, gridOptions=grid_opts, height=400, theme="streamlit", key="tab3_instruct_grid")

        st.divider()

        # ── STEP 2: 배송리스트 다운로드 ──
        st.subheader("STEP 2: 배송리스트 다운로드")

        if _instruct_all.empty:
            st.info("상품준비중 주문이 없습니다.")
        else:
            _dl_orders = _instruct_all.copy()

            # 계정별 건수 표시
            _acct_counts = _dl_orders.groupby("계정").size().reset_index(name="건수")
            st.dataframe(_acct_counts, hide_index=True)

            # 쿠팡 DeliveryList 형식 (40컬럼) 생성
            _dl_rows = []
            for _idx, (_i, _row) in enumerate(_dl_orders.iterrows(), 1):
                _dl_rows.append({
                    "번호": _idx,
                    "묶음배송번호": int(_row["묶음배송번호"]),
                    "주문번호": int(_row["주문번호"]),
                    "택배사": "한진택배",
                    "운송장번호": "",
                    "분리배송 Y/N": "분리배송가능" if _row.get("분리배송가능") else "분리배송불가",
                    "분리배송 출고예정일": "",
                    "주문시 출고예정일": _row.get("주문시출고예정일", ""),
                    "출고일(발송일)": "",
                    "주문일": _row.get("주문일시", _row.get("주문일", "")),
                    "등록상품명": _row.get("상품명", ""),
                    "등록옵션명": _row.get("옵션명", ""),
                    "노출상품명(옵션명)": f"{_row.get('상품명', '')}, {_row.get('옵션명', '')}",
                    "노출상품ID": str(_row.get("_seller_product_id", "")),
                    "옵션ID": str(_row.get("_vendor_item_id", "")),
                    "최초등록등록상품명/옵션명": _row.get("최초등록상품옵션명", ""),
                    "업체상품코드": _row.get("업체상품코드", ""),
                    "바코드": "",
                    "결제액": int(_row.get("결제금액", 0)),
                    "배송비구분": _row.get("배송비구분", ""),
                    "배송비": _row.get("배송비", 0),
                    "도서산간 추가배송비": int(_row.get("도서산간추가배송비", 0)),
                    "구매수(수량)": int(_row.get("수량", 0)),
                    "옵션판매가(판매단가)": int(_row.get("판매단가", 0) or _row.get("결제금액", 0)),
                    "구매자": _row.get("구매자", ""),
                    "구매자전화번호": _row.get("구매자전화번호", ""),
                    "수취인이름": _row.get("수취인", ""),
                    "수취인전화번호": _row.get("수취인전화번호", ""),
                    "우편번호": _row.get("우편번호", ""),
                    "수취인 주소": _row.get("수취인주소", ""),
                    "배송메세지": _row.get("배송메세지", ""),
                    "상품별 추가메시지": "",
                    "주문자 추가메시지": "",
                    "배송완료일": "",
                    "구매확정일자": "",
                    "개인통관번호(PCCC)": _row.get("개인통관번호", ""),
                    "통관용수취인전화번호": _row.get("통관용전화번호", ""),
                    "기타": "",
                    "결제위치": _row.get("결제위치", ""),
                    "배송유형": "판매자 배송",
                })

            _dl_df = pd.DataFrame(_dl_rows)

            # 엑셀 생성 — 묶음배송번호/주문번호를 텍스트로 저장 (지수 표기 방지)
            _dl_buf = io.BytesIO()
            with pd.ExcelWriter(_dl_buf, engine="openpyxl") as writer:
                _dl_df.to_excel(writer, sheet_name="Delivery", index=False)
                ws = writer.sheets["Delivery"]
                from openpyxl.utils import get_column_letter
                for col_name in ["묶음배송번호", "주문번호", "노출상품ID", "옵션ID"]:
                    if col_name in _dl_df.columns:
                        col_idx = _dl_df.columns.get_loc(col_name)  # 0-based
                        col_letter = get_column_letter(col_idx + 1)  # 1-based for Excel
                        for row_idx in range(2, len(_dl_df) + 2):  # 2~N+1 (header=1)
                            cell = ws[f"{col_letter}{row_idx}"]
                            cell.value = str(int(cell.value)) if cell.value is not None else ""
                            cell.number_format = "@"
            _dl_buf.seek(0)

            st.download_button(
                f"통합 배송리스트 다운로드 ({len(_dl_orders)}건)",
                _dl_buf.getvalue(),
                file_name=f"DeliveryList({date.today().isoformat()})_통합.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="tab3_dl_delivery_list",
                type="primary",
                use_container_width=True,
            )
            st.caption("한진택배 프로그램에 업로드 → 송장번호 받은 뒤 → STEP 3에서 등록")

        st.divider()

        # ── STEP 3: 송장 엑셀 업로드 ──
        st.subheader("STEP 3: 송장 엑셀 업로드")
        st.caption("한진택배에서 송장번호를 받은 엑셀을 업로드하면 각 계정별로 자동 등록됩니다.")

        _inv_file = st.file_uploader("송장 엑셀 파일 (운송장번호 포함)", type=["xlsx", "xls"], key="tab3_inv_file_upload")

        if _inv_file is not None:
            try:
                _inv_df = pd.read_excel(_inv_file)

                # 컬럼명 확인
                _need_cols = ["묶음배송번호", "주문번호", "운송장번호"]
                _missing = [c for c in _need_cols if c not in _inv_df.columns]
                if _missing:
                    st.error(f"필수 컬럼 누락: {_missing}")
                else:
                    # 운송장번호가 있는 행만
                    _inv_filled = _inv_df[_inv_df["운송장번호"].notna() & (_inv_df["운송장번호"] != "")].copy()

                    if _inv_filled.empty:
                        st.warning("운송장번호가 입력된 행이 없습니다.")
                    else:
                        st.success(f"송장번호 입력된 주문: {len(_inv_filled)}건")

                        # 옵션ID 컬럼 확인
                        _has_option_id = "옵션ID" in _inv_filled.columns

                        # INSTRUCT 주문과 매칭하여 계정 정보 연결
                        _inv_merged = _inv_filled.copy()
                        if not _instruct_all.empty:
                            _match_cols = _instruct_all[["묶음배송번호", "_account_id", "_vendor_item_id", "주문번호"]].copy()
                            _match_cols["묶음배송번호"] = _match_cols["묶음배송번호"].astype(str)
                            _inv_merged["묶음배송번호"] = _inv_merged["묶음배송번호"].astype(str)
                            _inv_merged["주문번호"] = _inv_merged["주문번호"].astype(str)
                            _match_cols["주문번호"] = _match_cols["주문번호"].astype(str)

                            _inv_merged = _inv_merged.merge(
                                _match_cols.drop_duplicates(subset=["묶음배송번호", "주문번호"]),
                                on=["묶음배송번호", "주문번호"], how="left",
                            )

                        # 계정별로 분리하여 표시
                        if "_account_id" not in _inv_merged.columns:
                            st.error("상품준비중 주문과 매칭할 수 없습니다. 먼저 배송리스트를 다운로드하세요.")
                        else:
                            _matched = _inv_merged[_inv_merged["_account_id"].notna()]
                            _unmatched = _inv_merged[_inv_merged["_account_id"].isna()]

                            if not _unmatched.empty:
                                st.warning(f"매칭 안 된 주문: {len(_unmatched)}건 (이미 발송됐거나 취소된 주문)")

                            if _matched.empty:
                                st.info("등록할 송장이 없습니다.")
                            else:
                                # 계정별 건수
                                _acct_id_map = dict(zip(accounts_df["id"].astype(int), accounts_df["account_name"]))
                                _matched["계정"] = _matched["_account_id"].astype(int).map(_acct_id_map)
                                _acct_summary = _matched.groupby("계정").size().reset_index(name="송장건수")
                                st.dataframe(_acct_summary, hide_index=True)

                                if st.button(f"전체 송장 등록 ({len(_matched)}건)", key="tab3_btn_bulk_invoice", type="primary"):
                                    _total_success = 0
                                    _total_fail = 0

                                    for _aid, _grp in _matched.groupby("_account_id"):
                                        _aid = int(_aid)
                                        _acct_row = accounts_df[accounts_df["id"] == _aid]
                                        if _acct_row.empty:
                                            continue
                                        _acct_row = _acct_row.iloc[0]
                                        _client = create_wing_client(_acct_row)
                                        if not _client:
                                            st.error(f"[{_acct_row['account_name']}] API 클라이언트 생성 실패")
                                            continue

                                        _inv_data = []
                                        for _, _r in _grp.iterrows():
                                            _vid = int(_r["_vendor_item_id"]) if pd.notna(_r.get("_vendor_item_id")) else 0
                                            if not _vid and _has_option_id:
                                                _vid = int(_r["옵션ID"]) if pd.notna(_r.get("옵션ID")) else 0
                                            _inv_data.append({
                                                "shipmentBoxId": int(_r["묶음배송번호"]),
                                                "orderId": int(_r["주문번호"]),
                                                "vendorItemId": _vid,
                                                "deliveryCompanyCode": "HANJIN",
                                                "invoiceNumber": str(_r["운송장번호"]).strip(),
                                                "splitShipping": False,
                                                "preSplitShipped": False,
                                                "estimatedShippingDate": "",
                                            })

                                        try:
                                            _result = _client.upload_invoice(_inv_data)
                                            _s_cnt = 0
                                            _f_cnt = 0
                                            if isinstance(_result, dict) and "data" in _result:
                                                for _ri in _result["data"].get("responseList", []):
                                                    if _ri.get("succeed"):
                                                        _s_cnt += 1
                                                    else:
                                                        _f_cnt += 1
                                                        st.error(f"  [{_acct_row['account_name']}] {_ri.get('shipmentBoxId')}: {_ri.get('resultMessage', '')}")
                                            else:
                                                _s_cnt = len(_inv_data)
                                            _total_success += _s_cnt
                                            _total_fail += _f_cnt
                                            st.info(f"[{_acct_row['account_name']}] 성공 {_s_cnt}건" + (f", 실패 {_f_cnt}건" if _f_cnt else ""))
                                        except Exception as e:
                                            _total_fail += len(_inv_data)
                                            st.error(f"[{_acct_row['account_name']}] API 오류: {e}")

                                    if _total_success > 0:
                                        st.success(f"송장 등록 완료: 총 {_total_success}건 성공" + (f", {_total_fail}건 실패" if _total_fail else ""))
                                        _clear_order_caches()
                                        st.rerun()
                                    elif _total_fail > 0:
                                        st.error(f"전체 실패: {_total_fail}건")

            except Exception as e:
                st.error(f"엑셀 파일 읽기 오류: {e}")


    # ══════════════════════════════════════
    # 탭4: 출고/극동
    # ══════════════════════════════════════
    with _ord_tab4:
            st.caption("WING API 실시간 출고 주문 → 극동 프로그램용 엑셀 다운로드")

            _gk_col1, _gk_col2, _gk_col3 = st.columns([2, 2, 1])
            with _gk_col1:
                _gk_date_from = st.date_input("시작일", value=date.today() - timedelta(days=1), key="tab4_gk_date_from")
            with _gk_col2:
                _gk_date_to = st.date_input("종료일", value=date.today(), key="tab4_gk_date_to")
            with _gk_col3:
                _gk_status = st.selectbox("상태", ["INSTRUCT", "DEPARTURE", "DELIVERING", "FINAL_DELIVERY"], key="tab4_gk_status")

            # WING API 실시간 조회 (INSTRUCT는 상단 캐시 재사용)
            _gk_from_str = _gk_date_from.isoformat()
            _gk_to_str = _gk_date_to.isoformat()

            _gk_api_rows = []
            if _gk_status == "INSTRUCT" and not _instruct_all.empty:
                # 상단에서 이미 조회한 INSTRUCT 데이터 재사용
                for _, _row in _instruct_all.iterrows():
                    _gk_api_rows.append({
                        "옵션명": _row.get("옵션명", ""),
                        "상품명": _row.get("상품명", ""),
                        "수량": int(_row.get("수량", 0)),
                        "결제금액": int(_row.get("결제금액", 0)),
                        "_seller_product_id": _row.get("_seller_product_id", ""),
                        "계정": _row.get("계정", ""),
                    })
            else:
                with st.spinner(f"{_gk_status} 주문 조회 중..."):
                    for _, _gk_acct in accounts_df.iterrows():
                        _gk_client = create_wing_client(_gk_acct)
                        if not _gk_client:
                            continue
                        try:
                            _gk_result = _gk_client.get_all_ordersheets(_gk_from_str, _gk_to_str, status=_gk_status)
                            for _gk_os in _gk_result:
                                _gk_items = _gk_os.get("orderItems", [])
                                if not _gk_items:
                                    _gk_items = [_gk_os]
                                for _gk_item in _gk_items:
                                    _gk_api_rows.append({
                                        "옵션명": _gk_item.get("vendorItemName", ""),
                                        "상품명": _gk_item.get("sellerProductName") or _gk_os.get("sellerProductName", ""),
                                        "수량": int(_gk_item.get("shippingCount", 0) or 0),
                                        "결제금액": int(_gk_item.get("orderPrice", 0) or 0),
                                        "_seller_product_id": _gk_item.get("sellerProductId") or _gk_os.get("sellerProductId", ""),
                                        "계정": _gk_acct["account_name"],
                                    })
                        except Exception:
                            continue

            # API 결과 → DataFrame + DB에서 ISBN/도서명/출판사 매칭
            _gk_orders = pd.DataFrame(_gk_api_rows) if _gk_api_rows else pd.DataFrame()

            if not _gk_orders.empty:
                # sellerProductId → listings.isbn → books 매칭
                _gk_isbn_lookup = query_df_cached("""
                    SELECT l.coupang_product_id,
                           l.isbn as "ISBN",
                           b.title as "DB도서명",
                           l.product_name as 리스팅도서명,
                           b.list_price as 정가,
                           COALESCE(b.author, '') as 저자,
                           b.year as 출판년도,
                           pub.name as 출판사,
                           pub.supply_rate as 공급률
                    FROM listings l
                    LEFT JOIN books b ON l.isbn = b.isbn AND l.isbn IS NOT NULL AND l.isbn != ''
                    LEFT JOIN publishers pub ON b.publisher_id = pub.id
                    WHERE l.coupang_product_id IS NOT NULL
                """)
                _gk_map = {}
                if not _gk_isbn_lookup.empty:
                    for _, _r in _gk_isbn_lookup.iterrows():
                        _gk_map[str(_r["coupang_product_id"])] = {
                            "ISBN": str(_r["ISBN"]) if pd.notna(_r["ISBN"]) else "",
                            "DB도서명": str(_r["DB도서명"]) if pd.notna(_r["DB도서명"]) else "",
                            "리스팅도서명": str(_r["리스팅도서명"]) if pd.notna(_r["리스팅도서명"]) else "",
                            "정가": _r["정가"] if pd.notna(_r["정가"]) else 0,
                            "저자": str(_r["저자"]) if pd.notna(_r["저자"]) else "",
                            "출판년도": _r["출판년도"] if pd.notna(_r["출판년도"]) else None,
                            "출판사": str(_r["출판사"]) if pd.notna(_r["출판사"]) else "",
                            "공급률": _r["공급률"] if pd.notna(_r["공급률"]) else None,
                        }

                def _gk_enrich(row):
                    info = _gk_map.get(str(row.get("_seller_product_id", "")), {})
                    return pd.Series({
                        "ISBN": info.get("ISBN", ""),
                        "DB도서명": info.get("DB도서명", ""),
                        "리스팅도서명": info.get("리스팅도서명", ""),
                        "정가": info.get("정가", 0),
                        "저자": info.get("저자", ""),
                        "출판년도": info.get("출판년도", None),
                        "출판사": info.get("출판사", ""),
                        "공급률": info.get("공급률", None),
                    })
                _gk_extra = _gk_orders.apply(_gk_enrich, axis=1)
                _gk_orders = pd.concat([_gk_orders, _gk_extra], axis=1)

            if _gk_orders.empty:
                st.info(f"{_gk_from_str} ~ {_gk_to_str} 에 {_gk_status} 주문이 없습니다. 주문 동기화가 필요할 수 있습니다.")
            else:
                # 사은품 필터링
                _gk_before = len(_gk_orders)
                _gk_orders = _gk_orders[~_gk_orders["옵션명"].apply(lambda x: is_gift_item(str(x)))].copy()
                _gk_gift_cnt = _gk_before - len(_gk_orders)
                if _gk_gift_cnt > 0:
                    st.caption(f"사은품/증정품 {_gk_gift_cnt}건 제외됨")

                if _gk_orders.empty:
                    st.info("사은품 제외 후 주문이 없습니다.")
                else:
                    # 도서명 정리: 1) books.title 2) listing.product_name 3) 옵션명
                    def _resolve_gk_title(r):
                        if pd.notna(r.get("DB도서명")) and r["DB도서명"]:
                            return str(r["DB도서명"]).strip()
                        if pd.notna(r.get("리스팅도서명")) and r["리스팅도서명"]:
                            return str(r["리스팅도서명"]).strip()
                        return str(r["옵션명"]).strip()
                    _gk_orders["도서명"] = _gk_orders.apply(_resolve_gk_title, axis=1)
                    _gk_orders["ISBN_clean"] = _gk_orders["ISBN"].apply(lambda x: str(x).strip() if pd.notna(x) and x else "")

                    # ISBN 기반 그룹핑
                    _gk_orders["_key"] = _gk_orders.apply(lambda r: r["ISBN_clean"] if r["ISBN_clean"] else r["도서명"], axis=1)
                    _gk_agg = _gk_orders.groupby("_key").agg(
                        상품바코드=("ISBN_clean", "first"),
                        상품명=("도서명", "first"),
                        정가=("정가", "first"),
                        수량=("수량", "sum"),
                        공급률=("공급률", "first"),
                        출판사=("출판사", "first"),
                        저자=("저자", "first"),
                        출판년도=("출판년도", "first"),
                    ).reset_index(drop=True)

                    # KPI
                    _gk_total_amount = int(_gk_agg.apply(
                        lambda r: (r["정가"] * r["공급률"] * r["수량"]) if pd.notna(r["공급률"]) and r["공급률"] and pd.notna(r["정가"]) else 0,
                        axis=1
                    ).sum()) if not _gk_agg.empty else 0
                    _gk_k1, _gk_k2, _gk_k3 = st.columns(3)
                    _gk_k1.metric("출고 품목", f"{len(_gk_agg)}종")
                    _gk_k2.metric("출고 수량", f"{int(_gk_agg['수량'].sum())}권")
                    _gk_k3.metric("총 금액", f"₩{_ord_fmt_krw(_gk_total_amount)}")

                    # 테이블 표시
                    _gk_show = _gk_agg[["상품바코드", "상품명", "수량", "출판사"]].copy()
                    st.dataframe(_gk_show, hide_index=True, width="stretch")

                    # 극동 형식 엑셀 생성
                    _gk_result = pd.DataFrame()
                    _gk_result["NO."] = range(1, len(_gk_agg) + 1)
                    _gk_result["상품바코드"] = _gk_agg["상품바코드"].values
                    _gk_result["상품명"] = _gk_agg["상품명"].values
                    _gk_result["#"] = ""
                    _gk_result["정 가"] = _gk_agg["정가"].apply(lambda x: int(x) if pd.notna(x) else 0).values
                    _gk_result["수 량"] = _gk_agg["수량"].values
                    _gk_result["%"] = _gk_agg["공급률"].apply(lambda x: f"{x*100:.0f}" if pd.notna(x) and x else "").values
                    _gk_result["단 가"] = _gk_agg.apply(
                        lambda r: int(r["정가"] * r["공급률"]) if pd.notna(r["공급률"]) and r["공급률"] and pd.notna(r["정가"]) else (int(r["정가"]) if pd.notna(r["정가"]) else 0),
                        axis=1
                    ).values
                    _gk_result["금 액"] = (_gk_result["단 가"] * _gk_result["수 량"]).values
                    _gk_result[""] = ""
                    _gk_result["출판사"] = _gk_agg["출판사"].apply(lambda x: str(x) if pd.notna(x) else "").values
                    _gk_result["저자"] = _gk_agg["저자"].apply(lambda x: str(x) if pd.notna(x) else "").values
                    _gk_result["출판년도"] = _gk_agg["출판년도"].apply(lambda x: str(int(x)) if pd.notna(x) else "").values

                    _gk_buf = io.BytesIO()
                    with pd.ExcelWriter(_gk_buf, engine="openpyxl") as writer:
                        _gk_result.to_excel(writer, sheet_name="극동", index=False)
                    _gk_buf.seek(0)

                    st.download_button(
                        f"극동 엑셀 다운로드 ({len(_gk_agg)}종 / {int(_gk_agg['수량'].sum())}권)",
                        _gk_buf.getvalue(),
                        file_name=f"극동_{_gk_date_from.strftime('%m%d')}_{_gk_date_to.strftime('%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="gk_xlsx_dl",
                        type="primary",
                        use_container_width=True,
                    )

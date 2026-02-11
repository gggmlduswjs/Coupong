"""
극동 프로그램용 엑셀 내보내기
==============================
당일 출고 완료(DEPARTURE) 주문을 극동 형식 엑셀로 생성

극동 컬럼: NO. | 상품바코드 | 상품명 | # | 정가 | 수량 | % | 단가 | 금액 | | 출판사 | 저자 | 출판년도

사용법:
    python scripts/export_geukdong.py
    python scripts/export_geukdong.py --date 2026-02-11
    python scripts/export_geukdong.py --status INSTRUCT
"""
import sys
import argparse
from pathlib import Path
from datetime import date, timedelta

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd
from sqlalchemy import text
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database import engine
from app.constants import is_gift_item


def query_df(sql: str, params: dict = None) -> pd.DataFrame:
    if params:
        return pd.read_sql(text(sql), engine, params=params)
    return pd.read_sql(sql, engine)


def export_geukdong(target_date: str = None, status: str = "DEPARTURE",
                     account_name: str = None, output: str = None):
    """극동 형식 엑셀 생성"""
    if not target_date:
        target_date = date.today().isoformat()

    # 최신 주문 동기화
    print("주문 데이터 동기화 중...")
    from scripts.sync_orders import OrderSync
    syncer = OrderSync()
    syncer.sync_all(days=7, account_name=account_name)
    print("동기화 완료.\n")

    # WHERE 절
    acct_where = ""
    params = {}
    if account_name:
        acct_where = "AND o.account_id = (SELECT id FROM accounts WHERE account_name = :acct_name)"
        params["acct_name"] = account_name

    # 주문 + 도서 정보 조인
    orders = query_df(f"""
        SELECT
            o.vendor_item_name as 옵션명,
            o.seller_product_name as 상품명,
            o.shipping_count as 수량,
            COALESCE(l.isbn, b.isbn) as ISBN,
            b.title as DB도서명,
            b.list_price as 정가,
            b.year as 출판년도,
            pub.name as 출판사,
            pub.supply_rate as 공급률
        FROM orders o
        LEFT JOIN listings l ON o.listing_id = l.id
        LEFT JOIN products p ON l.product_id = p.id
        LEFT JOIN books b ON p.book_id = b.id
        LEFT JOIN publishers pub ON b.publisher_id = pub.id
        WHERE o.status = :status
        AND DATE(o.ordered_at) = :target_date
        {acct_where}
        ORDER BY pub.name, o.vendor_item_name
    """, {**params, "status": status, "target_date": target_date})

    if orders.empty:
        print(f"해당 날짜({target_date})에 {status} 주문이 없습니다.")
        return

    # 사은품 필터링
    before = len(orders)
    orders = orders[~orders["옵션명"].apply(lambda x: is_gift_item(str(x)))].copy()
    filtered = before - len(orders)
    if filtered > 0:
        print(f"사은품/증정품 {filtered}건 제외됨")

    if orders.empty:
        print("사은품 제외 후 주문이 없습니다.")
        return

    # 도서명 정리: DB도서명 우선
    orders["도서명"] = orders.apply(
        lambda r: str(r["DB도서명"]).strip() if pd.notna(r.get("DB도서명")) and r["DB도서명"] else str(r["옵션명"]).strip(),
        axis=1
    )
    orders["ISBN_clean"] = orders["ISBN"].apply(lambda x: str(x).strip() if pd.notna(x) and x else "")

    # ISBN 기반 그룹핑 (같은 도서 수량 합산)
    orders["_key"] = orders.apply(lambda r: r["ISBN_clean"] if r["ISBN_clean"] else r["도서명"], axis=1)
    agg = orders.groupby("_key").agg(
        상품바코드=("ISBN_clean", "first"),
        상품명=("도서명", "first"),
        정가=("정가", "first"),
        수량=("수량", "sum"),
        공급률=("공급률", "first"),
        출판사=("출판사", "first"),
        저자=("저자", "first"),
        출판년도=("출판년도", "first"),
    ).reset_index(drop=True)

    # 계산 컬럼
    agg["#"] = ""  # 극동 형식의 # 컬럼 (빈값)
    agg["정가"] = agg["정가"].apply(lambda x: int(x) if pd.notna(x) else 0)
    agg["%"] = agg["공급률"].apply(lambda x: f"{x*100:.0f}" if pd.notna(x) and x else "")
    agg["단가"] = agg.apply(
        lambda r: int(r["정가"] * r["공급률"]) if pd.notna(r["공급률"]) and r["공급률"] else r["정가"],
        axis=1
    )
    agg["금액"] = agg["단가"] * agg["수량"]
    agg[""] = ""  # 빈 컬럼 (극동 형식)

    # 극동 형식 컬럼 순서
    result = pd.DataFrame()
    result["NO."] = range(1, len(agg) + 1)
    result["상품바코드"] = agg["상품바코드"].values
    result["상품명"] = agg["상품명"].values
    result["#"] = agg["#"].values
    result["정 가"] = agg["정가"].values
    result["수 량"] = agg["수량"].values
    result["%"] = agg["%"].values
    result["단 가"] = agg["단가"].values
    result["금 액"] = agg["금액"].values
    result[""] = agg[""].values
    result["출판사"] = agg["출판사"].apply(lambda x: str(x) if pd.notna(x) else "").values
    result["저자"] = agg["저자"].apply(lambda x: str(x) if pd.notna(x) else "").values
    result["출판년도"] = agg["출판년도"].apply(lambda x: str(int(x)) if pd.notna(x) else "").values

    # 출력 파일명
    if not output:
        output = f"극동_{target_date[5:7]}{target_date[8:10]}.xlsx"

    # 엑셀 생성
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="극동", index=False, startrow=1)
        ws = writer.sheets["극동"]

        # 타이틀
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
        ws.cell(row=1, column=1).value = f"극동 출고 목록 ({target_date})"
        ws.cell(row=1, column=1).font = Font(bold=True, size=13)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # 헤더 스타일
        hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
        for ci in range(1, 14):
            c = ws.cell(row=2, column=ci)
            c.fill = hf
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.alignment = Alignment(horizontal="center")
            c.border = bdr

        # 데이터 테두리
        for ri in range(3, 3 + len(result)):
            for ci in range(1, 14):
                ws.cell(row=ri, column=ci).border = bdr

        # 합계 행
        sr = 3 + len(result)
        sf = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        ws.cell(row=sr, column=1, value="합계").font = Font(bold=True)
        ws.cell(row=sr, column=1).fill = sf
        ws.cell(row=sr, column=6, value=int(result["수 량"].sum())).font = Font(bold=True)
        ws.cell(row=sr, column=6).fill = sf
        ws.cell(row=sr, column=9, value=int(result["금 액"].sum())).font = Font(bold=True)
        ws.cell(row=sr, column=9).fill = sf
        for ci in range(1, 14):
            ws.cell(row=sr, column=ci).border = bdr

        # 컬럼 너비
        ws.column_dimensions["A"].width = 5   # NO.
        ws.column_dimensions["B"].width = 16  # 상품바코드
        ws.column_dimensions["C"].width = 45  # 상품명
        ws.column_dimensions["D"].width = 4   # #
        ws.column_dimensions["E"].width = 10  # 정가
        ws.column_dimensions["F"].width = 6   # 수량
        ws.column_dimensions["G"].width = 5   # %
        ws.column_dimensions["H"].width = 10  # 단가
        ws.column_dimensions["I"].width = 12  # 금액
        ws.column_dimensions["J"].width = 2   # 빈칸
        ws.column_dimensions["K"].width = 14  # 출판사
        ws.column_dimensions["L"].width = 14  # 저자
        ws.column_dimensions["M"].width = 8   # 출판년도

    total_qty = int(result["수 량"].sum())
    total_amt = int(result["금 액"].sum())
    print(f"극동 엑셀 생성 완료: {output}")
    print(f"  날짜: {target_date} | 상태: {status}")
    print(f"  총 품목: {len(result)}종 / {total_qty}권")
    print(f"  총 금액: {total_amt:,}원")


def main():
    parser = argparse.ArgumentParser(description="극동 프로그램용 엑셀 내보내기")
    parser.add_argument("--date", type=str, default=None, help="조회 날짜 (YYYY-MM-DD, 기본: 오늘)")
    parser.add_argument("--status", type=str, default="DEPARTURE", help="주문 상태 (기본: DEPARTURE)")
    parser.add_argument("--account", type=str, default=None, help="특정 계정만")
    parser.add_argument("--output", type=str, default=None, help="출력 파일명")
    args = parser.parse_args()

    export_geukdong(
        target_date=args.date,
        status=args.status,
        account_name=args.account,
        output=args.output,
    )


if __name__ == "__main__":
    main()

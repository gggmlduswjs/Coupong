#!/usr/bin/env python3
"""
쿠팡 WING 상품정보 수정요청 엑셀 생성기

_filled.xlsx에서 판매중 + 바코드(ISBN) 있는 행 →
알라딘 API로 메타데이터(저자, 출판사, 시리즈 등) 조회 →
원본 템플릿 형식 그대로 (231컬럼) 수정용 .xlsx 생성

수정 컬럼:
  col5(제조사), col6(브랜드), col7(검색어), col8(대량수정=Y),
  검색옵션 값(코드 기반 자동 매칭), col231(바코드)

사용법:
  python scripts/generate_wing_update_csv.py              # 전체
  python scripts/generate_wing_update_csv.py --account 007-book  # 특정 계정
  python scripts/generate_wing_update_csv.py --skip-api    # API 스킵 (캐시만)
  python scripts/generate_wing_update_csv.py --dry-run     # 테스트
"""

import os
import re
import sys
import io
import json
import time
import argparse
from pathlib import Path

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

# Windows 인코딩
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", line_buffering=True)

import pandas as pd
import requests
from openpyxl import load_workbook, Workbook

# ── 설정 ──
FILLED_FILES = {
    "007-book": "C:/Users/MSI/Desktop/Coupong/007-book_filled.xlsx",
    "007-bm": "C:/Users/MSI/Desktop/Coupong/007-bm_filled.xlsx",
    "007-ez": "C:/Users/MSI/Desktop/Coupong/007-ez_filled.xlsx",
    "002-bm": "C:/Users/MSI/Desktop/Coupong/002-bm_filled.xlsx",
    "big6ceo": "C:/Users/MSI/Desktop/Coupong/big6ceo_filled.xlsx",
}
SHEET = "Template"
TOTAL_COLS = 231
OUTPUT_DIR = Path(project_root) / "output"
CACHE_FILE = OUTPUT_DIR / "aladin_cache.json"
SEARCH_CACHE_FILE = OUTPUT_DIR / "aladin_search_cache.json"
TTB_KEYS = ["ttbsjrnf57491614001", "ttbsjrnf57490005001", "ttbsjrnf57490005003"]
ISBN_PATTERN = re.compile(r"^97[89]\d{10}$")
API_INTERVAL = 0.2
SET_KEYWORDS = ["세트", "전2권", "전3권", "전4권", "전5권", "전6권", "전7권", "전8권"]

# 검색옵션 코드 → 매핑 필드
SEARCH_OPTION_MAP = {
    "728": "publisher",     # 제조사명
    "7937": "author",       # 저자
    "7939": "isbn",         # ISBN (검색옵션에도 있을 수 있음)
    "10524": "series",      # 시리즈명
    "11163": "year",        # 사용연도
    "11165": "semester",    # 학기구분 (skip)
    "11927": "language",    # 발행언어
    "12008": "series",      # 시리즈명 (대체코드)
}

# 노이즈 패턴 (검색어 생성용)
NOISE_PATTERNS = [
    (r"\+\s*사은품.*", ""), (r"\+\s*선물.*", ""), (r"\+\s*미니수첩.*", ""),
    (r"사은품\s*증정\)?", ""), (r"사은품", ""), (r"증정", ""),
    (r"!사은품!", ""), (r"!전\d+권!", ""),
    (r"\*+[^*]*\*+", ""), (r"\*+", ""), (r"\#\w+", ""), (r"\|+", ""),
    (r"!+", ""),
    (r"슝슝오늘출발\!*", ""), (r"오늘출발", ""), (r"당일발송", ""),
    (r"평일\d+시\w+", ""), (r"비닐포장", ""), (r"상세설명참조", ""),
    (r"^\s*\(선물\)\s*", ""), (r"^\s*\(사은품\)\s*", ""), (r"^\s*\(사은 증정\)\s*", ""),
    (r"^\s*\(선물 증정\)\s*", ""), (r"^\s*\(당일발송\)\s*", ""),
    (r"\s*:\s*슝슝.*$", ""), (r"\s*:\s*내신.*$", ""),
    (r"\s*:\s*오늘출발.*$", ""), (r"\s*:\s*평일.*$", ""),
    (r"//+.*$", ""),
]

STOP_WORDS = {
    "세트", "set", "권", "원", "년", "판", "개정", "개정판", "최신", "신판",
    "전", "후", "상", "하", "중", "편", "부", "the", "and", "of", "for",
    "a", "an", "in", "to", "is", "기출", "문제집", "교재",
}

TITLE_PLUS_PATTERNS = [
    "개념+유형", "기본+응용", "유형+내신", "기본+실력", "Q+Q",
    "개념+연산", "유형+실력", "교과서+익힘",
]


# ════════════════════════════════════════
# ISBN 정규화
# ════════════════════════════════════════

def normalize_isbn(val) -> str | None:
    """바코드 값을 ISBN-13 문자열로 정규화"""
    if val is None:
        return None
    try:
        s = str(int(val)) if isinstance(val, (float, int)) else str(val)
        s = re.sub(r"\D", "", s)
        if ISBN_PATTERN.match(s):
            return s
    except (ValueError, OverflowError):
        pass
    return None


# ════════════════════════════════════════
# 알라딘 API
# ════════════════════════════════════════

def load_cache() -> dict:
    if CACHE_FILE.exists():
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(cache: dict):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


class RateLimitError(Exception):
    pass


_current_key_idx = 0


def fetch_aladin_metadata(isbn: str) -> dict | None:
    """알라딘 ItemLookUp API (멀티키, 한도초과 자동 전환)"""
    global _current_key_idx

    tried = 0
    while tried < len(TTB_KEYS):
        key = TTB_KEYS[_current_key_idx]
        try:
            r = requests.get(
                "http://www.aladin.co.kr/ttb/api/ItemLookUp.aspx",
                params={
                    "ttbkey": key, "itemIdType": "ISBN", "ItemId": isbn,
                    "output": "js", "Version": "20131101",
                },
                timeout=15,
            )
            data = r.json()

            if data.get("errorCode") == 10:
                print(f"  키 {_current_key_idx + 1}/{len(TTB_KEYS)} 한도 초과, 다음 키로 전환")
                _current_key_idx = (_current_key_idx + 1) % len(TTB_KEYS)
                tried += 1
                continue

            items = data.get("item", [])
            if not items:
                return None

            item = items[0]
            series_name = ""
            si = item.get("seriesInfo")
            if isinstance(si, dict):
                series_name = si.get("seriesName", "")

            return {
                "author": item.get("author", ""),
                "publisher": item.get("publisher", ""),
                "pubDate": item.get("pubDate", ""),
                "seriesName": series_name,
                "categoryName": item.get("categoryName", ""),
            }
        except Exception:
            return None

    raise RateLimitError("모든 API 키 한도 초과")


def batch_fetch_aladin(isbns: list[str], cache: dict, skip_api: bool = False) -> dict:
    to_fetch = [isbn for isbn in isbns if isbn not in cache]
    cached = len(isbns) - len(to_fetch)

    print(f"\n알라딘 API 조회: 전체 {len(isbns)}건, 캐시 {cached}건, 조회필요 {len(to_fetch)}건")

    if skip_api:
        print("  --skip-api: API 호출 건너뜀")
        return cache

    if not to_fetch:
        print("  전부 캐시 히트!")
        return cache

    ok, fail = 0, 0
    for i, isbn in enumerate(to_fetch, 1):
        try:
            meta = fetch_aladin_metadata(isbn)
        except RateLimitError:
            print(f"\n  모든 API 키 한도 초과! {i-1}/{len(to_fetch)}건에서 중단")
            print(f"  내일 재실행하면 캐시된 {ok}건은 스킵됩니다")
            break

        if meta:
            cache[isbn] = meta
            ok += 1
        else:
            cache[isbn] = None
            fail += 1

        if i % 100 == 0 or i == len(to_fetch):
            print(f"  [{i}/{len(to_fetch)}] 성공={ok} 실패={fail}")

        if i % 100 == 0:
            save_cache(cache)

        time.sleep(API_INTERVAL)

    save_cache(cache)
    print(f"  API 조회 완료: 성공 {ok}, 실패 {fail}")
    return cache


# ════════════════════════════════════════
# 세트 상품 구성품 ISBN 검색
# ════════════════════════════════════════

def load_search_cache() -> dict:
    if SEARCH_CACHE_FILE.exists():
        with open(SEARCH_CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_search_cache(cache: dict):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(SEARCH_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


KYOBO_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
}


def search_kyobo(keyword: str) -> str | None:
    """교보문고 검색 → ISBN 추출 (검색 페이지 data-bid)"""
    from urllib.parse import quote
    try:
        r = requests.get(
            f"https://search.kyobobook.co.kr/search?keyword={quote(keyword)}&target=total",
            headers=KYOBO_HEADERS, timeout=15,
        )
        r.raise_for_status()
        m = re.search(r'data-bid="(\d{13})"', r.text)
        if m and ISBN_PATTERN.match(m.group(1)):
            return m.group(1)
    except Exception:
        pass
    return None


def search_aladin_by_title(query: str) -> str | None:
    """알라딘 키워드 검색 → 첫 번째 결과 ISBN13 반환"""
    global _current_key_idx
    tried = 0
    while tried < len(TTB_KEYS):
        key = TTB_KEYS[_current_key_idx]
        try:
            r = requests.get(
                "http://www.aladin.co.kr/ttb/api/ItemSearch.aspx",
                params={
                    "ttbkey": key, "Query": query, "QueryType": "Title",
                    "MaxResults": 1, "output": "js", "Version": "20131101",
                },
                timeout=15,
            )
            data = r.json()
            if data.get("errorCode") == 10:
                _current_key_idx = (_current_key_idx + 1) % len(TTB_KEYS)
                tried += 1
                continue
            items = data.get("item", [])
            if items:
                raw = items[0].get("isbn13", "")
                # ISBN 정규화 (float→int→str 변환)
                return normalize_isbn(raw)
            return None
        except Exception:
            return None
    raise RateLimitError("모든 API 키 한도 초과")


def is_set_product(name: str) -> bool:
    return any(k in name for k in SET_KEYWORDS)


def clean_search_query(q: str) -> str:
    """알라딘 검색 전 쿼리 정리 — 노이즈 제거 + 간결화"""
    s = q.strip()
    # 대괄호 태그 제거: [선물], [연합도서], [매스티안], [전3권] 등
    s = re.sub(r"\[[^\]]*\]", "", s)
    # 꺾쇠 태그 제거: <2024 최신판>
    s = re.sub(r"<[^>]*>", "", s)
    # 연도 태그 제거 (검색에 방해)
    s = re.sub(r"\(?\d{4}년?\)?", "", s)
    # 교육과정/부제목 노이즈
    s = re.sub(r"\(?\d{2,4}\s*개정\s*교육과정?\)?", "", s)
    s = re.sub(r"\d+제\b", "", s)  # 462제
    s = re.sub(r"상세설명참조", "", s)
    s = re.sub(r"최신판", "", s)
    # 선물+, (선물) 접두사
    s = re.sub(r"^\s*\(?\s*선물\s*\)?\s*\+?\s*", "", s)
    s = re.sub(r"사은품\s*증정", "", s)
    s = re.sub(r"증정", "", s)
    # 영어 교재 노이즈
    s = re.sub(r"\b(?:SB|WB|StudentBook|WorkBook|Student\s*Book|Work\s*Book)\b", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\b\d+(?:st|nd|rd|th)\s*edition\b", "", s, flags=re.IGNORECASE)
    # 불필요한 수식어
    s = re.sub(r"문제집", "", s)
    s = re.sub(r"기출문제집", "", s)
    s = re.sub(r"개정판", "", s)
    # 약칭 → 정식명 (알라딘/교보 검색 일치율 향상)
    s = re.sub(r"통사(\d)", r"통합사회\1", s)  # 통사2 → 통합사회2
    s = re.sub(r"\b사문\b", "사회문화", s)  # 사문 → 사회문화
    s = re.sub(r"\b생윤\b", "생활과윤리", s)
    s = re.sub(r"\b확통\b", "확률과통계", s)
    s = re.sub(r"\b컴활\b", "컴퓨터활용능력", s)
    # 전각 특수문자 → 반각
    s = s.replace("＋", "+").replace("＆", "&")
    # 학교급 정규화: "중등" → "중학" (알라딘은 "중학" 표기가 많음)
    s = re.sub(r"중학\s+중등", "중학", s)
    s = re.sub(r"중등", "중학", s)
    # "중2-1" → "2-1" (학교급 접두사 제거)
    s = re.sub(r"[중초고](\d+-\d+)", r"\1", s)
    # "초등 초4" → "초등 4"
    s = re.sub(r"초(\d)(?!\d)", r"\1", s)
    s = re.sub(r"중(\d)(?!\d)", r"\1", s)
    s = re.sub(r"고(\d)(?!\d)", r"\1", s)
    # 괄호 안 짧은 설명 제거: (바로배송), (총3권), (네오라이트 형광펜제공)
    s = re.sub(r"\([^)]{0,20}배송[^)]*\)", "", s)
    s = re.sub(r"\(총\d+권\)", "", s)
    s = re.sub(r"\([^)]*제공[^)]*\)", "", s)
    s = re.sub(r"\([^)]*미포함[^)]*\)", "", s)
    # 후행 출판사명/저자명 제거 (알라딘 검색에 방해)
    s = re.sub(r"\s+(?:비상교육|좋은책신사고|동아출판|수경출판사|이퓨처|이퓨쳐|길벗|에듀|매스티안|소마사고력수학|이지스에듀|영진닷컴|아카데미소프트)\s*$", "", s)
    s = re.sub(r"\s+(?:비상교육|좋은책신사고|동아출판|수경출판사|이퓨처|이퓨쳐)\s+편집부\s*$", "", s)
    s = re.sub(r"\s+이홍섭\s*$", "", s)  # 저자명
    # 쓰레기 패턴 제거
    s = re.sub(r"\s*패키지\s*", "", s)  # "패키지 쎈" → "쎈"
    s = re.sub(r"\s*자체브랜드\s*$", "", s)
    s = re.sub(r"\s*팩토\s*수학\s*$", "", s)  # 중복 "팩토 수학"
    s = re.sub(r"\s*수학독해\s*$", "", s)  # 중복 "수학독해"
    s = re.sub(r"\s+\d+번\b", "", s)  # "1-6번" → "1-6"
    # 후행 닫는괄호 잔재: "적용)", "용)"
    s = re.sub(r"\s*(?:적용|고\d)\)\s*$", "", s)
    s = re.sub(r"\s*용\)\s*$", "", s)
    s = re.sub(r"\)\s*$", "", s)
    # 슝슝오늘출발 등 마케팅
    s = re.sub(r"슝슝오늘출발!?", "", s)
    s = re.sub(r"fast\s*post\)?", "", s, flags=re.IGNORECASE)
    # "/ " 구분자 뒤 잡글 제거
    s = re.sub(r"\s*/\s*(?:세이펜|자체브랜드|에듀|soma).*$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s*-\s*/.*$", "", s)
    # 미완성 괄호 제거: "(2권", "(3-1 3-2", "개정교육과정)" 등
    s = re.sub(r"\([^)]*$", "", s)  # 열린 괄호만 있는 경우
    s = re.sub(r"^[^(]*\)", "", s)  # 닫힌 괄호만 있는 경우
    # 잔여 괄호/특수문자 정리
    s = re.sub(r"\(\s*\)", "", s)
    s = re.sub(r"\[\s*\]", "", s)
    s = re.sub(r"[~:;,|]", " ", s)
    # 후행 고립 숫자 제거: "오투 과학 5-1 5" → "오투 과학 5-1"
    s = re.sub(r"(\d+-\d+)\s+\d+\s*$", r"\1", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"^\s*[-+]\s*", "", s)
    s = re.sub(r"\s*[-+]\s*$", "", s)
    return s.strip()


def parse_set_components(name: str) -> list[str]:
    """세트 상품명 → 구성품 검색 쿼리 리스트"""
    # 기본 노이즈 제거
    s = re.sub(r"\(사은품\)|\(선물\)", "", name)
    # "사은품+" 접두사 제거 (뒤에 오는 내용 보존!)
    s = re.sub(r"^\s*사은품\s*\+\s*", "", s)
    # "선물+" 접두사 제거
    s = re.sub(r"^\s*\(?\s*선물\s*\)?\s*\+?\s*", "", s)
    # 후행 "+사은품...", "+미니수첩..." 제거
    s = re.sub(r"\+\s*사은품.*", "", s)
    s = re.sub(r"\+\s*미니수첩.*", "", s)
    s = re.sub(r"\+\s*바빠연습장.*", "", s)
    s = re.sub(r"\+\s*쁘띠수첩.*", "", s)
    s = re.sub(r"\+\s*영어노트.*", "", s)
    s = re.sub(r"\+\s*형광펜.*", "", s)
    s = re.sub(r"\[전\d+권\]", "", s)
    s = re.sub(r"\[.*?\]", "", s)  # 모든 대괄호 태그
    s = re.sub(r"<[^>]*>", "", s)  # 꺾쇠 태그
    s = re.sub(r"\(\d+권세트\)", "", s)  # (2권세트)
    s = re.sub(r"전\s*\d+\s*권", "", s)
    s = re.sub(r"세트", "", s)
    # 후행 학년/학교급 태그: "초3", "중1", "고1", "초4" 단독
    s = re.sub(r"\s+[초중고]\d\s*$", "", s)
    s = re.sub(r"\s+[초중고]\d\s+", " ", s)
    # 후행 출판사명
    s = re.sub(r"\s+(?:비상교육|좋은책신사고|동아출판|수경출판사|이퓨처|이퓨쳐|길벗)\s*$", "", s)
    s = re.sub(r"\s*상세설명참조", "", s)
    # 전각 → 반각
    s = s.replace("＋", "+").replace("，", ",")
    # 괄호 안 학기/상하 표기 제거: "(3상 3하, 3-1 3-2)", "(2상 + 2하 2-1 + 2-2)" 등
    s = re.sub(r"\([^)]*상[^)]*하[^)]*\)", "", s)
    # 쉼표 구분 학기: "3-1,3-2" → "3-1 3-2"
    s = re.sub(r"(\d+-\d+)\s*,\s*(\d+-\d+)", r"\1 \2", s)
    s = re.sub(r"\s+", " ", s).strip()

    # 패턴1: "X 1-1+1-2" → ["X 1-1", "X 1-2"]
    m = re.match(r"^(.+?)\s+(\d+-\d+)\s*\+\s*(\d+-\d+)(.*)$", s)
    if m:
        base = m.group(1).strip()
        return [clean_search_query(f"{base} {m.group(2)}"),
                clean_search_query(f"{base} {m.group(3)}")]

    # 패턴1b: "X 상+하" → ["X 상", "X 하"]
    m = re.match(r"^(.+?)\s+(상)\s*\+\s*(하)(.*)$", s)
    if m:
        base = m.group(1).strip()
        return [clean_search_query(f"{base} {m.group(2)}"),
                clean_search_query(f"{base} {m.group(3)}")]

    # 패턴2: "X 1-1 1-2" (공백 구분 학기, 뒤에 텍스트 허용)
    m = re.match(r"^(.+?)\s+(\d+-\d+)\s+(\d+-\d+)\b(.*)$", s)
    if m:
        base = m.group(1).strip()
        return [clean_search_query(f"{base} {m.group(2)}"),
                clean_search_query(f"{base} {m.group(3)}")]

    # 패턴3: "X 3학년 1, 2학기" or "X 1, 2학기"
    m = re.match(r"^(.+?)\s+(\d+)학년\s+(\d+)\s*,\s*(\d+)\s*학기(.*)$", s)
    if m:
        base = m.group(1).strip()
        grade = m.group(2)
        return [clean_search_query(f"{base} {grade}-{m.group(3)}"),
                clean_search_query(f"{base} {grade}-{m.group(4)}")]

    # 패턴3b: "X N학년 1학기+2학기 (N-1, N-2)" → 괄호 안 학기 추출
    m = re.match(r"^(.+?)\s+\d+학년.*\((\d+-\d+)\s*,?\s*(\d+-\d+)\)(.*)$", s)
    if m:
        base = re.sub(r"\s+\d+학년.*$", "", m.group(1)).strip()
        return [clean_search_query(f"{base} {m.group(2)}"),
                clean_search_query(f"{base} {m.group(3)}")]

    # 패턴3c: "X N학년 1 2학기" (쉼표 없이)
    m = re.match(r"^(.+?)\s+(\d+)학년\s+(\d+)\s+(\d+)\s*학기(.*)$", s)
    if m:
        base = m.group(1).strip()
        grade = m.group(2)
        return [clean_search_query(f"{base} {grade}-{m.group(3)}"),
                clean_search_query(f"{base} {grade}-{m.group(4)}")]

    # 패턴4: + 로 분리 (과목 등)
    parts = re.split(r"\s*\+\s*", s)
    if len(parts) >= 2:
        result = []
        # 학년 정보 추출 (어디든 있으면 공유): "3-1", "초등", "중학" 등
        grade_suffix = ""
        for p in parts:
            gm = re.search(r"(\d+-\d+)", p)
            if gm:
                grade_suffix = gm.group(1)
                break
        base = re.sub(r"\s*\d+-\d+\s*$", "", parts[0]).strip()
        for p in parts:
            p = p.strip()
            p = re.sub(r"\(\s*\)", "", p).strip()
            if not p or len(p) < 3:
                continue
            # 축약형 (숫자-숫자만): base 붙이기
            if re.match(r"^\d+-\d+$", p):
                result.append(clean_search_query(f"{base} {p}"))
            elif re.match(r"^\d+$", p) and len(p) <= 2:
                result.append(clean_search_query(f"{base} {p}"))
            else:
                cleaned = clean_search_query(p)
                # 학년 정보가 없는 짧은 파트에 grade_suffix 전파
                if grade_suffix and not re.search(r"\d+-\d+", cleaned) and len(cleaned) < 20:
                    cleaned = f"{cleaned} {grade_suffix}"
                result.append(cleaned)
        # 쓰레기 컴포넌트 필터 (출판사명만, 워크북만, 사은품 등)
        GARBAGE = {"이퓨처", "이퓨쳐", "본교재 워크북", "워크북", "사은품", "증정", "메모장"}
        filtered = []
        for r in result:
            if len(r) <= 5:
                continue
            if r.strip() in GARBAGE or any(r.strip().startswith(g) for g in GARBAGE):
                continue
            if re.match(r"^[/\s]+", r):  # "/ 이퓨처..." 같은 잔여
                continue
            filtered.append(r)
        return filtered

    # 패턴5: "X 1~3" → ["X 1", "X 2", "X 3"]
    m = re.match(r"^(.+?)\s+(\d+)\s*~\s*(\d+)\s*$", s)
    if m:
        base = m.group(1).strip()
        start, end = int(m.group(2)), int(m.group(3))
        if 1 <= end - start <= 5:
            return [clean_search_query(f"{base} {i}") for i in range(start, end + 1)]

    # 분리 불가 → 전체를 clean한 하나의 쿼리로
    cleaned = clean_search_query(s)
    return [cleaned] if len(cleaned) > 5 else []


def batch_search_set_isbns(set_targets: list, search_cache: dict,
                           skip_api: bool = False) -> dict:
    """세트 상품 구성품 ISBN 일괄 검색

    set_targets: [(df_idx, product_name), ...]
    반환: {df_idx: [isbn1, isbn2, ...]}
    """
    result = {}
    queries_needed = []

    for df_idx, pname in set_targets:
        components = parse_set_components(pname)
        isbns = []
        for comp in components:
            comp_key = comp.strip().lower()
            if comp_key in search_cache and search_cache[comp_key] is not None:
                isbns.append(search_cache[comp_key])
            elif comp_key not in search_cache:
                # 새 쿼리
                queries_needed.append((df_idx, comp, comp_key))
            else:
                # 캐시에 None (이전 실패) → clean 버전으로 재시도
                cleaned = clean_search_query(comp)
                cleaned_key = cleaned.strip().lower()
                if cleaned_key != comp_key and cleaned_key not in search_cache:
                    queries_needed.append((df_idx, cleaned, cleaned_key))
                elif cleaned_key in search_cache and search_cache[cleaned_key] is not None:
                    isbns.append(search_cache[cleaned_key])
        if isbns:
            result[df_idx] = isbns

    # 중복 쿼리 제거
    seen_keys = set()
    unique_queries = []
    for item in queries_needed:
        if item[2] not in seen_keys:
            seen_keys.add(item[2])
            unique_queries.append(item)
    queries_needed = unique_queries

    print(f"  세트 구성품: 조회 필요 {len(queries_needed)}건")

    if skip_api or not queries_needed:
        return result

    aladin_ok, kyobo_ok, fail = 0, 0, 0
    for i, (df_idx, comp, comp_key) in enumerate(queries_needed, 1):
        isbn = None
        # 1차: 알라딘 검색
        try:
            isbn = search_aladin_by_title(comp)
        except RateLimitError:
            print(f"\n  알라딘 API 키 한도 초과! 교보문고로 전환")

        if isbn:
            aladin_ok += 1
        else:
            # 2차: 교보문고 fallback
            isbn = search_kyobo(comp)
            if isbn:
                kyobo_ok += 1
            else:
                fail += 1
            time.sleep(0.3)  # 교보 rate limit

        search_cache[comp_key] = isbn
        if isbn:
            result.setdefault(df_idx, []).append(isbn)

        if i % 50 == 0 or i == len(queries_needed):
            print(f"    [{i}/{len(queries_needed)}] 알라딘={aladin_ok} 교보={kyobo_ok} 실패={fail}")

        if i % 100 == 0:
            save_search_cache(search_cache)

        time.sleep(API_INTERVAL)

    save_search_cache(search_cache)
    total_ok = aladin_ok + kyobo_ok
    print(f"  세트 ISBN 검색 완료: 알라딘 {aladin_ok} + 교보 {kyobo_ok} = {total_ok}건, 실패 {fail}건")
    return result


# ════════════════════════════════════════
# 검색어 생성
# ════════════════════════════════════════

def generate_keywords(product_name: str, brand: str = "",
                      author: str = "", series: str = "",
                      max_kw: int = 8) -> str:
    if not product_name:
        return ""

    txt = product_name
    for pat, repl in NOISE_PATTERNS:
        txt = re.sub(pat, repl, txt, flags=re.IGNORECASE)

    parens = re.findall(r"\(([^)]+)\)", txt)
    txt = re.sub(r"\([^)]*\)", " ", txt)

    preserved = {}
    for i, tp in enumerate(TITLE_PLUS_PATTERNS):
        ph = f"__TP{i}__"
        if tp in txt:
            txt = txt.replace(tp, ph)
            preserved[ph] = tp

    if "+" in txt:
        parts = [p.strip() for p in txt.split("+")]
        txt = next((p for p in parts if len(p) >= 4), parts[0])

    for ph, orig in preserved.items():
        txt = txt.replace(ph, orig)

    txt = re.sub(r"[~\\/:;,.\[\]{}]", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()

    tokens = txt.split()
    for p in parens:
        tokens.extend(p.split())

    priority = []
    if series:
        priority.extend(series.split())
    if brand:
        priority.append(brand)
    if author:
        a = re.sub(r"\s*\(.*?\)", "", author).strip()
        if a:
            priority.extend(a.split(",")[0].strip().split())

    kws, seen = [], set()
    for t in tokens + priority:
        t = t.strip()
        # 쿠팡 검색어 특수문자 제거: [] ? ' & ㆍ ｜ + 등
        t = re.sub(r"[\[\]?'\"&ㆍ｜|+]", "", t).strip()
        if not t or len(t) < 2 or re.match(r"^\d{10,13}$", t):
            continue
        if t.lower() in STOP_WORDS or t.lower() in seen or t.startswith("#"):
            continue
        seen.add(t.lower())
        kws.append(t)

    return ",".join(kws[:max_kw])


# ════════════════════════════════════════
# 사용연도 / 발행언어
# ════════════════════════════════════════

YEAR_PATTERN = re.compile(r"(?:^|\D)(202[3-9])(?:년|\)|$|\s|\D)")


def extract_year(product_name: str, pub_date: str = "") -> str:
    m = YEAR_PATTERN.search(product_name)
    if m:
        return m.group(1)
    if pub_date and len(pub_date) >= 4:
        y = pub_date[:4]
        if y.isdigit() and 2023 <= int(y) <= 2029:
            return y
    return ""


def detect_language(category_name: str) -> str:
    if not category_name:
        return "한국어"
    cat = category_name.lower()
    if "외국" in cat or "영어" in cat or "english" in cat:
        return "영어"
    if "일본" in cat or "일어" in cat:
        return "일본어"
    if "중국" in cat or "중어" in cat:
        return "중국어"
    return "한국어"


# ════════════════════════════════════════
# 등록상품명 정리 (송장용)
# ════════════════════════════════════════

def clean_product_name(name: str) -> str:
    """등록상품명을 송장에서 읽기 쉽게 간결하게 정리

    '(선물) 비상교육 라이트 완성ON 국어 문법 2-1 (2026년) +사은품 미니수첩'
    → '라이트 완성ON 국어 문법 2-1 (2026년)'

    '큐브 개념 초등 수학 6-1 (2026년) - 2022 개정 교육과정, 교과서 개념을 다잡는 기본서'
    → '큐브 개념 초등 수학 6-1 (2026년)'
    """
    if not name:
        return name
    s = name.strip()

    # 1) 마케팅 노이즈 제거
    for pat, repl in NOISE_PATTERNS:
        s = re.sub(pat, repl, s, flags=re.IGNORECASE)

    # 2) 부제목/설명 제거 ( - 이후, : 이후, , 이후의 긴 설명)
    #    단, "기본+응용", "1-2", "3-1" 같은 패턴은 보존
    #    " - " 뒤에 한글 설명이 오면 제거
    s = re.sub(r"\s+-\s+(?:\d{4}\s*개정|개정\s*교육|교육과정|with|비상|능률|좋은책|신사고).*$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+-\s+(?:[가-힣]{2,}\s+[가-힣]{2,}\s+[가-힣]{2,}).*$", "", s)  # 한글 3어절 이상 설명
    s = re.sub(r",\s*(?:[가-힣]{2,}\s+[가-힣]{2,}\s+[가-힣]{2,}).*$", "", s)      # 콤마 뒤 긴 설명
    s = re.sub(r"\s*:\s*(?:[가-힣]{2,}\s+[가-힣]{2,}).*$", "", s)                 # 콜론 뒤 설명

    # 3) 빈 괄호 제거
    s = re.sub(r"\(\s*\)", "", s)
    # 4) 잔여 특수문자 정리
    s = re.sub(r"\s*~+\s*", " ", s)
    s = re.sub(r"\s*-\s*$", "", s)
    s = re.sub(r"\s*[/|]\s*$", "", s)
    s = re.sub(r"\s*:\s*$", "", s)
    s = re.sub(r"\s{2,}", " ", s).strip()

    return s


# ════════════════════════════════════════
# 엑셀 생성 (원본 파일 복사 → in-place 수정)
# ════════════════════════════════════════

def process_account(account: str, filled_path: str, target_rows: list,
                    cache: dict, output_dir: Path, dry_run: bool = False,
                    set_isbns: dict = None) -> int:
    """원본 _filled.xlsx를 그대로 복사 → 판매중+바코드 행만 셀 수정

    서식(색상, 간격, 열너비) 100% 보존.
    target_rows: [(df_idx, isbn), ...] — pandas 인덱스(0-based)와 ISBN 쌍
    set_isbns: {df_idx: [isbn1, isbn2, ...]} — 세트상품 구성품 ISBN
    """
    if not target_rows and not set_isbns:
        print(f"  {account}: 대상 없음 → 건너뜀")
        return 0

    if not set_isbns:
        set_isbns = {}
    count = len(target_rows)
    set_count = len(set_isbns)
    print(f"\n  {account}: {count}건 (바코드) + {set_count}건 (세트ISBN) 처리 중...")

    if dry_run:
        print(f"  [DRY RUN]")
        return count + set_count

    # 원본 파일 복사 (서식 보존)
    import shutil
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"{account}_update.xlsx"
    shutil.copy2(filled_path, str(out_path))

    # 복사본 열기 (in-place 수정)
    wb = load_workbook(str(out_path))
    ws = wb[SHEET]

    # pandas 인덱스 → 엑셀 행번호 변환 (header=3 → 데이터 시작 = 행5, idx0 = 행5)
    DATA_START_ROW = 5
    target_set = {df_idx: isbn for df_idx, isbn in target_rows}

    # 전체 행 col8(대량수정)을 "N"으로 초기화 (이전 스크립트의 잔여 Y 제거)
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=8).value == "Y":
            ws.cell(row=r, column=8, value="N")

    # 대상 행만 수정
    modified = 0
    for df_idx, isbn in target_rows:
        excel_row = DATA_START_ROW + df_idx

        # 상품명 (col2)
        product_name = str(ws.cell(row=excel_row, column=2).value or "").strip()
        brand = str(ws.cell(row=excel_row, column=6).value or "").strip()

        # 알라딘 메타
        meta = cache.get(isbn)
        if meta is None:
            meta = {"author": "", "publisher": brand, "pubDate": "",
                    "seriesName": "", "categoryName": ""}

        publisher = meta.get("publisher", "") or brand
        author = meta.get("author", "")
        series = meta.get("seriesName", "")
        pub_date = meta.get("pubDate", "")
        category = meta.get("categoryName", "")
        year = extract_year(product_name, pub_date)
        language = detect_language(category)

        # ── 셀 수정 ──
        # col2(등록상품명), col3(노출상품명): 원본 유지 (건드리지 않음)
        # col5: 제조사
        ws.cell(row=excel_row, column=5, value=publisher)
        # col6: 브랜드
        ws.cell(row=excel_row, column=6, value=publisher)
        # col7: 검색어
        kw = generate_keywords(product_name, brand=publisher, author=author, series=series)
        ws.cell(row=excel_row, column=7, value=kw)
        # col8: 대량상품수정
        ws.cell(row=excel_row, column=8, value="Y")
        # col231: 바코드 (텍스트 포맷 필수 — General이면 Excel이 과학적 표기법으로 변환)
        bc_cell = ws.cell(row=excel_row, column=231, value=isbn)
        bc_cell.number_format = '@'

        # ── 검색옵션 값 업데이트 ──
        for c in range(30, 229, 2):
            type_val = ws.cell(row=excel_row, column=c).value
            if not type_val:
                continue
            m = re.match(r"\[(\d+)\]", str(type_val))
            if not m:
                continue
            code = m.group(1)
            field = SEARCH_OPTION_MAP.get(code)
            if not field:
                continue

            val_col = c + 1
            if field == "publisher":
                ws.cell(row=excel_row, column=val_col, value=publisher)
            elif field == "author":
                # 저자 옵션값 최대 30자 제한
                ws.cell(row=excel_row, column=val_col, value=author[:30])
            elif field == "isbn":
                c_isbn = ws.cell(row=excel_row, column=val_col, value=isbn)
                c_isbn.number_format = '@'
            elif field == "series" and series:
                ws.cell(row=excel_row, column=val_col, value=series)
            elif field == "year" and year:
                ws.cell(row=excel_row, column=val_col, value=year)
            elif field == "language":
                ws.cell(row=excel_row, column=val_col, value=language)

        modified += 1

    # ── 세트상품 ISBN 기입 (빈 검색옵션 슬롯에 [7939]ISBN 추가) ──
    set_modified = 0
    if set_isbns:
        for df_idx, isbns in set_isbns.items():
            excel_row = DATA_START_ROW + df_idx
            if not isbns:
                continue

            # 현재 행의 빈 검색옵션 슬롯 찾기
            empty_slots = []
            for c in range(30, 229, 2):
                type_val = ws.cell(row=excel_row, column=c).value
                if not type_val or str(type_val).strip() == "":
                    empty_slots.append(c)

            # 이미 [7939]ISBN이 있는 슬롯도 찾기 (기존 값 업데이트 or 스킵)
            existing_isbn_slots = []
            for c in range(30, 229, 2):
                type_val = ws.cell(row=excel_row, column=c).value
                if type_val and "[7939]" in str(type_val):
                    existing_isbn_slots.append(c)

            # 기존 ISBN 슬롯 + 빈 슬롯에 ISBN 채우기
            all_slots = existing_isbn_slots + empty_slots
            filled = 0
            for i, isbn in enumerate(isbns):
                if i >= len(all_slots):
                    break  # 슬롯 부족
                slot_col = all_slots[i]
                ws.cell(row=excel_row, column=slot_col, value="[7939]ISBN")
                isbn_cell = ws.cell(row=excel_row, column=slot_col + 1, value=isbn)
                isbn_cell.number_format = '@'
                filled += 1

            if filled > 0:
                ws.cell(row=excel_row, column=8, value="Y")  # 대량수정 플래그
                set_modified += 1

    if set_modified:
        print(f"    + 세트상품 ISBN: {set_modified}건")

    wb.save(str(out_path))
    total = modified + set_modified
    print(f"  → {out_path} ({modified}건 메타 + {set_modified}건 세트ISBN)")

    return total


# ════════════════════════════════════════
# 메인
# ════════════════════════════════════════

def collect_isbns_pandas(targets: dict) -> tuple[set, dict, dict]:
    """pandas로 빠르게 ISBN 수집 + 계정별 대상 행 인덱스 + 세트 타겟"""
    all_isbns = set()
    account_targets = {}  # acc → [(df_idx, isbn), ...]
    account_set_targets = {}  # acc → [(df_idx, product_name), ...]

    for acc, path in targets.items():
        if not os.path.exists(path):
            continue
        print(f"    {acc} 스캔 중...")
        df = pd.read_excel(path, sheet_name=SHEET, header=3)
        cols = df.columns.tolist()
        # cols[9]=판매상태, cols[230]=바코드
        rows = []
        set_rows = []
        for idx in df.index:
            if df.at[idx, cols[9]] != "판매중":
                continue
            pname = str(df.at[idx, cols[1]]) if pd.notna(df.at[idx, cols[1]]) else ""
            if not pname.strip():
                continue
            isbn = normalize_isbn(df.at[idx, cols[230]])
            if isbn:
                rows.append((idx, isbn))
                all_isbns.add(isbn)
            elif is_set_product(pname):
                # 바코드 없는 세트상품 → ISBN 검색 대상
                set_rows.append((idx, pname))
        account_targets[acc] = rows
        account_set_targets[acc] = set_rows
        print(f"    {acc}: {len(rows)}건 (바코드), 세트 {len(set_rows)}건 (ISBN 검색 대상)")

    return all_isbns, account_targets, account_set_targets


def main():
    parser = argparse.ArgumentParser(description="WING 상품정보 수정 엑셀 생성")
    parser.add_argument("--account", help="특정 계정만")
    parser.add_argument("--skip-api", action="store_true", help="API 호출 스킵 (캐시만)")
    parser.add_argument("--dry-run", action="store_true", help="테스트 (파일 생성 안 함)")
    args = parser.parse_args()

    print("=" * 60)
    print("  WING 상품정보 수정 엑셀 생성기 (원본 템플릿 형식)")
    print("=" * 60)

    targets = {args.account: FILLED_FILES[args.account]} if args.account else FILLED_FILES

    # Step 1: pandas로 빠르게 ISBN 수집 + 대상 행 식별
    print("\n[Step 1] ISBN 수집 (pandas)")
    all_isbns, account_targets, account_set_targets = collect_isbns_pandas(targets)
    total_rows = sum(len(v) for v in account_targets.values())
    total_sets = sum(len(v) for v in account_set_targets.values())
    print(f"  전체: {total_rows}건 (바코드), 유니크 ISBN: {len(all_isbns)}건")
    print(f"  세트상품 (바코드 없음): {total_sets}건")

    # Step 2: 알라딘 API 일괄 조회 (ISBN 메타데이터)
    print(f"\n[Step 2] 알라딘 API 조회 (ISBN 메타)")
    cache = load_cache()
    cache = batch_fetch_aladin(list(all_isbns), cache, skip_api=args.skip_api)

    hit = sum(1 for isbn in all_isbns if cache.get(isbn) is not None)
    miss = len(all_isbns) - hit
    print(f"  메타데이터 보유: {hit}건, 미보유(fallback): {miss}건")

    # Step 3: 세트상품 구성품 ISBN 검색
    all_set_isbns = {}  # acc → {df_idx: [isbn1, isbn2, ...]}
    if total_sets > 0:
        print(f"\n[Step 3] 세트상품 구성품 ISBN 검색")
        search_cache = load_search_cache()
        for acc, set_rows in account_set_targets.items():
            if not set_rows:
                continue
            print(f"\n  {acc}: {len(set_rows)}건 세트상품 분석")
            acc_set_isbns = batch_search_set_isbns(set_rows, search_cache, skip_api=args.skip_api)
            all_set_isbns[acc] = acc_set_isbns
            found = sum(1 for v in acc_set_isbns.values() if v)
            print(f"    ISBN 발견: {found}/{len(set_rows)}건")
        save_search_cache(search_cache)
    else:
        print(f"\n[Step 3] 세트상품 없음 → 건너뜀")

    # Step 4: 계정별 엑셀 생성 (원본 231컬럼 형식)
    print(f"\n[Step 4] 엑셀 생성 (원본 231컬럼 형식)")
    grand_total = 0

    for acc, path in targets.items():
        rows = account_targets.get(acc, [])
        set_isbns = all_set_isbns.get(acc, {})
        count = process_account(acc, path, rows, cache, OUTPUT_DIR,
                                dry_run=args.dry_run, set_isbns=set_isbns)
        grand_total += count

    print(f"\n{'=' * 60}")
    print(f"  완료! 총 {grand_total}건 처리 (바코드 {total_rows} + 세트 {total_sets})")
    if not args.dry_run:
        print(f"  출력 위치: {OUTPUT_DIR}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()

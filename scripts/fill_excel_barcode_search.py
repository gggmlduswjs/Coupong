#!/usr/bin/env python3
"""
엑셀 대량수정 템플릿: 바코드(ISBN) + 검색어 일괄 채우기

2단계:
  Phase 1: DB 매칭 (속성ISBN → DB cpid → DB 상품명)
  Phase 2: 미발견 상품명을 알라딘 API + 교보문고 크롤링으로 ISBN 검색
           → 결과를 name_isbn 맵에 추가 후 전 계정에 적용

사용법:
  python scripts/fill_excel_barcode_search.py              # 전체 (크롤링 포함)
  python scripts/fill_excel_barcode_search.py --db-only     # DB 매칭만 (빠름)
  python scripts/fill_excel_barcode_search.py --dry-run     # 테스트
  python scripts/fill_excel_barcode_search.py --account 007-book  # 특정 계정
"""

import os
import re
import sys
import time
import io
from pathlib import Path

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

# Windows 인코딩 + 버퍼링 제거
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", line_buffering=True)

import pandas as pd
from dotenv import load_dotenv
load_dotenv()

from sqlalchemy import text, create_engine

# ── 설정 ──
EXCEL_FILES = {
    "007-book": "C:/Users/MSI/Desktop/Coupong/007-book.xlsx",
    "007-bm": "C:/Users/MSI/Desktop/Coupong/007-bm.xlsx",
    "007-ez": "C:/Users/MSI/Desktop/Coupong/007-ez.xlsx",
    "002-bm": "C:/Users/MSI/Desktop/Coupong/002-bm.xlsx",
    "big6ceo": "C:/Users/MSI/Desktop/Coupong/big6ceo.xlsx",
}
SHEET = "Template"
HEADER_ROW = 3
ISBN_PATTERN = re.compile(r"^97[89]\d{10}$")

STOP_WORDS = {
    "세트", "set", "권", "원", "년", "판", "개정", "개정판", "최신", "신판",
    "전", "후", "상", "하", "중", "편", "부", "the", "and", "of", "for",
    "a", "an", "in", "to", "is", "기출", "문제집", "교재",
}

# 책 제목에 +가 포함된 알려진 패턴 (분리하면 안 됨)
TITLE_PLUS_PATTERNS = [
    "개념+유형", "기본+응용", "유형+내신", "기본+실력", "Q+Q",
    "개념+연산", "유형+실력", "교과서+익힘",
]

# 상품명 정제 패턴 (검색용) — 순서 중요: 위에서부터 적용
NOISE_PATTERNS = [
    # 사은품/선물 관련 (+ 뒤에 오는 것 포함)
    (r"\+\s*사은품.*", ""), (r"\+\s*선물.*", ""), (r"\+\s*미니수첩.*", ""),
    (r"사은품\s*증정\)?", ""), (r"사은품", ""), (r"증정", ""),
    (r"!사은품!", ""), (r"!전\d+권!", ""),
    # 장식문자
    (r"\*+[^*]*\*+", ""), (r"\*+", ""), (r"\#\w+", ""), (r"\|+", ""),
    (r"!+", ""),
    # 배송/마케팅 노이즈
    (r"슝슝오늘출발\!*", ""), (r"오늘출발", ""), (r"당일발송", ""),
    (r"평일\d+시\w+", ""), (r"비닐포장", ""), (r"상세설명참조", ""),
    # 선물/사은 접두사
    (r"^\s*\(선물\)\s*", ""), (r"^\s*\(사은품\)\s*", ""), (r"^\s*\(사은 증정\)\s*", ""),
    (r"^\s*\(선물 증정\)\s*", ""), (r"^\s*\(당일발송\)\s*", ""),
    # 콜론 뒤 마케팅 문구
    (r"\s*:\s*슝슝.*$", ""), (r"\s*:\s*내신.*$", ""),
    (r"\s*:\s*오늘출발.*$", ""), (r"\s*:\s*평일.*$", ""),
    (r"//+.*$", ""),
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
}


# ════════════════════════════════════════
# DB 로드
# ════════════════════════════════════════

def load_db_isbn():
    """DB에서 ISBN 데이터 로드"""
    engine = create_engine(os.getenv("DATABASE_URL"))
    cpid_isbn = {}
    name_isbn = {}

    with engine.connect() as conn:
        for r in conn.execute(text(
            "SELECT coupang_product_id, isbn FROM listings "
            "WHERE isbn IS NOT NULL AND isbn != '' AND coupang_product_id IS NOT NULL"
        )).fetchall():
            isbn = str(r[1]).split(",")[0].strip()
            if ISBN_PATTERN.match(isbn):
                cpid_isbn[int(r[0])] = isbn

        for r in conn.execute(text(
            "SELECT product_name, isbn FROM listings "
            "WHERE isbn IS NOT NULL AND isbn != '' AND product_name IS NOT NULL"
        )).fetchall():
            isbn = str(r[1]).split(",")[0].strip()
            if ISBN_PATTERN.match(isbn):
                name_isbn[r[0].strip()] = isbn

    print(f"DB 로드: cpid→ISBN {len(cpid_isbn)}건, 상품명→ISBN {len(name_isbn)}건")
    return cpid_isbn, name_isbn


# ════════════════════════════════════════
# 크롤링: 알라딘 API + 교보문고
# ════════════════════════════════════════

def clean_for_search(name: str) -> str:
    """상품명을 검색 키워드로 정제 (v2: 책제목+보존, 노이즈 공격적 제거)"""
    s = name.strip()

    # 1) 노이즈 패턴 제거
    for pat, repl in NOISE_PATTERNS:
        s = re.sub(pat, repl, s, flags=re.IGNORECASE)

    # 2) 괄호 안 내용 제거 (단, 연도/학년은 보존)
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"\s+", " ", s).strip()

    # 3) 책 제목의 + 보존 (임시 치환)
    preserved = {}
    for i, tp in enumerate(TITLE_PLUS_PATTERNS):
        placeholder = f"__TITLEPLUS{i}__"
        if tp in s:
            s = s.replace(tp, placeholder)
            preserved[placeholder] = tp

    # 4) 세트 분리: 남은 +는 실제 세트 구분자
    if "+" in s:
        parts = [p.strip() for p in s.split("+")]
        # 첫 번째 의미있는 파트 선택 (2글자 이상)
        s = next((p for p in parts if len(p) >= 4), parts[0]).strip()

    # 5) 보존된 제목 + 복원
    for placeholder, original in preserved.items():
        s = s.replace(placeholder, original)

    # 6) 잔여 정리
    s = re.sub(r"[~\\/:;,.\[\]{}]", " ", s)  # 특수문자 제거
    s = re.sub(r"\s+", " ", s).strip()
    return s[:60]


def search_aladin(keyword: str, ttb_key: str) -> str | None:
    """알라딘 API로 ISBN 검색 (1회 호출)"""
    import requests
    try:
        r = requests.get(
            "http://www.aladin.co.kr/ttb/api/ItemSearch.aspx",
            params={
                "ttbkey": ttb_key, "Query": keyword, "QueryType": "Keyword",
                "SearchTarget": "Book", "MaxResults": 3,
                "output": "js", "Version": "20131101",
            },
            timeout=10,
        )
        data = r.json()
        for item in data.get("item", []):
            isbn = item.get("isbn13", "")
            if ISBN_PATTERN.match(isbn):
                return isbn
    except Exception:
        pass
    return None


def search_kyobo(keyword: str) -> str | None:
    """교보문고 검색 → ISBN 추출 (검색 페이지 data-bid)"""
    import requests
    from urllib.parse import quote
    try:
        r = requests.get(
            f"https://search.kyobobook.co.kr/search?keyword={quote(keyword)}&target=total",
            headers=HEADERS, timeout=15,
        )
        r.raise_for_status()
        # data-bid 에서 ISBN 추출 (첫 번째 결과)
        m = re.search(r'data-bid="(\d{13})"', r.text)
        if m and ISBN_PATTERN.match(m.group(1)):
            return m.group(1)
    except Exception:
        pass
    return None


def crawl_missing_isbns(missing_names: list, name_isbn: dict) -> dict:
    """미발견 상품명에 대해 알라딘 → 교보 순서로 ISBN 검색"""
    ttb_key = os.getenv("ALADIN_TTB_KEY", "")
    found = {}
    total = len(missing_names)

    print(f"\n{'='*60}")
    print(f"  Phase 2: 크롤링 ({total}개 유니크 상품명)")
    print(f"{'='*60}")
    if not ttb_key:
        print("  경고: ALADIN_TTB_KEY 없음 → 교보문고만 사용")

    aladin_ok, kyobo_ok, fail = 0, 0, 0

    for i, name in enumerate(missing_names, 1):
        search_q = clean_for_search(name)
        if len(search_q) < 4:
            fail += 1
            continue

        isbn = None

        # 알라딘 먼저
        if ttb_key:
            isbn = search_aladin(search_q, ttb_key)
            if isbn:
                aladin_ok += 1

        # 교보문고 fallback
        if not isbn:
            isbn = search_kyobo(search_q)
            time.sleep(0.5)  # 교보 rate limit
            if isbn:
                kyobo_ok += 1

        if isbn:
            found[name] = isbn
            name_isbn[name] = isbn  # 다음 계정에도 적용되도록
        else:
            fail += 1

        # 진행률 (100건마다)
        if i % 100 == 0 or i == total:
            print(f"  [{i}/{total}] 알라딘={aladin_ok} 교보={kyobo_ok} 실패={fail}")

    print(f"\n  크롤링 완료: 알라딘 {aladin_ok} + 교보 {kyobo_ok} = {len(found)}건 발견, {fail}건 실패")
    return found


# ════════════════════════════════════════
# 헬퍼
# ════════════════════════════════════════

def extract_isbn_from_attr(row, cols):
    """구매옵션 속성에서 ISBN 추출"""
    for i in range(13, 29, 2):
        attr_name = str(row[cols[i]]) if pd.notna(row[cols[i]]) else ""
        attr_val = str(row[cols[i + 1]]) if pd.notna(row[cols[i + 1]]) else ""
        if "ISBN" in attr_name.upper() and attr_val:
            digits = re.sub(r"\D", "", attr_val)
            if ISBN_PATTERN.match(digits):
                return digits
    model = str(row[cols[229]]) if pd.notna(row[cols[229]]) else ""
    digits = re.sub(r"\D", "", model)
    if ISBN_PATTERN.match(digits):
        return digits
    return None


def generate_search_keywords(product_name: str, brand: str = "", max_kw: int = 20) -> str:
    """상품명에서 검색 키워드 생성"""
    if not product_name:
        return ""
    txt = product_name
    parens = re.findall(r"\(([^)]+)\)", txt)
    txt = re.sub(r"\([^)]*\)", " ", txt)
    txt = re.sub(r"[+&*#@!~`|\\/<>{}=\-:,.]", " ", txt)
    tokens = txt.split()
    for p in parens:
        tokens.extend(p.split())
    if brand and brand not in tokens:
        tokens.append(brand)
    kws, seen = [], set()
    for t in tokens:
        t = t.strip()
        if not t or len(t) < 2 or t.lower() in STOP_WORDS or t.lower() in seen:
            continue
        seen.add(t.lower())
        kws.append(t)
    return ",".join(kws[:max_kw])


def has_valid_barcode(val):
    """바코드 값이 유효한 ISBN인지"""
    if pd.isna(val):
        return False
    try:
        s = str(int(val)) if isinstance(val, float) else str(val)
        return bool(ISBN_PATTERN.match(re.sub(r"\D", "", s)))
    except (ValueError, OverflowError):
        return False


# ════════════════════════════════════════
# 엑셀 처리
# ════════════════════════════════════════

def collect_missing_names(cpid_isbn, name_isbn):
    """전 계정 엑셀에서 ISBN 미발견 유니크 상품명 수집"""
    missing = set()
    for acc, path in EXCEL_FILES.items():
        if not os.path.exists(path):
            continue
        df = pd.read_excel(path, sheet_name=SHEET, header=HEADER_ROW)
        cols = df.columns.tolist()
        active = df[df[cols[9]] == "판매중"]
        for _, row in active.iterrows():
            if has_valid_barcode(row[cols[230]]):
                continue
            if extract_isbn_from_attr(row, cols):
                continue
            cpid = row[cols[0]]
            if pd.notna(cpid):
                try:
                    if int(float(cpid)) in cpid_isbn:
                        continue
                except (ValueError, OverflowError):
                    pass
            pname = str(row[cols[1]]).strip() if pd.notna(row[cols[1]]) else ""
            if pname in name_isbn:
                continue
            if pname:
                missing.add(pname)
    return list(missing)


def fill_excel(account_name, excel_path, cpid_isbn, name_isbn, dry_run=False):
    """엑셀 채우기 (name_isbn에 크롤링 결과도 포함됨)"""
    print(f"\n{'='*60}")
    print(f"  {account_name}")
    print(f"{'='*60}")

    df = pd.read_excel(excel_path, sheet_name=SHEET, header=HEADER_ROW)
    cols = df.columns.tolist()
    COL_CPID, COL_NAME, COL_BRAND = cols[0], cols[1], cols[5]
    COL_SEARCH, COL_STATUS = cols[6], cols[9]
    COL_BARCODE, COL_BULK_YN = cols[230], cols[7]

    active_mask = df[COL_STATUS] == "판매중"
    stats = {"attr": 0, "db_cpid": 0, "db_name": 0, "crawl": 0,
             "already": 0, "missing": 0, "search_new": 0, "search_ok": 0, "modified": 0}

    for idx in df[active_mask].index:
        row = df.loc[idx]
        modified = False

        # ── 바코드 ──
        if has_valid_barcode(row[COL_BARCODE]):
            stats["already"] += 1
        else:
            isbn = extract_isbn_from_attr(row, cols)
            src = "attr"
            if not isbn:
                cpid = row[COL_CPID]
                if pd.notna(cpid):
                    try:
                        isbn = cpid_isbn.get(int(float(cpid)))
                        src = "db_cpid"
                    except (ValueError, OverflowError):
                        pass
            if not isbn:
                pname = str(row[COL_NAME]).strip() if pd.notna(row[COL_NAME]) else ""
                isbn = name_isbn.get(pname)
                # DB 매칭이면 db_name, 크롤링이면 crawl
                src = "db_name"

            if isbn:
                df.at[idx, COL_BARCODE] = int(isbn)
                modified = True
                # 소스 구분 (크롤링으로 추가된 건 crawl로 카운트)
                if src == "attr":
                    stats["attr"] += 1
                elif src == "db_cpid":
                    stats["db_cpid"] += 1
                else:
                    stats["db_name"] += 1
            else:
                stats["missing"] += 1

        # ── 검색어 ──
        if pd.notna(row[COL_SEARCH]) and str(row[COL_SEARCH]).strip():
            stats["search_ok"] += 1
        else:
            pname = str(row[COL_NAME]) if pd.notna(row[COL_NAME]) else ""
            brand = str(row[COL_BRAND]) if pd.notna(row[COL_BRAND]) else ""
            kw = generate_search_keywords(pname, brand)
            if kw:
                df.at[idx, COL_SEARCH] = kw
                modified = True
                stats["search_new"] += 1

        if modified:
            df.at[idx, COL_BULK_YN] = "Y"
            stats["modified"] += 1

    total_active = active_mask.sum()
    barcode_filled = stats["attr"] + stats["db_cpid"] + stats["db_name"]
    print(f"  판매중: {total_active} | 바코드채움: {barcode_filled} (속성{stats['attr']} DB-cpid{stats['db_cpid']} DB-name{stats['db_name']}) | 미발견: {stats['missing']} | 검색어: {stats['search_new']}건 생성 | 수정행: {stats['modified']}")

    # 저장
    if not dry_run and stats["modified"] > 0:
        out_path = excel_path.replace(".xlsx", "_filled.xlsx")
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb[SHEET]
        data_start_row = HEADER_ROW + 2

        for idx in df.index:
            row_num = data_start_row + idx
            bv = df.at[idx, COL_BARCODE]
            if pd.notna(bv):
                try:
                    ws.cell(row=row_num, column=231, value=int(bv))
                except (ValueError, OverflowError):
                    ws.cell(row=row_num, column=231, value=str(bv))
            sv = df.at[idx, COL_SEARCH]
            if pd.notna(sv):
                ws.cell(row=row_num, column=7, value=str(sv))
            yv = df.at[idx, COL_BULK_YN]
            if pd.notna(yv):
                ws.cell(row=row_num, column=8, value=str(yv))

        wb.save(out_path)
        print(f"  → {out_path}")
    elif dry_run:
        print(f"  [DRY RUN]")

    return stats


# ════════════════════════════════════════
# 메인
# ════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(description="엑셀 바코드+검색어 채우기")
    parser.add_argument("--dry-run", action="store_true", help="테스트 모드 (저장 안 함)")
    parser.add_argument("--db-only", action="store_true", help="DB 매칭만 (크롤링 안 함)")
    parser.add_argument("--account", help="특정 계정만")
    args = parser.parse_args()

    mode = "DRY RUN" if args.dry_run else ("DB ONLY" if args.db_only else "FULL (DB+크롤링)")
    print(f"엑셀 바코드/검색어 채우기 [{mode}]")

    # Phase 1: DB 로드
    cpid_isbn, name_isbn = load_db_isbn()

    # Phase 2: 크롤링 (--db-only가 아닐 때)
    if not args.db_only:
        print("\n미발견 상품명 수집 중...")
        missing = collect_missing_names(cpid_isbn, name_isbn)
        print(f"  유니크 미발견: {len(missing)}개")
        if missing:
            crawl_missing_isbns(missing, name_isbn)
            print(f"  name_isbn 맵 갱신: {len(name_isbn)}건")

    # Phase 3: 엑셀 채우기 + 저장
    targets = {args.account: EXCEL_FILES[args.account]} if args.account else EXCEL_FILES
    all_stats = []
    for acc, path in targets.items():
        if not os.path.exists(path):
            print(f"\n  {acc}: 파일 없음")
            continue
        s = fill_excel(acc, path, cpid_isbn, name_isbn, dry_run=args.dry_run)
        all_stats.append(s)

    # 요약
    print(f"\n{'='*60}")
    t_filled = sum(s["attr"] + s["db_cpid"] + s["db_name"] for s in all_stats)
    t_search = sum(s["search_new"] for s in all_stats)
    t_miss = sum(s["missing"] for s in all_stats)
    t_mod = sum(s["modified"] for s in all_stats)
    print(f"  바코드 채움: {t_filled} | 검색어 생성: {t_search} | 바코드 미발견: {t_miss} | 수정행: {t_mod}")


if __name__ == "__main__":
    main()

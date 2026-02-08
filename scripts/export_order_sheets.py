"""
거래처별 발주서 Excel 내보내기
==============================
주문 데이터를 거래처(총판)별로 그룹핑하여 Excel 파일 생성

사용법:
    python scripts/export_order_sheets.py --days 7
    python scripts/export_order_sheets.py --days 14 --account 007-book
    python scripts/export_order_sheets.py --all-status --output 발주서.xlsx
"""
import sys
import re
import argparse
from pathlib import Path
from datetime import date, timedelta

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd
from sqlalchemy import text
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.database import engine
from app.constants import resolve_distributor, match_publisher_from_text


def query_df(sql: str, params: dict = None) -> pd.DataFrame:
    if params:
        return pd.read_sql(text(sql), engine, params=params)
    return pd.read_sql(sql, engine)


def get_publisher_names() -> list:
    """활성 출판사 목록 (긴 이름 우선)"""
    df = query_df("SELECT name FROM publishers WHERE is_active = 1 ORDER BY LENGTH(name) DESC")
    return df["name"].tolist() if not df.empty else []


def _match_pub(item_name: str, product_name: str, pub_names: list) -> str:
    """옵션명 → 상품명 순서로 출판사 매칭 (2-pass: DB이름 + 시리즈명)"""
    result = match_publisher_from_text(str(item_name or ""), pub_names)
    if not result and product_name:
        result = match_publisher_from_text(str(product_name), pub_names)
    return result


def clean_book_name(name: str) -> str:
    """상품명에서 불필요한 태그 제거"""
    name = str(name or "").strip()
    name = re.sub(r'^\(사은품\)\s*', '', name)
    name = re.sub(r'\s*사은품증정\s*', '', name)
    name = re.sub(r'\s*수첩형메모지\+형광펜\s*증정\s*', '', name)
    name = re.sub(r'\s*\+사은품\s*', '', name)
    name = re.sub(r'\s*\+선물\s*', '', name)
    return name.strip()


def split_set_name(name: str) -> list:
    """세트 상품명 → 개별 도서명 리스트"""
    if '+' not in name:
        return [name]
    clean = name
    clean = re.sub(r'\(사은품\)\s*', '', clean)
    clean = re.sub(r'\+선물', '', clean)
    clean = re.sub(r'\+사은품', '', clean)
    clean = re.sub(r'\+증정', '', clean)
    clean = re.sub(r'\(전\d+권\)', '', clean)
    clean = re.sub(r'전\d+권', '', clean)
    clean = re.sub(r'\(\d{4}년?\)', '', clean)
    clean = re.sub(r'\(2022\s*개정[^)]*\)', '', clean)
    clean = re.sub(r'사은품증정', '', clean)
    # "세트" 뒤 공유 접미사 추출
    _shared_suffix = ""
    _set_match = re.search(r'세트\s*[-–]?\s*(.+)$', clean)
    if _set_match:
        sc = _set_match.group(1).strip()
        sc = re.sub(r'\d{4}년', '', sc).strip()
        sc = re.sub(r'\(.*?\)', '', sc).strip()
        sc = sc.rstrip('-').strip()
        sc = re.sub(r'\s*-\s*\S+제공.*$', '', sc)
        sc = re.sub(r'\s*-\s*\S+적용.*$', '', sc)
        if 2 <= len(sc) <= 15 and not re.match(r'^[\d\s\-]+$', sc):
            _shared_suffix = sc
    clean = re.sub(r'\s*세트\s*[-–]?\s*.*$', '', clean)
    clean = re.sub(r'\d{4}년', '', clean)
    clean = clean.strip().rstrip('-').strip()
    if '+' not in clean:
        return [name]
    parts = [p.strip() for p in clean.split('+') if p.strip()]
    if len(parts) < 2:
        return [name]
    result = [parts[0]]
    for p in parts[1:]:
        if re.match(r'^[\d\-]+$', p.strip()):
            prefix = re.sub(r'\s*\d+[\-]\d+\s*$', '', parts[0]).strip()
            if not prefix:
                prefix = re.sub(r'\s+\d+\s*$', '', parts[0]).strip()
            result.append(f"{prefix} {p.strip()}")
        else:
            result.append(p)
    if _shared_suffix:
        result = [f"{r} {_shared_suffix}" if _shared_suffix not in r else r for r in result]
    # 짧은 파트에 공통 접두사 상속 (예: "자이스토리 영어 독해 기본+완성" → 완성이 단독이면 안 됨)
    if len(result) >= 2:
        first = result[0]
        # 첫 파트에서 마지막 단어를 제거하여 base prefix 추출
        words = first.rsplit(' ', 1)
        base_prefix = words[0] if len(words) > 1 else ""
        if base_prefix and len(base_prefix) > 4:
            for i in range(1, len(result)):
                if len(result[i]) <= 4 and not re.search(r'\d', result[i]):
                    result[i] = f"{base_prefix} {result[i]}"
    return result


def _style_sheet(ws, num_cols: int, num_data_rows: int, title: str, show_sum: bool = True):
    """시트 공통 스타일링: 타이틀 + 헤더 + 합계"""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    # 타이틀 행
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1)
    cell.value = title
    cell.font = Font(bold=True, size=13)
    cell.alignment = Alignment(horizontal="center")

    # 헤더 행 (row 2)
    for ci in range(1, num_cols + 1):
        c = ws.cell(row=2, column=ci)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    # 데이터 행 테두리
    for ri in range(3, 3 + num_data_rows):
        for ci in range(1, num_cols + 1):
            ws.cell(row=ri, column=ci).border = thin_border

    # 합계 행
    if show_sum and num_data_rows > 0:
        sum_row = 3 + num_data_rows
        sum_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        ws.cell(row=sum_row, column=1, value="합계").font = Font(bold=True, size=11)
        ws.cell(row=sum_row, column=1).fill = sum_fill
        # 마지막 열 = 수량
        total = 0
        for ri in range(3, 3 + num_data_rows):
            v = ws.cell(row=ri, column=num_cols).value
            if isinstance(v, (int, float)):
                total += int(v)
        ws.cell(row=sum_row, column=num_cols, value=total).font = Font(bold=True, size=11)
        ws.cell(row=sum_row, column=num_cols).fill = sum_fill
        for ci in range(1, num_cols + 1):
            ws.cell(row=sum_row, column=ci).border = thin_border


def export_order_sheets(days: int = 7, account_name: str = None,
                        all_status: bool = False, output: str = None,
                        split_set: bool = True):
    """거래처별 발주서 Excel 생성 (간결 형식: 도서명 | 출판사 | 주문수량)"""
    # 최신 주문 데이터 동기화
    print("주문 데이터 동기화 중...")
    from scripts.sync_orders import OrderSync
    syncer = OrderSync()
    syncer.sync_all(days=days, account_name=account_name)
    print("동기화 완료.\n")

    date_to = date.today()
    date_from = date_to - timedelta(days=days)

    # WHERE 절
    acct_where = ""
    params = {}
    if account_name:
        acct_where = "AND o.account_id = (SELECT id FROM accounts WHERE account_name = :acct_name)"
        params["acct_name"] = account_name

    status_where = "" if all_status else "AND o.status = 'INSTRUCT'"
    date_where = f"AND DATE(o.ordered_at) >= '{date_from.isoformat()}' AND DATE(o.ordered_at) <= '{date_to.isoformat()}'"

    # DB JOIN으로 실제 출판사 가져오기 (orders → listings → products → books → publishers)
    orders = query_df(f"""
        SELECT
            o.seller_product_name as 상품명,
            o.vendor_item_name as 옵션명,
            o.shipping_count as 수량,
            pub.name as DB출판사
        FROM orders o
        LEFT JOIN listings l ON o.listing_id = l.id
        LEFT JOIN products p ON l.product_id = p.id
        LEFT JOIN books b ON p.book_id = b.id
        LEFT JOIN publishers pub ON b.publisher_id = pub.id
        WHERE 1=1 {acct_where} {status_where} {date_where}
        ORDER BY o.ordered_at DESC
    """, params)

    if orders.empty:
        print("해당 조건의 주문이 없습니다.")
        return

    # 출판사 매칭: DB JOIN 우선 → 텍스트 매칭 fallback
    pub_names = get_publisher_names()

    def _resolve_publisher(row):
        db_pub = row.get("DB출판사")
        if pd.notna(db_pub) and db_pub:
            return str(db_pub)
        return _match_pub(row.get("옵션명", ""), row.get("상품명", ""), pub_names)

    orders["출판사"] = orders.apply(_resolve_publisher, axis=1)

    # 도서명 정리: seller_product_name 기반
    orders["도서명"] = orders["상품명"].apply(clean_book_name)

    # 세트 분리
    if split_set:
        expanded = []
        for _, row in orders.iterrows():
            bname = row["도서명"]
            parts = split_set_name(bname)
            if len(parts) >= 2:
                for part in parts:
                    new_row = row.copy()
                    new_row["도서명"] = part.strip()
                    expanded.append(new_row)
            else:
                expanded.append(row.copy())
        orders = pd.DataFrame(expanded)

    # 거래처 매핑
    orders["거래처"] = orders["출판사"].apply(resolve_distributor)

    # ===== 도서별 합산 =====
    agg = orders.groupby(["거래처", "출판사", "도서명"]).agg(
        주문수량=("수량", "sum"),
    ).reset_index().sort_values(["거래처", "출판사", "도서명"])

    # 출력 파일명
    if not output:
        output = f"주문{date_to.strftime('%m%d')}.xlsx"

    # ===== Excel 생성 =====
    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # --- 1) 전체 목록 시트 (원본 옵션명 | 수량) ---
        raw_agg = orders.groupby("옵션명").agg(수량=("수량", "sum")).reset_index()
        raw_agg = raw_agg.sort_values("옵션명")
        raw_agg.to_excel(writer, sheet_name="전체", index=False, startrow=1)
        ws_all = writer.sheets["전체"]
        _style_sheet(ws_all, 2, len(raw_agg),
                     f"전체 주문 목록 ({date_from} ~ {date_to})")
        ws_all.column_dimensions["A"].width = 70
        ws_all.column_dimensions["B"].width = 8

        # --- 2) 요약 시트 ---
        summary = agg.groupby("거래처").agg(
            품목수=("도서명", "count"),
            총수량=("주문수량", "sum"),
        ).reset_index().sort_values("총수량", ascending=False)
        summary.to_excel(writer, sheet_name="요약", index=False, startrow=1)
        ws_sum = writer.sheets["요약"]
        _style_sheet(ws_sum, 3, len(summary),
                     f"거래처별 요약 ({date_from} ~ {date_to})")
        ws_sum.column_dimensions["A"].width = 12
        ws_sum.column_dimensions["B"].width = 10
        ws_sum.column_dimensions["C"].width = 10

        # --- 3) 거래처별 시트 (도서명 | 출판사 | 주문수량) ---
        dist_order = ["제일", "대성", "일신", "서부", "북전", "동아", "강우사", "대원", "일반"]
        all_dists = sorted(agg["거래처"].unique(), key=lambda d: dist_order.index(d) if d in dist_order else 99)

        for dist_name in all_dists:
            sheet_df = agg[agg["거래처"] == dist_name][["도서명", "출판사", "주문수량"]].copy()
            if sheet_df.empty:
                continue
            sheet_df = sheet_df.sort_values(["출판사", "도서명"])

            safe_name = dist_name[:31].replace("/", "_").replace("\\", "_")
            sheet_df.to_excel(writer, sheet_name=safe_name, index=False, startrow=1)

            ws = writer.sheets[safe_name]
            _style_sheet(ws, 3, len(sheet_df),
                         f"[{dist_name}] 발주서 ({date_from} ~ {date_to})")
            ws.column_dimensions["A"].width = 45
            ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 10
            # 수량 열 가운데 정렬
            for ri in range(3, 3 + len(sheet_df)):
                ws.cell(row=ri, column=3).alignment = Alignment(horizontal="center")

        # --- 4) 거래처-출판사 매핑표 시트 ---
        from app.constants import DISTRIBUTOR_MAP
        map_rows = []
        for dist, pubs in DISTRIBUTOR_MAP.items():
            for i, pub in enumerate(pubs):
                map_rows.append({"거래처명": dist if i == 0 else "", "출판사": pub})
            map_rows.append({"거래처명": "", "출판사": ""})
        map_df = pd.DataFrame(map_rows)
        map_df.to_excel(writer, sheet_name="거래처매핑", index=False)
        ws_map = writer.sheets["거래처매핑"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for ci in range(1, 3):
            c = ws_map.cell(row=1, column=ci)
            c.fill = header_fill
            c.font = Font(bold=True, color="FFFFFF")
        ws_map.column_dimensions["A"].width = 12
        ws_map.column_dimensions["B"].width = 16

    print(f"발주서 생성 완료: {output}")
    print(f"  기간: {date_from} ~ {date_to}")
    total_items = int(agg["주문수량"].sum())
    print(f"  총 품목: {len(agg)}종 / {total_items}권")
    print(f"  거래처: {len(agg['거래처'].unique())}곳")
    for _, row in summary.iterrows():
        print(f"    {row['거래처']}: {int(row['품목수'])}종 / {int(row['총수량'])}권")


def main():
    parser = argparse.ArgumentParser(description="거래처별 발주서 Excel 내보내기")
    parser.add_argument("--days", type=int, default=7, help="조회 기간 (일, 기본 7)")
    parser.add_argument("--account", type=str, default=None, help="특정 계정만 (예: 007-book)")
    parser.add_argument("--all-status", action="store_true", help="전체 상태 포함 (기본: 미출고만)")
    parser.add_argument("--no-split", action="store_true", help="세트 상품 분리 안 함")
    parser.add_argument("--output", type=str, default=None, help="출력 파일명 (기본: 주문MMDD.xlsx)")
    args = parser.parse_args()

    export_order_sheets(
        days=args.days,
        account_name=args.account,
        all_status=args.all_status,
        output=args.output,
        split_set=not args.no_split,
    )


if __name__ == "__main__":
    main()

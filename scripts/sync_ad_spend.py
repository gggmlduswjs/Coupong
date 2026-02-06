"""
광고비 정산 동기화 스크립트
========================
쿠팡 광고비 정산 Excel 파일 → ad_spends 테이블

Excel 형식: {vendorId}-dailySettlement-{YYYYMMDD}-{YYYYMMDD}.xlsx
- 행 구조: 날짜 요약행(배송유형 있음) + 캠페인 상세행(광고유형 있음) 교대 반복
- 상세행 식별: 컬럼4(광고 유형)에 값이 있는 행 = 캠페인 레벨 데이터
- 날짜 전파: 상세행은 날짜가 NaN → 직전 요약행의 날짜를 forward-fill

사용법:
    python scripts/sync_ad_spend.py path/to/excel.xlsx
    python scripts/sync_ad_spend.py --dir path/to/folder   # 폴더 내 모든 xlsx
"""
import os
import sys
import re
import argparse
import logging
from datetime import datetime, date
from pathlib import Path
from typing import List, Optional, Tuple

from sqlalchemy import create_engine, text

# 프로젝트 루트
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv
load_dotenv(ROOT / ".env")

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)


class AdSpendSync:
    """광고비 정산 Excel 파싱 + DB 저장"""

    CREATE_TABLE_SQL = """
    CREATE TABLE IF NOT EXISTS ad_spends (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        account_id INTEGER NOT NULL REFERENCES accounts(id),
        ad_date DATE NOT NULL,
        campaign_id VARCHAR(50) NOT NULL,
        campaign_name VARCHAR(200),
        ad_type VARCHAR(20),
        ad_objective VARCHAR(50),
        daily_budget INTEGER DEFAULT 0,
        spent_amount INTEGER DEFAULT 0,
        adjustment INTEGER DEFAULT 0,
        spent_after_adjust INTEGER DEFAULT 0,
        over_spend INTEGER DEFAULT 0,
        billable_cost INTEGER DEFAULT 0,
        vat_amount INTEGER DEFAULT 0,
        total_charge INTEGER DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(account_id, ad_date, campaign_id)
    )
    """

    CREATE_INDEXES_SQL = [
        "CREATE INDEX IF NOT EXISTS ix_ad_account_date ON ad_spends(account_id, ad_date)",
        "CREATE INDEX IF NOT EXISTS ix_ad_date ON ad_spends(ad_date)",
    ]

    def __init__(self, db_path: str = None):
        if db_path is None:
            db_path = str(ROOT / "coupang_auto.db")
        self.engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False, "timeout": 30})
        self._ensure_table()

    def _ensure_table(self):
        """테이블이 없으면 생성"""
        with self.engine.connect() as conn:
            conn.execute(text(self.CREATE_TABLE_SQL))
            for idx_sql in self.CREATE_INDEXES_SQL:
                conn.execute(text(idx_sql))
            conn.commit()
        logger.info("ad_spends 테이블 확인 완료")

    def _safe_int(self, val) -> int:
        """안전한 정수 변환"""
        if val is None:
            return 0
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return 0

    def _parse_korean_date(self, text_val: str) -> Optional[date]:
        """'2026년 01월 04일' → date(2026, 1, 4)"""
        if not text_val:
            return None
        m = re.search(r"(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일", str(text_val))
        if m:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        return None

    def _extract_vendor_id(self, filepath: str) -> Optional[str]:
        """파일명에서 vendor_id 추출: A01234216-dailySettlement-... → A01234216"""
        fname = Path(filepath).stem
        # 파일명 첫 부분 (하이픈 전)
        parts = fname.split("-")
        if parts:
            return parts[0]
        return None

    def _find_account_id(self, vendor_id: str) -> Optional[int]:
        """vendor_id로 accounts 테이블에서 account_id 조회"""
        with self.engine.connect() as conn:
            row = conn.execute(
                text("SELECT id FROM accounts WHERE vendor_id = :vid LIMIT 1"),
                {"vid": vendor_id}
            ).fetchone()
            if row:
                return row[0]
        return None

    def parse_excel(self, filepath: str) -> Tuple[Optional[int], List[dict]]:
        """
        Excel 파싱 → (account_id, rows_list) 반환

        Returns:
            (account_id, [{"ad_date": date, "campaign_id": str, ...}, ...])
        """
        import openpyxl

        vendor_id = self._extract_vendor_id(filepath)
        if not vendor_id:
            logger.error(f"파일명에서 vendor_id 추출 실패: {filepath}")
            return None, []

        account_id = self._find_account_id(vendor_id)
        if not account_id:
            logger.error(f"vendor_id '{vendor_id}'에 해당하는 계정을 찾을 수 없음")
            return None, []

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        rows = []
        current_date = None
        current_vat = 0       # 요약행의 VAT
        current_charge = 0    # 요약행의 총청구

        for row_idx in range(2, ws.max_row + 1):  # 헤더(1행) 건너뛰기
            col1 = ws.cell(row=row_idx, column=1).value   # 날짜
            col2 = ws.cell(row=row_idx, column=2).value   # 배송유형
            col4 = ws.cell(row=row_idx, column=4).value   # 광고 유형

            # 날짜 요약행: col1에 날짜가 있는 행
            if col1 and str(col1).strip():
                parsed = self._parse_korean_date(str(col1))
                if parsed:
                    current_date = parsed
                    current_vat = self._safe_int(ws.cell(row=row_idx, column=16).value)
                    current_charge = self._safe_int(ws.cell(row=row_idx, column=17).value)
                continue  # 요약행은 스킵 (상세행만 저장)

            # 상세행: col4(광고유형)에 값이 있는 행 (PA 등)
            if col4 and str(col4).strip() and str(col4).strip() != "-":
                if current_date is None:
                    continue

                campaign_id = str(ws.cell(row=row_idx, column=6).value or "").strip()
                if not campaign_id:
                    continue

                rows.append({
                    "ad_date": current_date,
                    "campaign_id": campaign_id,
                    "campaign_name": str(ws.cell(row=row_idx, column=7).value or "").strip(),
                    "ad_type": str(col4).strip(),
                    "ad_objective": str(ws.cell(row=row_idx, column=5).value or "").strip(),
                    "daily_budget": self._safe_int(ws.cell(row=row_idx, column=10).value),
                    "spent_amount": self._safe_int(ws.cell(row=row_idx, column=11).value),
                    "adjustment": self._safe_int(ws.cell(row=row_idx, column=12).value),
                    "spent_after_adjust": self._safe_int(ws.cell(row=row_idx, column=13).value),
                    "over_spend": self._safe_int(ws.cell(row=row_idx, column=14).value),
                    "billable_cost": self._safe_int(ws.cell(row=row_idx, column=15).value),
                    "vat_amount": current_vat,
                    "total_charge": current_charge,
                })

        wb.close()
        logger.info(f"파싱 완료: {filepath} → {len(rows)}건 (vendor={vendor_id}, account_id={account_id})")
        return account_id, rows

    def save_to_db(self, account_id: int, rows: List[dict]) -> int:
        """파싱된 데이터를 DB에 INSERT OR REPLACE"""
        if not rows:
            return 0

        upserted = 0
        with self.engine.connect() as conn:
            for row in rows:
                try:
                    conn.execute(text("""
                        INSERT OR REPLACE INTO ad_spends
                        (account_id, ad_date, campaign_id, campaign_name,
                         ad_type, ad_objective, daily_budget,
                         spent_amount, adjustment, spent_after_adjust,
                         over_spend, billable_cost, vat_amount, total_charge)
                        VALUES
                        (:account_id, :ad_date, :campaign_id, :campaign_name,
                         :ad_type, :ad_objective, :daily_budget,
                         :spent_amount, :adjustment, :spent_after_adjust,
                         :over_spend, :billable_cost, :vat_amount, :total_charge)
                    """), {
                        "account_id": account_id,
                        "ad_date": row["ad_date"].isoformat(),
                        "campaign_id": row["campaign_id"],
                        "campaign_name": row["campaign_name"],
                        "ad_type": row["ad_type"],
                        "ad_objective": row["ad_objective"],
                        "daily_budget": row["daily_budget"],
                        "spent_amount": row["spent_amount"],
                        "adjustment": row["adjustment"],
                        "spent_after_adjust": row["spent_after_adjust"],
                        "over_spend": row["over_spend"],
                        "billable_cost": row["billable_cost"],
                        "vat_amount": row["vat_amount"],
                        "total_charge": row["total_charge"],
                    })
                    upserted += 1
                except Exception as e:
                    logger.debug(f"INSERT 스킵: {e}")

            conn.commit()

        logger.info(f"저장 완료: {upserted}/{len(rows)}건")
        return upserted

    def sync_file(self, filepath: str) -> dict:
        """단일 파일 동기화"""
        account_id, rows = self.parse_excel(filepath)
        if account_id is None:
            return {"file": filepath, "error": "계정 매칭 실패", "parsed": 0, "saved": 0}

        saved = self.save_to_db(account_id, rows)

        # 계정명 조회
        with self.engine.connect() as conn:
            name_row = conn.execute(
                text("SELECT account_name FROM accounts WHERE id = :aid"),
                {"aid": account_id}
            ).fetchone()
            account_name = name_row[0] if name_row else str(account_id)

        # 기간 정보
        dates = [r["ad_date"] for r in rows]
        date_range = f"{min(dates)} ~ {max(dates)}" if dates else "-"

        return {
            "file": Path(filepath).name,
            "account": account_name,
            "period": date_range,
            "parsed": len(rows),
            "saved": saved,
        }

    def sync_dir(self, dirpath: str) -> List[dict]:
        """폴더 내 모든 xlsx 파일 동기화"""
        p = Path(dirpath)
        files = sorted(p.glob("*-dailySettlement-*.xlsx"))
        results = []
        for f in files:
            result = self.sync_file(str(f))
            results.append(result)
        return results


def main():
    parser = argparse.ArgumentParser(description="광고비 정산 Excel → DB 동기화")
    parser.add_argument("path", nargs="?", help="Excel 파일 또는 폴더 경로")
    parser.add_argument("--dir", type=str, help="폴더 경로 (내부 xlsx 전체)")
    args = parser.parse_args()

    syncer = AdSpendSync()

    if args.dir:
        results = syncer.sync_dir(args.dir)
    elif args.path:
        results = [syncer.sync_file(args.path)]
    else:
        # 프로젝트 루트에서 dailySettlement 파일 찾기
        files = sorted(ROOT.glob("*-dailySettlement-*.xlsx"))
        if files:
            results = [syncer.sync_file(str(f)) for f in files]
        else:
            print("Excel 파일을 지정해주세요.")
            return

    # 리포트
    print("\n" + "=" * 70)
    print("광고비 정산 동기화 결과")
    print("=" * 70)
    for r in results:
        if "error" in r and r["error"]:
            print(f"  {r['file']:40s} | 오류: {r['error']}")
        else:
            print(f"  {r['file']:40s} | {r['account']:10s} | {r['period']} | 파싱 {r['parsed']:4d} | 저장 {r['saved']:4d}")
    print("=" * 70)

    # DB 확인
    eng = create_engine(f"sqlite:///{ROOT / 'coupang_auto.db'}")
    with eng.connect() as conn:
        cnt = conn.execute(text("SELECT COUNT(*) FROM ad_spends")).scalar()
        print(f"\nad_spends 총 레코드: {cnt:,}건")


if __name__ == "__main__":
    main()
